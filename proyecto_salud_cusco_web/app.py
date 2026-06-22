import sys, os, json, threading, subprocess, time, platform, math, re, traceback
from pathlib import Path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from html_report_builder import build_page1_html, build_suplementacion_html, build_tx_anemia_html, _sec_header, _render_seccion_table, _REPORT_CSS, _wrap_tables
from typing import Optional
from flask import Flask, render_template, request, jsonify, session
from flask_cors import CORS


if platform.system() == 'Windows':
    CREATE_NO_WINDOW = 0x08000000
else:
    CREATE_NO_WINDOW = 0
try:
    from .config import (
        BASE_DIR, PROJECT_ROOT, SCRIPTS_INGESTA, SCRIPTS_BI,
        SCRIPTS_MANTENIMIENTO, SCRIPTS_INSTALACION, SCRIPTS_SQL_REPORTES,
        SCRIPTS_SQL_VACUNAS, SCRIPTS_PADRONES, SCRIPTS_BI_LOAD,
        BOTONES_REPORTE_PREDETERMINADOS,
        SCRIPTS_MAESTROS_EDITABLES, EDITOR_BUTTONS_FILE
    )
except ImportError:
    from config import (
        BASE_DIR, PROJECT_ROOT, SCRIPTS_INGESTA, SCRIPTS_BI,
        SCRIPTS_MANTENIMIENTO, SCRIPTS_INSTALACION, SCRIPTS_SQL_REPORTES,
        SCRIPTS_SQL_VACUNAS, SCRIPTS_PADRONES, SCRIPTS_BI_LOAD,
        BOTONES_REPORTE_PREDETERMINADOS,
        SCRIPTS_MAESTROS_EDITABLES, EDITOR_BUTTONS_FILE
    )

app = Flask(__name__,
    template_folder=str(BASE_DIR / 'templates'),
    static_folder=str(BASE_DIR / 'static'),
    static_url_path='/static')
app.secret_key = 'sistema-salud-cusco-web-2026'
CORS(app)

# Disable cache for static files (dev speed)
@app.after_request
def _no_cache(resp):
    if resp.content_type and resp.content_type.startswith('text/html'):
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    elif resp.content_type and resp.content_type.startswith('text/javascript'):
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    return resp

# ---------- Helper Utilities ----------
_scripts_python_dir = str(PROJECT_ROOT / "scripts_python")
if _scripts_python_dir not in sys.path:
    sys.path.insert(0, _scripts_python_dir)
    sys.path.insert(0, str(PROJECT_ROOT))

# ---------- Shared Execution State (Polling-based) ----------
_exec_state = {}
_exec_state_lock = threading.Lock()
_active_process = None
_active_process_lock = threading.Lock()
import uuid as _uuid

MAX_LINES = 2000

def _nuevo_token():
    return str(_uuid.uuid4())

def _init_state(token, nombre):
    with _exec_state_lock:
        _exec_state[token] = {
            'token': token,
            'nombre': nombre,
            'status': 'starting',
            'lines': [],
            'line_count': 0,
            'progress': {'done': 0, 'total': 0, 'porcentaje': 0, 'eta': ''},
            'start_time': time.time(),
            'end_time': None,
            'exit_code': None,
            'error': None,
        }

def _append_line(token, line):
    with _exec_state_lock:
        state = _exec_state.get(token)
        if not state:
            return
        state['lines'].append(line)
        if len(state['lines']) > MAX_LINES:
            state['lines'] = state['lines'][-MAX_LINES//2:]
        state['line_count'] += 1

def _set_progress(token, total=None, done=None, porcentaje=None, eta=None):
    with _exec_state_lock:
        state = _exec_state.get(token)
        if not state:
            return
        p = state['progress']
        if total is not None:
            p['total'] = total
        if done is not None:
            p['done'] = done
        if porcentaje is not None:
            p['porcentaje'] = porcentaje
        if eta is not None:
            p['eta'] = eta

def _set_status(token, status, exit_code=None, error=None):
    with _exec_state_lock:
        state = _exec_state.get(token)
        if not state:
            return
        state['status'] = status
        if exit_code is not None:
            state['exit_code'] = exit_code
        if error is not None:
            state['error'] = error
        if status in ('completed', 'error', 'cancelled'):
            state['end_time'] = time.time()

def _get_state(token):
    with _exec_state_lock:
        s = _exec_state.get(token)
        if s is None:
            return None
        return dict(s)

def _cleanup_old_states():
    now = time.time()
    with _exec_state_lock:
        to_del = [t for t, s in _exec_state.items()
                  if s.get('end_time') and now - s['end_time'] > 300]
        for t in to_del:
            del _exec_state[t]

def _resolver_ruta_script(script_rel):
    if getattr(sys, 'frozen', False):
        appdata = Path(os.environ.get('APPDATA', os.path.expanduser('~'))) / "Proyecto_Salud_Cusco" / "data" / script_rel
        if appdata.exists():
            return str(appdata)
    return str(PROJECT_ROOT / script_rel)

def _cargar_config_editor():
    path = EDITOR_BUTTONS_FILE
    if path.exists():
        with open(str(path), 'r', encoding='utf-8') as f:
            return json.load(f)
    return {"botones": BOTONES_REPORTE_PREDETERMINADOS}

def _guardar_config_editor(config):
    with open(str(EDITOR_BUTTONS_FILE), 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

# ---------- Subprocess Execution with Polling ----------
def ejecutar_script(ruta_script, args=None, mostrar_progreso=False):
    global _active_process
    nombre = os.path.basename(ruta_script).replace('.py', '')
    token = _nuevo_token()
    _init_state(token, nombre)
    _set_status(token, 'starting')

    if getattr(sys, 'frozen', False):
        comando = [sys.executable, '--run-script', ruta_script]
    else:
        comando = [sys.executable, ruta_script]
    if args:
        comando.extend(args if isinstance(args, list) else [args])

    def hilo():
        global _active_process
        _set_status(token, 'running')
        try:
            _append_line(token, f"Iniciando: {' '.join(comando)}")

            env = os.environ.copy()
            env["PYTHONUTF8"] = "1"
            env["PYTHONUNBUFFERED"] = "1"

            proc = subprocess.Popen(
                comando,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding='utf-8',
                errors='replace',
                env=env,
                creationflags=CREATE_NO_WINDOW,
            )

            with _active_process_lock:
                _active_process = proc

            total = None
            start_time = time.time()

            for line in iter(proc.stdout.readline, ''):
                if not line:
                    break

                state = _get_state(token)
                if state and state['status'] == 'cancelled':
                    proc.terminate()
                    time.sleep(0.5)
                    if proc.poll() is None:
                        proc.kill()
                    break

                line = line.rstrip('\n\r')
                _append_line(token, line)

                if mostrar_progreso and line.startswith('[PROGRESS]'):
                    payload = line[len('[PROGRESS]'):]
                    if payload.startswith('TOTAL='):
                        total = int(payload.split('=')[1])
                        _set_progress(token, total=total)
                    elif payload.startswith('DONE='):
                        parts = payload.split('|')
                        progress_data = {'token': token}
                        for p in parts:
                            if '=' in p:
                                k, v = p.split('=', 1)
                                progress_data[k.lower()] = v
                        done = int(progress_data.get('done', 0))
                        total_val = total or 100
                        pct = min(done / total_val, 1.0)
                        elapsed = time.time() - start_time
                        eta = ''
                        if done > 0:
                            eta_secs = (elapsed / done) * (total_val - done)
                            if eta_secs > 3600:
                                eta = f"{eta_secs/3600:.1f}h"
                            elif eta_secs > 60:
                                eta = f"{eta_secs/60:.0f}m {eta_secs%60:.0f}s"
                            else:
                                eta = f"{eta_secs:.0f}s"
                        _set_progress(token, done=done, porcentaje=pct, eta=eta)

            proc.wait()
            rc = proc.returncode
            with _active_process_lock:
                _active_process = None

            if rc == 0:
                _set_status(token, 'completed', exit_code=rc)
                _append_line(token, '✅ Completado correctamente')
            else:
                _set_status(token, 'error', exit_code=rc)
                _append_line(token, f'❌ Error (código {rc})')

        except Exception as e:
            with _active_process_lock:
                _active_process = None
            _set_status(token, 'error', error=str(e))
            _append_line(token, f'❌ Error: {e}')

    thread = threading.Thread(target=hilo, daemon=True)
    thread.start()
    return {'token': token, 'mensaje': 'Ejecución iniciada'}

# ---------- ROUTES ----------
@app.route('/')
def index():
    return render_template('index.html')

# ==================== DB CONFIG ====================
@app.route('/api/db/config', methods=['GET', 'POST'])
def db_config():
    try:
        from db_config import get_db_config, update_db_config
    except ImportError:
        return jsonify({'error': 'No se pudo importar db_config'}), 500

    if request.method == 'GET':
        cfg = get_db_config()
        return jsonify({
            'host': cfg.host if hasattr(cfg, 'host') else 'localhost',
            'port': cfg.port if hasattr(cfg, 'port') else 5432,
            'database': cfg.database if hasattr(cfg, 'database') else 'ivan_proceso_his',
            'schema': cfg.schema if hasattr(cfg, 'schema') else 'es_ivan',
            'user': cfg.user if hasattr(cfg, 'user') else 'postgres',
            'password': '********',
            'has_password': bool(cfg.password if hasattr(cfg, 'password') else ''),
        })

    data = request.json
    update_db_config(
        host=data.get('host'),
        port=data.get('port'),
        database=data.get('database'),
        schema=data.get('schema'),
        password=data.get('password'),
    )
    return jsonify({'mensaje': 'Configuración guardada'})

@app.route('/api/db/detect', methods=['POST'])
def db_detect():
    try:
        from db_config import detectar_postgresql_existente
        result = detectar_postgresql_existente()
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e), 'instalado': False}), 500

@app.route('/api/db/verify', methods=['POST'])
def db_verify():
    try:
        from db_config import verificar_bd_esquema, get_db_config
        cfg = get_db_config()
        log_lines = []
        exito = verificar_bd_esquema(
            config=cfg,
            log=lambda m: log_lines.append(m),
            guardar_password=True
        )
        return jsonify({'exito': bool(exito), 'mensaje': 'Verificación completada', 'log': log_lines})
    except Exception as e:
        return jsonify({'exito': False, 'mensaje': str(e), 'log': []}), 500

@app.route('/api/db/install', methods=['POST'])
def db_install():
    ruta = str(SCRIPTS_INSTALACION / 'instalar_postgresql.py')
    result = ejecutar_script(ruta, mostrar_progreso=True)
    return jsonify(result)

@app.route('/api/db/start-service', methods=['POST'])
def db_start_service():
    try:
        import platform
        if platform.system() == "Linux":
            for cmd in [
                ["sudo", "-n", "systemctl", "start", "postgresql"],
                ["systemctl", "start", "postgresql"],
                ["pkexec", "systemctl", "start", "postgresql"],
            ]:
                try:
                    result = subprocess.run(
                        cmd, capture_output=True, text=True, timeout=30,
                    )
                    if result.returncode == 0:
                        time.sleep(2)
                        return jsonify({'exito': True, 'mensaje': "PostgreSQL iniciado correctamente"})
                except Exception:
                    continue
            return jsonify({
                'exito': False,
                'mensaje': "No se pudo iniciar PostgreSQL automáticamente. Ejecute manualmente: sudo systemctl start postgresql"
            })
        sys.path.insert(0, str(SCRIPTS_INSTALACION))
        from instalar_postgresql import iniciar_servicio_postgresql, esperar_servicio_activo
        from db_config import detectar_postgresql_existente
        info = detectar_postgresql_existente()
        exito = iniciar_servicio_postgresql(info.get("servicio_nombre"))
        if exito:
            esperar_servicio_activo(info.get("servicio_nombre"), timeout=30)
        return jsonify({'exito': exito, 'mensaje': "Servicio iniciado" if exito else "No se pudo iniciar"})
    except Exception as e:
        return jsonify({'exito': False, 'mensaje': str(e)}), 500

@app.route('/api/db/init', methods=['POST'])
def db_init():
    try:
        sys.path.insert(0, str(SCRIPTS_INSTALACION))
        from db_config import get_db_config, verificar_bd_esquema
        from instalar_postgresql import crear_base_datos_y_esquema
        cfg = get_db_config()
        log_lines = []
        exito = crear_base_datos_y_esquema(
            cfg.host, cfg.port, cfg.user, cfg.password,
            cfg.database, cfg.schema,
            log=lambda m: log_lines.append(m)
        )
        return jsonify({'exito': exito, 'log': log_lines})
    except Exception as e:
        return jsonify({'exito': False, 'mensaje': str(e)}), 500

@app.route('/api/db/recover-password', methods=['POST'])
def db_recover_password():
    try:
        from db_config import inicializar_base_datos, get_db_config
        cfg = get_db_config()
        log_lines = []
        exito = inicializar_base_datos(log=lambda m: log_lines.append(m))
        return jsonify({'exito': exito, 'log': log_lines})
    except Exception as e:
        return jsonify({'exito': False, 'mensaje': str(e)}), 500

# ==================== EJECUCION STATUS / CANCEL (Polling) ====================
@app.route('/api/ejecucion/status/<string:token>')
def ejecucion_status(token):
    state = _get_state(token)
    if state is None:
        return jsonify({'error': 'Ejecución no encontrada', 'status': 'unknown'}), 404
    return jsonify(state)

@app.route('/api/ejecucion/cancel', methods=['POST'])
def ejecucion_cancel():
    global _active_process
    data = request.json
    token = data.get('token', 0)

    if token:
        state = _get_state(token)
        if state and state['status'] in ('running', 'starting'):
            _set_status(token, 'cancelled')
            _append_line(token, '🛑 Proceso cancelado por el usuario')

    with _active_process_lock:
        if _active_process:
            try:
                _active_process.terminate()
                time.sleep(0.5)
                if _active_process.poll() is None:
                    _active_process.kill()
            except Exception:
                pass
            _active_process = None

    return jsonify({'mensaje': 'Proceso cancelado'})

# ==================== INGESTA ====================
def _meses_con_datos(anio):
    """Retorna lista de meses (int) que ya tienen datos en BD para el año."""
    try:
        from db_config import get_db_config
        import psycopg2
        cfg = get_db_config()
        conn = psycopg2.connect(
            dbname=cfg.database, user=cfg.user, password=cfg.password,
            host=cfg.host, port=cfg.port, connect_timeout=5
        )
        cur = conn.cursor()
        cur.execute(f"SELECT DISTINCT mes FROM {cfg.schema}.hisminsa24 WHERE anio = %s AND mes ~ '^[0-9]+$'", (anio,))
        meses = sorted(set(int(row[0]) for row in cur.fetchall() if row[0].isdigit()))
        cur.close(); conn.close()
        return meses
    except Exception:
        return []

def _borrar_anio_bd(anio):
    """Borra todos los registros de un año antes de importar."""
    from db_config import get_db_config
    import psycopg2
    cfg = get_db_config()
    conn = psycopg2.connect(
        dbname=cfg.database, user=cfg.user, password=cfg.password,
        host=cfg.host, port=cfg.port, connect_timeout=5
    )
    cur = conn.cursor()
    cur.execute(f"DELETE FROM {cfg.schema}.hisminsa24 WHERE anio = %s", (anio,))
    borrados = cur.rowcount
    conn.commit()
    cur.close(); conn.close()
    return borrados

@app.route('/api/ingesta/check', methods=['POST'])
def ingesta_check():
    """Verifica qué meses del año ya tienen datos en BD."""
    data = request.json
    anio = data.get('anio', '2024')
    meses_bd = _meses_con_datos(anio)
    return jsonify({'anio': anio, 'meses_con_datos': meses_bd, 'tiene_datos': len(meses_bd) > 0})

@app.route('/api/ingesta/import', methods=['POST'])
def ingesta_import():
    data = request.json
    anio = data.get('anio', '2024')
    meses = data.get('meses', [])
    ruta_crudos = data.get('ruta_crudos', '')
    modo = data.get('modo', 'reemplazar')  # 'reemplazar' o 'completar'

    # Si modo es completar, filtrar meses que ya tienen datos
    if modo == 'completar':
        meses_bd = _meses_con_datos(anio)
        meses = [m for m in meses if m not in meses_bd]
        if not meses:
            return jsonify({'error': 'Todos los meses ya tienen datos', 'skip': True}), 200

    # Si modo es reemplazar, borrar datos del año antes de importar
    if modo == 'reemplazar':
        try:
            borrados = _borrar_anio_bd(anio)
        except Exception:
            pass  # si falla el borrado, el script internamente limpia mes x mes

    # Igual que el desktop: si hay ruta personalizada,
    # primero prueba ruta/anio, si no existe usa ruta tal cual
    if ruta_crudos:
        carpeta_con_anio = os.path.join(ruta_crudos, anio)
        ruta_final = carpeta_con_anio if os.path.isdir(carpeta_con_anio) else ruta_crudos
    else:
        ruta_final = ''

    if not meses or len(meses) == 12:
        script = str(SCRIPTS_INGESTA / '01cargacvs_universal.py')
        args = [anio]
        if ruta_final:
            args.append(ruta_final)
    else:
        script = str(SCRIPTS_INGESTA / '01cargacvs_mensual.py')
        args = [anio] + [str(m) for m in meses]
        if ruta_final:
            args.append(ruta_final)

    return jsonify(ejecutar_script(script, args=args, mostrar_progreso=True))

@app.route('/api/ingesta/refresh', methods=['POST'])
def ingesta_refresh():
    script = str(SCRIPTS_INGESTA / 'actualizar_his_proceso_maestros.py')
    return jsonify(ejecutar_script(script, mostrar_progreso=True))

@app.route('/api/ingesta/delete', methods=['POST'])
def ingesta_delete():
    data = request.json
    modo = data.get('modo', 'todo')
    anio = data.get('anio', '')
    mes = data.get('mes', '')
    script = str(SCRIPTS_MANTENIMIENTO / '02_eliminar_datos.py')
    args = [f'modo={modo}']
    if modo == 'anio' and anio:
        args.append(f'anio={anio}')
    if modo == 'mes' and anio and mes:
        args.append(f'anio={anio}')
        args.append(f'mes={mes}')
    return jsonify(ejecutar_script(script, args=args, mostrar_progreso=True))

# ==================== REPORTES ====================
@app.route('/api/reportes/config', methods=['GET'])
def reportes_config():
    config = _cargar_config_editor()
    return jsonify(config)

@app.route('/api/reportes/run', methods=['POST'])
def reportes_run():
    data = request.json
    ruta_script = data.get('script', '')
    anio = data.get('anio', '2024')

    ruta_abs = _resolver_ruta_script(ruta_script)
    ext = Path(ruta_script).suffix.lower()

    if ext == '.sql':
        from db_config import get_db_config
        cfg = get_db_config()

        with open(ruta_abs, 'r', encoding='utf-8', errors='replace') as f:
            sql_content = f.read()

        sql_upper = sql_content.strip().upper()
        if any(sql_upper.startswith(kw) for kw in ('CREATE', 'INSERT', 'DROP', 'UPDATE', 'DELETE', 'ALTER', 'TRUNCATE')):
            executor = str(SCRIPTS_BI / '04_ejecutor_procedures.py')
        else:
            executor = str(SCRIPTS_BI / '04_generador_reportes.py')

        result = ejecutar_script(executor, args=[ruta_abs, anio])
    else:
        result = ejecutar_script(ruta_abs, args=[anio])

    return jsonify(result)

@app.route('/api/reportes/save-config', methods=['POST'])
def reportes_save_config():
    data = request.json
    _guardar_config_editor(data)
    return jsonify({'mensaje': 'Configuración guardada'})

@app.route('/api/reportes/new', methods=['POST'])
def reportes_new():
    data = request.json
    nombre = data.get('nombre', '')
    seccion = data.get('seccion', 'reportes')
    color = data.get('color', '#3498DB')

    safe_name = nombre.lower().replace(' ', '_').replace('\u20e3', '').strip()
    sql_path = SCRIPTS_SQL_REPORTES / f"{safe_name}.sql"
    with open(str(sql_path), 'w', encoding='utf-8') as f:
        f.write(f"-- {nombre}\n-- GERESA Cusco - Sistema de Monitoreo de Salud\n\nSELECT 1;\n")

    config = _cargar_config_editor()
    config['botones'].append({
        'nombre': nombre,
        'script': f"scripts_sql/reportes/{safe_name}.sql",
        'seccion': seccion,
        'color_bg': color,
        'custom': True,
    })
    _guardar_config_editor(config)
    return jsonify({'mensaje': f'Reporte "{nombre}" creado', 'ruta': str(sql_path)})

@app.route('/api/reportes/delete', methods=['POST'])
def reportes_delete():
    data = request.json
    idx = data.get('index', -1)
    config = _cargar_config_editor()
    if 0 <= idx < len(config['botones']):
        boton = config['botones'].pop(idx)
        if boton.get('custom'):
            ruta = str(PROJECT_ROOT / boton['script'])
            if os.path.exists(ruta):
                os.remove(ruta)
        _guardar_config_editor(config)
        return jsonify({'mensaje': f'Reporte "{boton["nombre"]}" eliminado'})
    return jsonify({'error': 'Índice inválido'}), 400

# ==================== MAESTROS ====================
@app.route('/api/maestros/tablas', methods=['GET'])
def maestros_tablas():
    try:
        import psycopg2
        from db_config import get_db_config
        cfg = get_db_config()
        conn = psycopg2.connect(
            host=cfg.host, port=cfg.port, user=cfg.user,
            password=cfg.password, dbname=cfg.database
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT table_name, table_schema
            FROM information_schema.tables
            WHERE table_schema = %s
            ORDER BY table_name
        """, (cfg.schema,))
        tablas = [{'nombre': r[0], 'esquema': r[1]} for r in cur.fetchall()]
        cur.close()
        conn.close()
        return jsonify(tablas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/maestros/ejecutar', methods=['POST'])
def maestros_ejecutar():
    data = request.json
    script = data.get('script', '')
    args = data.get('args', [])
    ruta = _resolver_ruta_script(script)
    return jsonify(ejecutar_script(ruta, args=args, mostrar_progreso=True))

@app.route('/api/maestros/csv-list', methods=['POST'])
def maestros_csv_list():
    data = request.json
    folder = data.get('folder', '')
    if not folder or not os.path.isdir(folder):
        return jsonify({'error': 'Carpeta inválida'}), 400
    try:
        csvs = sorted([f for f in os.listdir(folder) if f.lower().endswith('.csv')])
        return jsonify({'csvs': csvs, 'folder': folder})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/maestros/eliminar', methods=['POST'])
def maestros_eliminar():
    data = request.json
    tablas = data.get('tablas', [])
    if not tablas:
        return jsonify({'error': 'No se especificaron tablas'}), 400
    try:
        import psycopg2
        from db_config import get_db_config
        cfg = get_db_config()
        conn = psycopg2.connect(host=cfg.host, port=cfg.port, user=cfg.user, password=cfg.password, dbname=cfg.database)
        cur = conn.cursor()
        resultados = []
        for t in tablas:
            try:
                cur.execute(f'DROP TABLE IF EXISTS "{cfg.schema}"."{t}" CASCADE;')
                conn.commit()
                resultados.append({'tabla': t, 'ok': True})
            except Exception as e:
                resultados.append({'tabla': t, 'ok': False, 'error': str(e)})
        cur.close()
        conn.close()
        return jsonify({'resultados': resultados})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/maestros/eliminar-todos', methods=['POST'])
def maestros_eliminar_todos():
    try:
        import psycopg2
        from db_config import get_db_config
        cfg = get_db_config()
        conn = psycopg2.connect(host=cfg.host, port=cfg.port, user=cfg.user, password=cfg.password, dbname=cfg.database)
        cur = conn.cursor()
        cur.execute(f"""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = '{cfg.schema}'
            AND (table_name LIKE 'maestro%%' OR table_name = 'eess2025' OR table_name LIKE '%%susalud%%')
        """)
        tablas = [r[0] for r in cur.fetchall()]
        for t in tablas:
            cur.execute(f'DROP TABLE IF EXISTS "{cfg.schema}"."{t}" CASCADE;')
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({'ok': True, 'eliminadas': len(tablas)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/maestros/descriptions', methods=['GET'])
def maestros_descriptions():
    descs = {
        "maestro_paciente": "Paciente (DNI, nombre, fecha nac., género, etnia)",
        "maestro_personal": "Personal de salud (colegio profesional)",
        "eess2025": "Establecimiento (red, microred, provincia, distrito)",
        "maestro_his_establecimiento": "Establecimientos crudos HIS (fuente para reconstruir eess2025)",
        "maestro_his_cie_cpms": "Diagnósticos CIE / Procedimientos CPT",
        "maestro_his_etnia": "Etnias (descripción)",
        "maestro_his_ups": "Unidades Productoras de Servicios (UPS)",
        "maestro_his_colegio": "Colegios profesionales",
        "maestro_his_actividad": "Actividades HIS",
        "maestro_his_centro_poblado": "Centros poblados",
        "maestro_his_condicion_contrato": "Condición de contrato del personal",
        "maestro_his_dosis": "Dosis de vacunas",
        "maestro_his_financiador": "Financiadores (SIS, ESSALUD, etc.)",
        "maestro_his_gruporiesgo_lab": "Grupos de riesgo (lab)",
        "maestro_his_institucion_edu": "Instituciones educativas",
        "maestro_his_lab": "Laboratorio (parámetros)",
        "maestro_his_otra_condicion": "Otra condición clínica",
        "maestro_his_pais": "Países (código y nombre)",
        "maestro_his_profesion": "Profesiones del personal",
        "maestro_his_sistema": "Sistemas de salud",
        "maestro_his_tipo_doc": "Tipos de documento de identidad",
        "maestro_his_ubigeo": "Ubigeos INEI / RENIEC",
        "maestro_his_susalud": "SUSALUD (establecimientos supervisados)",
        "maestro_eess_susalud": "SUSALUD crudo (fuente para reconstruir eess2025)",
    }
    return jsonify(descs)

# ==================== PADRON NOMINAL ====================
def _db_cursor():
    from db_config import get_db_config
    import psycopg2
    cfg = get_db_config()
    conn = psycopg2.connect(host=cfg.host, port=cfg.port, user=cfg.user, password=cfg.password, dbname=cfg.database)
    return conn

def _tabla_existe(nombre_tabla):
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        from db_config import get_db_config
        cfg = get_db_config()
        cur.execute("SELECT EXISTS(SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s)", (cfg.schema, nombre_tabla))
        res = cur.fetchone()[0]
        cur.close(); conn.close()
        return res
    except:
        return False

@app.route('/api/padron/status', methods=['GET'])
def padron_status():
    existe = _tabla_existe('padron_nominal')
    total = 0
    if existe:
        try:
            conn = _db_cursor()
            cur = conn.cursor()
            from db_config import get_db_config
            cfg = get_db_config()
            cur.execute(f"SELECT COUNT(*) FROM {cfg.schema}.padron_nominal")
            total = cur.fetchone()[0]
            cur.close(); conn.close()
        except:
            pass
    return jsonify({'existe': existe, 'total': total})

@app.route('/api/padron/cargar', methods=['POST'])
def padron_cargar():
    data = request.json or {}
    ruta_carpeta = data.get('ruta_carpeta', '')
    script = str(SCRIPTS_BI / 'cargar_padron_nominal.py')
    return jsonify(ejecutar_script(script, args=[ruta_carpeta] if ruta_carpeta else None, mostrar_progreso=True))

@app.route('/api/padron/consulta', methods=['POST'])
def padron_consulta():
    data = request.json
    pagina = data.get('pagina', 1)
    por_pagina = data.get('por_pagina', 50)
    busqueda = data.get('busqueda', '')
    offset = (pagina - 1) * por_pagina

    try:
        conn = _db_cursor()
        cur = conn.cursor()
        from db_config import get_db_config
        cfg = get_db_config()
        tabla = f"{cfg.schema}.padron_nominal"

        if busqueda:
            filtro = f"WHERE (nombre_nino ILIKE %s OR apellido_pat_nino ILIKE %s OR apellido_mat_nino ILIKE %s OR doc_madre ILIKE %s)"
            param = f'%{busqueda}%'
            params = [param, param, param, param]
            cur.execute(f"SELECT COUNT(*) FROM {tabla} {filtro}", params)
            total = cur.fetchone()[0]
            cur.execute(f"SELECT * FROM {tabla} {filtro} ORDER BY anio DESC, nro LIMIT %s OFFSET %s", params + [por_pagina, offset])
        else:
            cur.execute(f"SELECT COUNT(*) FROM {tabla}")
            total = cur.fetchone()[0]
            cur.execute(f"SELECT * FROM {tabla} ORDER BY anio DESC, nro LIMIT %s OFFSET %s", [por_pagina, offset])

        cols = [desc[0] for desc in cur.description]
        filas = [dict(zip(cols, row)) for row in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({'filas': filas, 'total': total, 'pagina': pagina, 'por_pagina': por_pagina})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/padron/geojson', methods=['GET'])
def padron_geojson():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        from db_config import get_db_config
        cfg = get_db_config()
        cur.execute(f"""
            SELECT latitud, longitud, nombre_nino, apellido_pat_nino,
                   apellido_mat_nino, direccion, dist_nino, ubigeo_pn
            FROM {cfg.schema}.padron_nominal
            WHERE latitud IS NOT NULL AND longitud IS NOT NULL
            AND latitud BETWEEN -14 AND -12
            AND longitud BETWEEN -73 AND -70
            LIMIT 5000
        """)
        features = []
        for row in cur.fetchall():
            lat, lng, nombre, apat, amat, dir, dist, ubigeo = row
            if lat and lng:
                features.append({
                    'type': 'Feature',
                    'geometry': {'type': 'Point', 'coordinates': [lng, lat]},
                    'properties': {
                        'nombre': f"{nombre or ''} {apat or ''} {amat or ''}",
                        'direccion': dir or '',
                        'distrito': dist or '',
                        'ubigeo': ubigeo or '',
                    }
                })
        cur.close(); conn.close()
        return jsonify({'type': 'FeatureCollection', 'features': features})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== CNV ====================
@app.route('/api/cnv/status', methods=['GET'])
def cnv_status():
    existe = _tabla_existe('cnv_cusco')
    total = 0
    if existe:
        try:
            conn = _db_cursor()
            cur = conn.cursor()
            from db_config import get_db_config
            cfg = get_db_config()
            cur.execute(f"SELECT COUNT(*) FROM {cfg.schema}.cnv_cusco")
            total = cur.fetchone()[0]
            cur.close(); conn.close()
        except:
            pass
    return jsonify({'existe': existe, 'total': total})

@app.route('/api/cnv/cargar', methods=['POST'])
def cnv_cargar():
    data = request.json or {}
    ruta_carpeta = data.get('ruta_carpeta', '')
    script = str(SCRIPTS_BI / 'cargar_cnv.py')
    return jsonify(ejecutar_script(script, args=[ruta_carpeta] if ruta_carpeta else None, mostrar_progreso=True))

@app.route('/api/cnv/consulta', methods=['POST'])
def cnv_consulta():
    data = request.json
    pagina = data.get('pagina', 1)
    por_pagina = data.get('por_pagina', 50)
    busqueda = data.get('busqueda', '')
    offset = (pagina - 1) * por_pagina

    try:
        conn = _db_cursor()
        cur = conn.cursor()
        from db_config import get_db_config
        cfg = get_db_config()
        tabla = f"{cfg.schema}.cnv_cusco"

        if busqueda:
            filtro = "WHERE (pri_ape_madre ILIKE %s OR seg_ape_madre ILIKE %s OR prenom_madre ILIKE %s OR nu_doc_madre ILIKE %s OR nu_cnv ILIKE %s)"
            param = f'%{busqueda}%'
            params = [param]*5
            cur.execute(f"SELECT COUNT(*) FROM {tabla} {filtro}", params)
            total = cur.fetchone()[0]
            cur.execute(f"SELECT * FROM {tabla} {filtro} ORDER BY fe_crea DESC LIMIT %s OFFSET %s", params + [por_pagina, offset])
        else:
            cur.execute(f"SELECT COUNT(*) FROM {tabla}")
            total = cur.fetchone()[0]
            cur.execute(f"SELECT * FROM {tabla} ORDER BY fe_crea DESC LIMIT %s OFFSET %s", [por_pagina, offset])

        cols = [desc[0] for desc in cur.description]
        filas = [dict(zip(cols, row)) for row in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({'filas': filas, 'total': total, 'pagina': pagina, 'por_pagina': por_pagina})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== SELECTOR DE CARPETA NATIVO ====================
@app.route('/api/seleccionar-carpeta', methods=['POST'])
def seleccionar_carpeta():
    script = str(SCRIPTS_BI / 'selector_carpeta.py')
    try:
        if getattr(sys, 'frozen', False):
            cmd = [sys.executable, '--run-script', script]
        else:
            cmd = [sys.executable, script]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        if result.returncode == 0 and result.stdout.strip():
            return jsonify({'ruta': result.stdout.strip()})
        return jsonify({'ruta': ''})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== LISTAR CARPETA (WEB) ====================
@app.route('/api/listar-carpeta', methods=['POST'])
def listar_carpeta():
    data = request.json or {}
    ruta = data.get('ruta', '')
    if not ruta:
        unidades = []
        for letra in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ':
            p = f'{letra}:\\'
            if os.path.exists(p):
                unidades.append(p)
        return jsonify({'carpetas': unidades, 'archivos': [], 'actual': ''})
    try:
        ruta_p = Path(ruta)
        if not ruta_p.exists() or not ruta_p.is_dir():
            return jsonify({'error': 'Ruta no válida'}), 400
        carpetas = sorted([str(p) for p in ruta_p.iterdir() if p.is_dir()])
        archivos = sorted([
            str(p) for p in ruta_p.iterdir()
            if p.is_file() and p.suffix.lower() in ('.csv', '.xlsx', '.xls')
        ])
        return jsonify({'carpetas': carpetas, 'archivos': archivos, 'actual': str(ruta_p)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== MAPA / GEOLOCALIZACION ====================
@app.route('/api/mapa/kde', methods=['POST'])
def mapa_kde():
    """Genera KDE gaussiano sin SciPy y hotspots por percentil 90."""
    try:
        data = request.json or {}
        anio = str(data.get('anio', '2024'))
        filtro_tipo = str(data.get('filtro', 'pn')).lower()  # 'pn', 'iras', 'edas'

        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()

        if filtro_tipo == 'pn':
            source_table = 'padron_nominal'
            data_year = None
            fallback = False
            cur.execute(f"""
                SELECT latitud::text as lat_txt,
                       longitud::text as lng_txt,
                       COUNT(*)::double precision as weight
                FROM {_qtable(esquema, 'padron_nominal')}
                WHERE latitud IS NOT NULL AND longitud IS NOT NULL
                  AND TRIM(latitud::text) != '' AND TRIM(longitud::text) != ''
                GROUP BY 1, 2
            """)
        elif filtro_tipo in ('iras', 'edas'):
            source_table = 'tabla_iras_edas'
            if not _table_exists(cur, esquema, source_table):
                return jsonify({'error': 'No existe tabla_iras_edas'}), 400
            data_year = anio
            fallback = False
            if not _table_has_year(cur, esquema, source_table, data_year) and _table_has_year(cur, esquema, source_table, '2024'):
                data_year = '2024'
                fallback = anio != data_year

            if filtro_tipo == 'iras':
                code_filter = "AND UPPER(t.codigo_item::text) LIKE 'J%%'"
            else:
                code_filter = "AND (UPPER(t.codigo_item::text) LIKE 'A0%%' OR UPPER(t.codigo_item::text) LIKE 'K52%%')"

            cur.execute(f"""
                SELECT pn.latitud::text as lat_txt,
                       pn.longitud::text as lng_txt,
                       COUNT(*)::double precision as weight
                FROM {_qtable(esquema, source_table)} t
                JOIN {_qtable(esquema, 'padron_nominal')} pn
                  ON pn.cnv_dni::text = t.dni_paciente::text
                WHERE t.anio::text = %s
                  {code_filter}
                  AND pn.latitud IS NOT NULL AND pn.longitud IS NOT NULL
                  AND TRIM(pn.latitud::text) != '' AND TRIM(pn.longitud::text) != ''
                GROUP BY 1, 2
            """, (data_year,))
        else:
            return jsonify({'error': 'Filtro no válido. Use pn, iras o edas.'}), 400

        coords = _normalized_coords_from_rows(cur.fetchall())
        cur.close(); conn.close()

        if len(coords) < 3:
            return jsonify({'error': f'Muy pocas coordenadas ({len(coords)})', 'coords_count': len(coords)}), 400
        warning = None
        if len(coords) < 5:
            warning = f'KDE calculado con pocas coordenadas ({len(coords)} puntos fuente); revisar completitud geográfica.'

        result = _compute_python_kde(coords, grid_size=100)
        result.update({
            'coords_count': int(sum(c[2] for c in coords)),
            'source_points': len(coords),
            'source_table': source_table,
            'filtro': filtro_tipo,
            'anio_solicitado': anio,
            'anio_datos': data_year,
            'fallback': fallback,
            'warning': warning,
        })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mapa/establecimientos', methods=['GET'])
def mapa_establecimientos():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        # Use the eess2025 table which has the most complete data
        cur.execute(f"""
            SELECT cod_eess, nombre_eess, cat, red, microred, distrito, provincia,
                   ubigueo_eess
            FROM {esquema}.eess2025
            WHERE ubigueo_eess IS NOT NULL AND ubigueo_eess != ''
            ORDER BY nombre_eess
        """)
        rows = [{'codigo': r[0], 'nombre': r[1], 'categoria': r[2],
                 'red': r[3], 'microred': r[4], 'distrito': r[5],
                 'provincia': r[6], 'ubigueo': r[7]} for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(rows)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/mapa/pacientes-por-distrito', methods=['GET'])
def mapa_pacientes_por_distrito():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        from db_config import get_db_config
        cfg = get_db_config()
        cur.execute(f"""
            SELECT dist_nino, COUNT(*) as total,
                   COUNT(*) FILTER (WHERE latitud IS NOT NULL) as con_coord
            FROM {cfg.schema}.padron_nominal
            GROUP BY dist_nino
            ORDER BY total DESC
        """)
        rows = [{'distrito': r[0], 'total': r[1], 'con_coord': r[2]} for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify(rows)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== DASHBOARDS ====================
def _get_esquema():
    from db_config import get_db_config
    return get_db_config().schema

def _ejecutar(conn, sql, params=None):
    cur = conn.cursor()
    cur.execute(sql, params or ())
    return cur

_IDENT_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')

def _qident(value):
    value = str(value)
    if not _IDENT_RE.match(value):
        raise ValueError(f'Identificador SQL no válido: {value}')
    return f'"{value}"'

def _qcol(column):
    return '"' + str(column).replace('"', '""') + '"'

def _qtable(schema, table):
    return f'{_qident(schema)}.{_qident(table)}'

def _table_exists(cur, schema, table):
    cur.execute("""
        SELECT EXISTS (
            SELECT 1 FROM information_schema.tables
            WHERE table_schema=%s AND table_name=%s
        )
    """, (schema, table))
    return bool(cur.fetchone()[0])

def _table_has_year(cur, schema, table, anio):
    cur.execute(f"SELECT 1 FROM {_qtable(schema, table)} WHERE anio::text=%s LIMIT 1", (str(anio),))
    return cur.fetchone() is not None

def _resolve_year_table(cur, schema, template, requested_year, fallback_year='2024'):
    requested_year = str(requested_year or fallback_year)
    requested_table = template.format(anio=requested_year)
    if _table_exists(cur, schema, requested_table) and _table_has_year(cur, schema, requested_table, requested_year):
        return requested_table, requested_year, False

    fallback_table = template.format(anio=fallback_year)
    if _table_exists(cur, schema, fallback_table) and _table_has_year(cur, schema, fallback_table, fallback_year):
        return fallback_table, fallback_year, requested_year != fallback_year

    if _table_exists(cur, schema, requested_table):
        return requested_table, requested_year, False
    raise ValueError(f'No se encontró tabla para {template.format(anio="AAAA")}')

def _columns_for_table(cur, schema, table, min_ordinal=1):
    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema=%s AND table_name=%s AND ordinal_position >= %s
        ORDER BY ordinal_position
    """, (schema, table, min_ordinal))
    return [r[0] for r in cur.fetchall()]

def _sum_expr(cols):
    cols = [c for c in cols if c]
    return '+'.join(f'COALESCE(SUM({_qcol(c)}),0)' for c in cols) if cols else '0'

def _pct_value(numerador, denominador):
    numerador = float(numerador or 0)
    denominador = float(denominador or 0)
    if denominador <= 0:
        return None
    return round(numerador * 100.0 / denominador, 1)

def _semaforo(pct):
    if pct is None:
        return 'sin_dato'
    if pct >= 95:
        return 'verde'
    if pct >= 80:
        return 'amarillo'
    return 'rojo'

def _coverage_item(nombre, numerador, denominador, denominador_nombre, meta=95, metodo='estimado'):
    numerador = int(numerador or 0)
    denominador = int(denominador or 0)
    pct = _pct_value(numerador, denominador)
    return {
        'nombre': nombre,
        'numerador': numerador,
        'denominador': denominador,
        'denominador_nombre': denominador_nombre,
        'porcentaje': pct,
        'semaforo': _semaforo(pct),
        'meta': meta,
        'metodo': metodo,
    }

def _parse_coord(value):
    if value is None:
        return None
    s = str(value).strip().replace(',', '.')
    if not s:
        return None
    try:
        f = float(s)
        if -180 <= f <= 180:
            return f
    except ValueError:
        pass

    sign = -1 if s.startswith('-') else 1
    digits = ''.join(ch for ch in s if ch.isdigit())
    if len(digits) < 4:
        return None
    # Coordenadas del proyecto: lat/lng de Perú usan 2 dígitos enteros (-13, -72).
    try:
        f = sign * float(digits[:2] + '.' + digits[2:])
        if -180 <= f <= 180:
            return f
    except ValueError:
        return None
    return None

def _normalized_coords_from_rows(rows):
    acc = {}
    for raw_lat, raw_lng, weight in rows:
        lat = _parse_coord(raw_lat)
        lng = _parse_coord(raw_lng)
        if lat is None or lng is None:
            continue
        if not (-14 <= lat <= -12 and -73 <= lng <= -70):
            continue
        key = (round(lat, 3), round(lng, 3))
        acc[key] = acc.get(key, 0.0) + float(weight or 0)
    return [(lat, lng, weight) for (lat, lng), weight in sorted(acc.items(), key=lambda item: item[1], reverse=True)]

def _percentile(values, pct):
    values = sorted(float(v) for v in values if v is not None)
    if not values:
        return 0.0
    k = (len(values) - 1) * pct / 100.0
    lo = math.floor(k)
    hi = math.ceil(k)
    if lo == hi:
        return values[int(k)]
    return values[lo] * (hi - k) + values[hi] * (k - lo)

def _compute_python_kde(coords, grid_size=100):
    """KDE gaussiano con regla de Scott; usa NumPy si existe y no requiere SciPy."""
    try:
        import numpy as np

        arr = np.asarray(coords, dtype=float)
        arr = arr[np.isfinite(arr).all(axis=1)]
        arr = arr[arr[:, 2] > 0]
        if arr.shape[0] > 5000:
            arr = arr[np.argsort(arr[:, 2])[-5000:]]
        lats = arr[:, 0]
        lngs = arr[:, 1]
        weights = arr[:, 2]
        weight_sum = float(weights.sum())
        if arr.shape[0] < 3 or weight_sum <= 0:
            raise ValueError('Muy pocas coordenadas válidas')

        lat_min, lat_max = float(lats.min()), float(lats.max())
        lng_min, lng_max = float(lngs.min()), float(lngs.max())
        lat_span = max(lat_max - lat_min, 0.01)
        lng_span = max(lng_max - lng_min, 0.01)
        lat_margin = max(lat_span * 0.05, 0.005)
        lng_margin = max(lng_span * 0.05, 0.005)

        lat_mean = float(np.average(lats, weights=weights))
        lng_mean = float(np.average(lngs, weights=weights))
        lat_std = float(np.sqrt(np.average((lats - lat_mean) ** 2, weights=weights)))
        lng_std = float(np.sqrt(np.average((lngs - lng_mean) ** 2, weights=weights)))
        n_eff = max((weight_sum ** 2) / float((weights ** 2).sum()), 3.0)
        scott = n_eff ** (-1.0 / 6.0)
        bw_lat = max(lat_std * scott, lat_span / 60.0, 0.01)
        bw_lng = max(lng_std * scott, lng_span / 60.0, 0.01)

        grid_lats = np.linspace(lat_min - lat_margin, lat_max + lat_margin, grid_size)
        grid_lngs = np.linspace(lng_min - lng_margin, lng_max + lng_margin, grid_size)
        gx, gy = np.meshgrid(grid_lats, grid_lngs, indexing='ij')
        flat_lat = gx.ravel()
        flat_lng = gy.ravel()
        z = np.zeros(flat_lat.shape, dtype=float)

        chunk = 300
        for start in range(0, arr.shape[0], chunk):
            end = start + chunk
            dl = (flat_lat[:, None] - lats[start:end][None, :]) / bw_lat
            dg = (flat_lng[:, None] - lngs[start:end][None, :]) / bw_lng
            z += np.exp(-0.5 * (dl * dl + dg * dg)).dot(weights[start:end])

        z = z / (2.0 * math.pi * bw_lat * bw_lng * weight_sum)
        max_z = float(z.max()) if z.size else 0.0
        norm = z / max_z if max_z > 0 else z
        positives = norm[norm > 0]
        threshold = float(np.percentile(positives, 90)) if positives.size else 0.0

        grid_points = []
        hotspot_candidates = []
        for idx, val in enumerate(norm):
            val = float(val)
            if val >= 0.02:
                point = [round(float(flat_lat[idx]), 6), round(float(flat_lng[idx]), 6), round(val, 4)]
                grid_points.append(point)
                if val >= threshold and threshold > 0:
                    hotspot_candidates.append(point)

        hotspots = sorted(hotspot_candidates, key=lambda p: p[2], reverse=True)[:25]
        return {
            'type': 'python_kde',
            'grid_size': f'{grid_size}x{grid_size}',
            'bandwidth': {'lat': round(bw_lat, 5), 'lng': round(bw_lng, 5), 'method': 'scott'},
            'threshold_p90': round(threshold, 4),
            'coordenadas': grid_points,
            'hotspots': [
                {'rank': i + 1, 'latitud': h[0], 'longitud': h[1], 'densidad': h[2]}
                for i, h in enumerate(hotspots)
            ],
        }
    except ImportError:
        pass

    coords = sorted([(float(a), float(b), float(c)) for a, b, c in coords if c and c > 0], key=lambda x: x[2], reverse=True)[:800]
    if len(coords) < 3:
        raise ValueError('Muy pocas coordenadas válidas')
    grid_size = min(grid_size, 60)
    lats = [c[0] for c in coords]
    lngs = [c[1] for c in coords]
    weights = [c[2] for c in coords]
    weight_sum = sum(weights)
    lat_min, lat_max = min(lats), max(lats)
    lng_min, lng_max = min(lngs), max(lngs)
    lat_span = max(lat_max - lat_min, 0.01)
    lng_span = max(lng_max - lng_min, 0.01)
    bw_lat = max(lat_span / 8.0, 0.01)
    bw_lng = max(lng_span / 8.0, 0.01)
    grid_points = []
    densities = []
    raw_points = []
    for i in range(grid_size):
        lat = (lat_min - lat_span * 0.05) + (lat_span * 1.1) * i / (grid_size - 1)
        for j in range(grid_size):
            lng = (lng_min - lng_span * 0.05) + (lng_span * 1.1) * j / (grid_size - 1)
            density = 0.0
            for p_lat, p_lng, w in coords:
                dl = (lat - p_lat) / bw_lat
                dg = (lng - p_lng) / bw_lng
                density += math.exp(-0.5 * (dl * dl + dg * dg)) * w
            density = density / max(weight_sum, 1.0)
            raw_points.append((lat, lng, density))
            densities.append(density)
    max_density = max(densities) if densities else 0.0
    norm_values = [d / max_density if max_density else 0.0 for d in densities]
    threshold = _percentile([v for v in norm_values if v > 0], 90)
    hotspots = []
    for (lat, lng, _density), val in zip(raw_points, norm_values):
        if val >= 0.02:
            point = [round(lat, 6), round(lng, 6), round(val, 4)]
            grid_points.append(point)
            if val >= threshold and threshold > 0:
                hotspots.append(point)
    hotspots = sorted(hotspots, key=lambda p: p[2], reverse=True)[:25]
    return {
        'type': 'python_kde_puro',
        'grid_size': f'{grid_size}x{grid_size}',
        'bandwidth': {'lat': round(bw_lat, 5), 'lng': round(bw_lng, 5), 'method': 'fallback'},
        'threshold_p90': round(threshold, 4),
        'coordenadas': grid_points,
        'hotspots': [
            {'rank': i + 1, 'latitud': h[0], 'longitud': h[1], 'densidad': h[2]}
            for i, h in enumerate(hotspots)
        ],
    }

@app.route('/api/dashboards/resumen', methods=['GET'])
def dashboards_resumen_old():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        resumen = {}
        for tabla, nombre in [('padron_nominal', 'Padron Nominal'),
                              ('cnv_cusco', 'CNV'),
                              ('tabla_vacunas', 'Vacunas'),
                              ('tabla_materno', 'Materno'),
                              ('tabla_iras_edas', 'IRAS/EDAS')]:
            try:
                cur.execute(f"SELECT COUNT(*) FROM {esquema}.{tabla}")
                resumen[nombre] = cur.fetchone()[0]
            except:
                resumen[nombre] = -1
        cur.close(); conn.close()
        return jsonify(resumen)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboards/iras-edas', methods=['GET'])
def dash_iras_edas():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        anio = request.args.get('anio', '2024')
        tabla_iras, anio_datos, fallback = _resolve_year_table(cur, esquema, 'iras_edas_{anio}', anio)
        filtros = ["anio::text=%s"]
        params_base = [anio_datos]
        if request.args.get('microred'):
            filtros.append("LOWER(microred) LIKE LOWER(%s)")
            params_base.append(f"%{request.args.get('microred')}%")
        if request.args.get('distrito'):
            filtros.append("LOWER(distrito) LIKE LOWER(%s)")
            params_base.append(f"%{request.args.get('distrito')}%")
        where_sql = " AND ".join(filtros)

        # Discover all indicator columns dynamically
        all_cols = _columns_for_table(cur, esquema, tabla_iras, min_ordinal=10)

        # Categorize columns
        eda_cols = [c for c in all_cols if c.startswith('eda_') or c.startswith('sosp_') or c.startswith('disent') or c.startswith('tto_')]
        ira_cols = [c for c in all_cols if c not in eda_cols]

        # IRA subcategories
        ira_subcats = {
            'ira_no_complicada': [c for c in ira_cols if 'IRA no complicada' in c or '3331101' in c],
            'faringoamigdalitis': [c for c in ira_cols if 'Faringoamigdalitis' in c or '3331102' in c],
            'otitis_media': [c for c in ira_cols if 'Otitis' in c or '3331103' in c],
            'sinusitis': [c for c in ira_cols if 'Sinusitis' in c or '3331104' in c],
            'neumonia': [c for c in ira_cols if 'Neumon' in c or '3331302' in c or '3331305' in c or 'IRAS con complicaciones' in c or '3331301' in c],
            'sob_asma': [c for c in ira_cols if 'SOB_ASMA' in c or 'SOB ASMA' in c],
        }
        # Remaining IRA cols
        ira_other = [c for c in ira_cols if not any(c in v for v in ira_subcats.values())]

        def quoted_sum(cols):
            if not cols:
                return '0'
            return '+'.join(f'COALESCE(SUM("{c}"),0)' for c in cols)

        # Monthly trend
        eda_expr = '+'.join(f'COALESCE(SUM("{c}"),0)' for c in eda_cols) if eda_cols else '0'
        ira_expr = '+'.join(f'COALESCE(SUM("{c}"),0)' for c in ira_cols) if ira_cols else '0'

        subcat_exprs = []
        for key, cols in ira_subcats.items():
            if cols:
                subcat_exprs.append(f'{quoted_sum(cols)} as {key}')
        if ira_other:
            subcat_exprs.append(f'{quoted_sum(ira_other)} as otras_iras')

        subcat_sql = ',' + ','.join(subcat_exprs) if subcat_exprs else ''

        cur.execute(f"""
            SELECT mes,
                {eda_expr} as total_eda,
                {ira_expr} as total_ira
                {subcat_sql}
            FROM {_qtable(esquema, tabla_iras)}
            WHERE {where_sql}
            GROUP BY mes ORDER BY mes
        """, params_base)
        monthly = []
        for r in cur.fetchall():
            entry = {'mes': int(r[0]), 'total_eda': int(r[1]), 'total_ira': int(r[2])}
            for i, (key, _) in enumerate(ira_subcats.items()):
                idx = 3 + i
                if idx < len(r) and r[idx] is not None:
                    entry[key] = int(r[idx])
            monthly.append(entry)

        # Top establishments
        cur.execute(f"""
            SELECT nombre_establecimiento,
                {eda_expr} as total_eda,
                {ira_expr} as total_ira
            FROM {_qtable(esquema, tabla_iras)} WHERE {where_sql}
            GROUP BY nombre_establecimiento
            ORDER BY ({eda_expr}+{ira_expr}) DESC LIMIT 10
        """, params_base)
        top_est = [{'nombre': r[0], 'total_eda': int(r[1]), 'total_ira': int(r[2])} for r in cur.fetchall()]

        # Summary
        cur.execute(f"""
            SELECT
                {eda_expr} as total_eda,
                {ira_expr} as total_ira,
                {quoted_sum(ira_subcats['neumonia'])} as total_neumonia
            FROM {_qtable(esquema, tabla_iras)} WHERE {where_sql}
        """, params_base)
        row = cur.fetchone()
        summary = {'total_eda': int(row[0]), 'total_ira': int(row[1]), 'total_neumonia': int(row[2])}
        cur.close(); conn.close()
        return jsonify({
            'monthly': monthly,
            'top_establecimientos': top_est,
            'summary': summary,
            'categorias': list(ira_subcats.keys()) + (['otras_iras'] if ira_other else []),
            'meta': {'anio_solicitado': str(anio), 'anio_datos': str(anio_datos), 'fallback': fallback, 'tabla': tabla_iras},
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboards/vacunacion', methods=['GET'])
def dash_vacunacion():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        anio = request.args.get('anio', '2024')
        tabla_pai, anio_datos, fallback = _resolve_year_table(cur, esquema, 'pai_{anio}', anio)
        filtros = ["anio::text=%s"]
        params_base = [anio_datos]
        for p in ['microred', 'distrito']:
            v = request.args.get(p)
            if v:
                filtros.append(f"LOWER({_qident(p)}) LIKE LOWER(%s)")
                params_base.append(f"%{v}%")
        where_sql = " AND ".join(filtros)

        # Get vaccination indicator columns dynamically from the resolved annual table.
        ind_cols = _columns_for_table(cur, esquema, tabla_pai, min_ordinal=11)

        # Build aggregated query dynamically
        agg_exprs = []
        col_alias = {}
        for idx, col in enumerate(ind_cols):
            alias = f'c{idx}'
            agg_exprs.append(f'COALESCE(SUM({_qcol(col)}),0) as {alias}')
            col_alias[col] = alias

        # Monthly trend
        agg_sql = ','.join(agg_exprs)
        cur.execute(f"SELECT mes,{agg_sql} FROM {_qtable(esquema, tabla_pai)} WHERE {where_sql} GROUP BY mes ORDER BY mes", params_base)
        rows = cur.fetchall()

        # Convert to per-vaccine monthly trend
        def cols_matching(*needles):
            needles = [n.lower() for n in needles]
            return [c for c in ind_cols if any(n in c.lower() for n in needles)]

        vaccine_groups = {
            'BCG': cols_matching('bcg'),
            'HVB': cols_matching('hvb'),
            'Penta': cols_matching('penta'),
            'IPV': cols_matching('ipv', 'apo_ipv'),
            'Neumococo': cols_matching('neumo'),
            'Rotavirus': cols_matching('rotav'),
            'Influenza': cols_matching('influenza'),
            'SPR': cols_matching('spr'),
            'DPT': cols_matching('dpt'),
            'Varicela': cols_matching('varicela'),
            'DT_pediatrico': cols_matching('dt pediatrico'),
            'Hepatitis_A': cols_matching('hepatitis'),
            'AMA': cols_matching('ama'),
            'Hib': cols_matching('hib'),
        }
        group_indexes = {
            vac_name: [ind_cols.index(col) for col in cols]
            for vac_name, cols in vaccine_groups.items()
        }

        monthly_trend = []
        for r in rows:
            entry = {'mes': int(r[0])}
            for vac_name, indexes in group_indexes.items():
                total = sum(int(r[idx + 1] or 0) for idx in indexes)
                if total > 0:
                    entry[vac_name] = total
            monthly_trend.append(entry)

        # Coverage by establishment (top 20)
        top_order = '+'.join([f'COALESCE(SUM({_qcol(c)}),0)' for c in ind_cols[:10]]) if ind_cols else '0'
        cur.execute(f"""
            SELECT nombre_establecimiento,{agg_sql}
            FROM {_qtable(esquema, tabla_pai)} WHERE {where_sql}
            GROUP BY nombre_establecimiento
            ORDER BY ({top_order}) DESC LIMIT 20
        """, params_base)
        est_rows = cur.fetchall()
        est_data = []
        for r in est_rows:
            total = sum(int(v) for v in r[1:] if v is not None)
            est_data.append({'nombre': r[0], 'total': total})

        # Summary (total doses)
        cur.execute(f"SELECT {agg_sql} FROM {_qtable(esquema, tabla_pai)} WHERE {where_sql}", params_base)
        totals = cur.fetchone()
        vaccine_summary = {}
        for vac_name, indexes in group_indexes.items():
            total = sum(int(totals[idx] or 0) for idx in indexes)
            if total > 0:
                vaccine_summary[vac_name] = total

        # Denominator available without creating tables: CNV of the data year.
        cur.execute(f"SELECT COUNT(*) FROM {_qtable(esquema, 'cnv_cusco')} WHERE periodo LIKE %s", (f'{anio_datos}%',))
        cnv_total = int(cur.fetchone()[0] or 0)
        def exact_cols(*names):
            wanted = {n.lower() for n in names}
            return [c for c in ind_cols if c.lower() in wanted]

        coverage_specs = {
            'BCG RN': exact_cols('bcg_12horas', 'bcg_12_24h', 'bcg_1_11m'),
            'HVB RN': exact_cols('hvb_12_24h', 'hvb_24h'),
            'Penta 3 <1a': exact_cols('Vacuna Penta3 men_1a'),
            'IPV 3 <1a': exact_cols('Vacuna Ipv3 men_1a'),
            'Neumococo 2 <1a': exact_cols('Vacuna neumo2 men_1a'),
            'Rotavirus 2 <1a': exact_cols('Vacuna rotav2 men_1a'),
            'SPR 1 1a': exact_cols('Vacuna SPR1 1a'),
        }
        coverage_summary = []
        for vac_name, cols in coverage_specs.items():
            if cols:
                total = sum(int(totals[ind_cols.index(col)] or 0) for col in cols)
                coverage_summary.append(_coverage_item(
                    vac_name,
                    total,
                    cnv_total,
                    f'CNV {anio_datos}',
                    metodo='estimado: dosis trazadora / nacimientos CNV'
                ))

        cur.close(); conn.close()
        return jsonify({
            'monthly_trend': monthly_trend,
            'by_establecimiento': est_data[:15],
            'vaccine_summary': vaccine_summary,
            'coverage_summary': coverage_summary,
            'denominador_cnv': cnv_total,
            'meta': {'anio_solicitado': str(anio), 'anio_datos': str(anio_datos), 'fallback': fallback, 'tabla': tabla_pai},
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboards/cred', methods=['GET'])
def dash_cred():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        anio = request.args.get('anio', '2024')
        tabla_cred, anio_datos, fallback = _resolve_year_table(cur, esquema, 'cred{anio}', anio)
        tabla_cred1, _, fallback1 = _resolve_year_table(cur, esquema, 'cred{anio}_1', anio)
        tabla_cred2, _, fallback2 = _resolve_year_table(cur, esquema, 'cred{anio}_2', anio)

        # Nutritional status from cred2024_1 (columns go up to 4a, except sobre_peso/obeso which include 5_11a)
        cur.execute(f"""
            SELECT
                COALESCE(SUM(desnutric_aguda_men1a+desnutric_aguda_1a+desnutric_aguda_2a+desnutric_aguda_3a+desnutric_aguda_4a),0) as desnutric_aguda,
                COALESCE(SUM(desnutric_cronica_men1a+desnutric_cronica_1a+desnutric_cronica_2a+desnutric_cronica_3a+desnutric_cronica_4a),0) as desnutric_cronica,
                COALESCE(SUM(desnutric_global_men1a+desnutric_global_1a+desnutric_global_2a+desnutric_global_3a+desnutric_global_4a),0) as desnutric_global,
                COALESCE(SUM(desnutric_severa_men1a+desnutric_severa_1a+desnutric_severa_2a+desnutric_severa_3a+desnutric_severa_4a),0) as desnutric_severa,
                COALESCE(SUM(sobre_peso_men1a+sobre_peso_1a+sobre_peso_2a+sobre_peso_3a+sobre_peso_4a+sobre_peso_5_11a),0) as sobre_peso,
                COALESCE(SUM(obeso_men1a+obeso_1a+obeso_2a+obeso_3a+obeso_4a+obeso_5_11a),0) as obeso
            FROM {_qtable(esquema, tabla_cred1)} WHERE anio::text=%s
        """, (anio_datos,))
        nut_row = cur.fetchone()
        nutritional = {
            'desnutric_aguda': int(nut_row[0]) if nut_row[0] else 0,
            'desnutric_cronica': int(nut_row[1]) if nut_row[1] else 0,
            'desnutric_global': int(nut_row[2]) if nut_row[2] else 0,
            'desnutric_severa': int(nut_row[3]) if nut_row[3] else 0,
            'sobre_peso': int(nut_row[4]) if nut_row[4] else 0,
            'obeso': int(nut_row[5]) if nut_row[5] else 0,
        }

        # Supplementation from cred2024 (use actual column names)
        cur.execute(f"""
            SELECT
                COALESCE(SUM(suplem_gest1+suplem_gest2+suplem_gest3+suplem_gest4+suplem_gest5+suplem_gest6+suplem_gest_ta),0) as suplem_gest,
                COALESCE(SUM(suplem_puer1+suplem_puer2+suplem_puer3+suplem_puer4+suplem_puer5+suplem_puer6+suplem_puer7),0) as suplem_puer,
                COALESCE(SUM(suplem_mef1+suplem_mef2+suplem_mef3+suplem_mef_ta),0) as suplem_mef,
                COALESCE(SUM(ta_suplem_bpn+suplem_4m_sano+suplem_5m_sano+ta_suplem_5_6m_sano+
                    suplem_1ra_6_11m+suplem_2da_6_11m+suplem_3ra_6_11m+suplem_4ta_6_11m+suplem_5ta_6_11m+suplem_6ta_6_11m+
                    suplem_1ra_1a+suplem_2da_1a+suplem_3ra_1a+suplem_4ta_1a+suplem_5ta_1a+suplem_6ta_1a+
                    suplem_1ra_2a+suplem_2da_2a+suplem_3ra_2a+suplem_4ta_2a+suplem_5ta_2a+suplem_6ta_2a+
                    suplem_1ra_3a+suplem_2da_3a+suplem_3ra_3a+suplem_4ta_3a+suplem_5ta_3a+suplem_6ta_3a+
                    suplem_1ra_4a+suplem_2da_4a+suplem_3ra_4a+suplem_4ta_4a+suplem_5ta_4a+suplem_6ta_4a+
                    suplem_1ra_5_11a+suplem_2da_5_11a+suplem_3ra_5_11a+
                    ta_suplem_1a+ta_suplem_2a+ta_suplem_3a+ta_suplem_4a+ta_suplem_5_11a+
                    suplem1_12_17+suplem2_12_17+suplem3_12_17+suplem4_12_17ta),0) as suplem_ninos
            FROM {_qtable(esquema, tabla_cred)} WHERE anio::text=%s
        """, (anio_datos,))
        sup_row = cur.fetchone()
        suplementacion = {
            'gestantes': int(sup_row[0]) if sup_row[0] else 0,
            'puerperas': int(sup_row[1]) if sup_row[1] else 0,
            'mef': int(sup_row[2]) if sup_row[2] else 0,
            'ninos': int(sup_row[3]) if sup_row[3] else 0,
        }

        # Developmental delays from cred2024_1 (columns: _m1a not _men1a)
        cur.execute(f"""
            SELECT
                COALESCE(SUM(retardo_desarrollo_len_m1a+retardo_desarrollo_len_1a+retardo_desarrollo_len_2a),0) as ret_len,
                COALESCE(SUM(retardo_desarrollo_mot_m1a+retardo_desarrollo_mot_1a+retardo_desarrollo_mot_2a),0) as ret_mot,
                COALESCE(SUM(retardo_desarrollo_soc_m1a+retardo_desarrollo_soc_1a+retardo_desarrollo_soc_2a),0) as ret_soc,
                COALESCE(SUM(retardo_desarrollo_coo_m1a+retardo_desarrollo_coo_1a+retardo_desarrollo_coo_2a),0) as ret_coo,
                COALESCE(SUM(retardo_desarrollo_cog_m1a+retardo_desarrollo_cog_1a+retardo_desarrollo_cog_2a),0) as ret_cog
            FROM {_qtable(esquema, tabla_cred1)} WHERE anio::text=%s
        """, (anio_datos,))
        dev_row = cur.fetchone()
        desarrollo = {
            'lenguaje': int(dev_row[0]) if dev_row[0] else 0,
            'motor': int(dev_row[1]) if dev_row[1] else 0,
            'social': int(dev_row[2]) if dev_row[2] else 0,
            'coordinacion': int(dev_row[3]) if dev_row[3] else 0,
            'cognitivo': int(dev_row[4]) if dev_row[4] else 0,
        }

        # Hemoglobin from cred2024_2
        cur.execute(f"""
            SELECT
                COALESCE(SUM(padron_total),0) as padron_total,
                COALESCE(SUM(total_ninos_bpn),0) as bpn,
                COALESCE(SUM(primer_dosaje_hb_bpn+segundo_dosaje_hb_bpn),0) as dosaje_bpn,
                COALESCE(SUM(hb_6_11m_primer+hb_6_11m_segundo),0) as hb_6_11m,
                COALESCE(SUM(hb_12_23m_primer+hb_12_23m_segundo+hb_12_23m_tercer),0) as hb_12_23m,
                COALESCE(SUM(hb_24_35m_primer+hb_24_35m_segundo),0) as hb_24_35m,
                COALESCE(SUM(hb_gestante_primer+hb_gestante_segundo+hb_gestante_tercero),0) as hb_gestante
            FROM {_qtable(esquema, tabla_cred2)} WHERE anio::text=%s
        """, (anio_datos,))
        hb_row = cur.fetchone()
        hemoglobina = {
            'padron_total': int(hb_row[0]), 'bpn': int(hb_row[1]),
            'dosaje_bpn': int(hb_row[2]), 'hb_6_11m': int(hb_row[3]),
            'hb_12_23m': int(hb_row[4]), 'hb_24_35m': int(hb_row[5]),
            'hb_gestante': int(hb_row[6]),
        }

        cur.execute(f"""
            SELECT COALESCE(MAX(total),0)::bigint
            FROM (
                SELECT mes, SUM(padron_total) as total
                FROM {_qtable(esquema, tabla_cred2)}
                WHERE anio::text=%s
                GROUP BY mes
            ) s
        """, (anio_datos,))
        padron_ref = int(cur.fetchone()[0] or 0)

        cred_cols = _columns_for_table(cur, esquema, tabla_cred)
        cred_final_cols = [c for c in [
            'cred11_men1a', 'cred6_1a', 'cred4_2a', 'cred4_3a', 'cred4_4a',
            'cred1_5a', 'cred1_6a', 'cred1_7a', 'cred1_8a', 'cred1_9a', 'cred1_10a', 'cred1_11a'
        ] if c in cred_cols]
        cur.execute(f"SELECT {_sum_expr(cred_final_cols)} FROM {_qtable(esquema, tabla_cred)} WHERE anio::text=%s", (anio_datos,))
        cred_final = int(cur.fetchone()[0] or 0)
        cobertura_cred = _coverage_item('CRED al dia estimado', cred_final, padron_ref, 'padron_total max mensual CRED', metodo='estimado: controles finales / padron_total max mensual')
        cobertura_hb = _coverage_item('Dosaje Hb ninos', hemoglobina['hb_6_11m'] + hemoglobina['hb_12_23m'] + hemoglobina['hb_24_35m'], padron_ref, 'padron_total max mensual CRED', metodo='estimado: dosajes Hb / padron_total max mensual')
        cobertura_sup = _coverage_item('Suplementacion ninos', suplementacion['ninos'], padron_ref, 'padron_total max mensual CRED', metodo='estimado: entregas / padron_total max mensual')
        cumplimiento_cred = {
            'al_dia_estimado': cred_final,
            'atrasado_estimado': max(padron_ref - cred_final, 0),
            'padron_referencia': padron_ref,
            'porcentaje': cobertura_cred['porcentaje'],
            'semaforo': cobertura_cred['semaforo'],
            'metodo': cobertura_cred['metodo'],
        }

        cur.close(); conn.close()
        return jsonify({
            'nutritional': nutritional,
            'suplementacion': suplementacion,
            'desarrollo': desarrollo,
            'hemoglobina': hemoglobina,
            'cumplimiento_cred': cumplimiento_cred,
            'coverage': [cobertura_cred, cobertura_hb, cobertura_sup],
            'meta': {'anio_solicitado': str(anio), 'anio_datos': str(anio_datos), 'fallback': fallback or fallback1 or fallback2, 'tablas': [tabla_cred, tabla_cred1, tabla_cred2]},
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboards/suplementacion', methods=['GET'])
def dash_suplementacion():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        anio = request.args.get('anio', '2024')
        tabla_cred, anio_datos, fallback = _resolve_year_table(cur, esquema, 'cred{anio}', anio)
        tabla_cred2, _, fallback2 = _resolve_year_table(cur, esquema, 'cred{anio}_2', anio)

        # Supplementation by establishment from cred2024
        cur.execute(f"""
            SELECT nombre_establecimiento,
                COALESCE(SUM(suplem_gest1+suplem_gest2+suplem_gest3+suplem_gest4+suplem_gest5+suplem_gest6+suplem_gest_ta),0) as suplem_gest,
                COALESCE(SUM(suplem_puer1+suplem_puer2+suplem_puer3+suplem_puer4+suplem_puer5+suplem_puer6+suplem_puer7),0) as suplem_puer,
                COALESCE(SUM(suplem_mef1+suplem_mef2+suplem_mef3+suplem_mef_ta),0) as suplem_mef,
                COALESCE(SUM(ta_suplem_bpn+suplem_4m_sano+suplem_5m_sano+ta_suplem_5_6m_sano+
                    suplem_1ra_6_11m+suplem_2da_6_11m+suplem_3ra_6_11m+suplem_4ta_6_11m+suplem_5ta_6_11m+suplem_6ta_6_11m+
                    suplem_1ra_1a+suplem_2da_1a+suplem_3ra_1a+suplem_4ta_1a+suplem_5ta_1a+suplem_6ta_1a+
                    suplem_1ra_2a+suplem_2da_2a+suplem_3ra_2a+suplem_4ta_2a+suplem_5ta_2a+suplem_6ta_2a+
                    suplem_1ra_3a+suplem_2da_3a+suplem_3ra_3a+suplem_4ta_3a+suplem_5ta_3a+suplem_6ta_3a+
                    suplem_1ra_4a+suplem_2da_4a+suplem_3ra_4a+suplem_4ta_4a+suplem_5ta_4a+suplem_6ta_4a+
                    suplem_1ra_5_11a+suplem_2da_5_11a+suplem_3ra_5_11a+
                    ta_suplem_1a+ta_suplem_2a+ta_suplem_3a+ta_suplem_4a+ta_suplem_5_11a+
                    suplem1_12_17+suplem2_12_17+suplem3_12_17+suplem4_12_17ta),0) as suplem_ninos
            FROM {_qtable(esquema, tabla_cred)} WHERE anio::text=%s
            GROUP BY nombre_establecimiento
            ORDER BY (COALESCE(SUM(suplem_gest1+suplem_gest2+suplem_gest3+suplem_gest4+suplem_gest5+suplem_gest6+suplem_gest_ta),0)+
                COALESCE(SUM(ta_suplem_bpn+suplem_4m_sano+suplem_5m_sano+ta_suplem_5_6m_sano+
                    suplem_1ra_6_11m+suplem_2da_6_11m+suplem_3ra_6_11m+suplem_4ta_6_11m+suplem_5ta_6_11m+suplem_6ta_6_11m+
                    suplem_1ra_1a+suplem_2da_1a+suplem_3ra_1a+suplem_4ta_1a+suplem_5ta_1a+suplem_6ta_1a+
                    suplem_1ra_2a+suplem_2da_2a+suplem_3ra_2a+suplem_4ta_2a+suplem_5ta_2a+suplem_6ta_2a+
                    suplem_1ra_3a+suplem_2da_3a+suplem_3ra_3a+suplem_4ta_3a+suplem_5ta_3a+suplem_6ta_3a+
                    suplem_1ra_4a+suplem_2da_4a+suplem_3ra_4a+suplem_4ta_4a+suplem_5ta_4a+suplem_6ta_4a+
                    suplem_1ra_5_11a+suplem_2da_5_11a+suplem_3ra_5_11a+
                    ta_suplem_1a+ta_suplem_2a+ta_suplem_3a+ta_suplem_4a+ta_suplem_5_11a+
                    suplem1_12_17+suplem2_12_17+suplem3_12_17+suplem4_12_17ta),0)) DESC LIMIT 20
        """, (anio_datos,))
        by_est = [{'nombre': r[0], 'gestantes': int(r[1]), 'puerperas': int(r[2]), 'mef': int(r[3]), 'ninos': int(r[4])} for r in cur.fetchall()]

        # Monthly trend
        cur.execute(f"""
            SELECT mes,
                COALESCE(SUM(suplem_gest1+suplem_gest2+suplem_gest3+suplem_gest4+suplem_gest5+suplem_gest6+suplem_gest_ta),0) as gest,
                COALESCE(SUM(ta_suplem_bpn+suplem_4m_sano+suplem_5m_sano+ta_suplem_5_6m_sano+
                    suplem_1ra_6_11m+suplem_2da_6_11m+suplem_3ra_6_11m+suplem_4ta_6_11m+suplem_5ta_6_11m+suplem_6ta_6_11m+
                    suplem_1ra_1a+suplem_2da_1a+suplem_3ra_1a+suplem_4ta_1a+suplem_5ta_1a+suplem_6ta_1a+
                    suplem_1ra_2a+suplem_2da_2a+suplem_3ra_2a+suplem_4ta_2a+suplem_5ta_2a+suplem_6ta_2a+
                    suplem_1ra_3a+suplem_2da_3a+suplem_3ra_3a+suplem_4ta_3a+suplem_5ta_3a+suplem_6ta_3a+
                    suplem_1ra_4a+suplem_2da_4a+suplem_3ra_4a+suplem_4ta_4a+suplem_5ta_4a+suplem_6ta_4a+
                    suplem_1ra_5_11a+suplem_2da_5_11a+suplem_3ra_5_11a+
                    ta_suplem_1a+ta_suplem_2a+ta_suplem_3a+ta_suplem_4a+ta_suplem_5_11a+
                    suplem1_12_17+suplem2_12_17+suplem3_12_17+suplem4_12_17ta),0) as ninos
            FROM {_qtable(esquema, tabla_cred)} WHERE anio::text=%s
            GROUP BY mes ORDER BY mes
        """, (anio_datos,))
        monthly = [{'mes': int(r[0]), 'gestantes': int(r[1]), 'ninos': int(r[2])} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT
                COALESCE(SUM(suplem_gest1+suplem_gest2+suplem_gest3+suplem_gest4+suplem_gest5+suplem_gest6+suplem_gest_ta),0) as gestantes,
                COALESCE(SUM(ta_suplem_bpn+suplem_4m_sano+suplem_5m_sano+ta_suplem_5_6m_sano+
                    suplem_1ra_6_11m+suplem_2da_6_11m+suplem_3ra_6_11m+suplem_4ta_6_11m+suplem_5ta_6_11m+suplem_6ta_6_11m+
                    suplem_1ra_1a+suplem_2da_1a+suplem_3ra_1a+suplem_4ta_1a+suplem_5ta_1a+suplem_6ta_1a+
                    suplem_1ra_2a+suplem_2da_2a+suplem_3ra_2a+suplem_4ta_2a+suplem_5ta_2a+suplem_6ta_2a+
                    suplem_1ra_3a+suplem_2da_3a+suplem_3ra_3a+suplem_4ta_3a+suplem_5ta_3a+suplem_6ta_3a+
                    suplem_1ra_4a+suplem_2da_4a+suplem_3ra_4a+suplem_4ta_4a+suplem_5ta_4a+suplem_6ta_4a+
                    suplem_1ra_5_11a+suplem_2da_5_11a+suplem_3ra_5_11a+
                    ta_suplem_1a+ta_suplem_2a+ta_suplem_3a+ta_suplem_4a+ta_suplem_5_11a+
                    suplem1_12_17+suplem2_12_17+suplem3_12_17+suplem4_12_17ta),0) as ninos
            FROM {_qtable(esquema, tabla_cred)} WHERE anio::text=%s
        """, (anio_datos,))
        total_sup = cur.fetchone()
        total_gest = int(total_sup[0] or 0)
        total_ninos = int(total_sup[1] or 0)

        cur.execute(f"""
            SELECT COALESCE(MAX(total),0)::bigint
            FROM (
                SELECT mes, SUM(padron_total) as total
                FROM {_qtable(esquema, tabla_cred2)}
                WHERE anio::text=%s
                GROUP BY mes
            ) s
        """, (anio_datos,))
        padron_ref = int(cur.fetchone()[0] or 0)

        gestantes_ref = 0
        if _table_exists(cur, esquema, 'tabla_materno'):
            cur.execute(f"""
                SELECT COUNT(DISTINCT dni_paciente)
                FROM {_qtable(esquema, 'tabla_materno')}
                WHERE anio::text=%s AND UPPER(COALESCE(condicion_gestante,''))='GESTANTE'
            """, (anio_datos,))
            gestantes_ref = int(cur.fetchone()[0] or 0)

        coverage = [
            _coverage_item('Suplementacion ninos', total_ninos, padron_ref, 'padron_total max mensual CRED', metodo='estimado: entregas / padron_total max mensual'),
            _coverage_item('Suplementacion gestantes', total_gest, gestantes_ref, 'gestantes unicas HIS', metodo='estimado: entregas / gestantes unicas HIS'),
        ]

        cur.close(); conn.close()
        return jsonify({
            'by_establecimiento': by_est,
            'monthly': monthly,
            'coverage': coverage,
            'summary': {'gestantes': total_gest, 'ninos': total_ninos, 'padron_ref': padron_ref, 'gestantes_ref': gestantes_ref},
            'meta': {'anio_solicitado': str(anio), 'anio_datos': str(anio_datos), 'fallback': fallback or fallback2, 'tablas': [tabla_cred, tabla_cred2]},
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboards/materno', methods=['GET'])
def dash_materno():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        anio = request.args.get('anio', '2024')
        params = [anio]
        filtros = ["anio=%s"]
        microred = request.args.get('microred')
        distrito = request.args.get('distrito')
        if microred:
            filtros.append("LOWER(microred) LIKE LOWER(%s)")
            params.append(f"%{microred}%")
        if distrito:
            filtros.append("LOWER(distrito) LIKE LOWER(%s)")
            params.append(f"%{distrito}%")
        where_sql = " AND ".join(filtros)

        # Monthly trend from tabla_materno
        cur.execute(f"""
            SELECT mes, COUNT(*) as total
            FROM {esquema}.tabla_materno WHERE {where_sql} AND mes IS NOT NULL
            GROUP BY mes ORDER BY mes
        """, params)
        monthly = [{'mes': int(r[0]), 'total': r[1]} for r in cur.fetchall()]

        # By establishment
        cur.execute(f"""
            SELECT nombre_establecimiento, COUNT(*) as total
            FROM {esquema}.tabla_materno WHERE {where_sql} AND nombre_establecimiento IS NOT NULL
            GROUP BY nombre_establecimiento ORDER BY total DESC LIMIT 15
        """, params)
        by_est = [{'nombre': r[0], 'total': r[1]} for r in cur.fetchall()]

        # Territory, age and diagnostic views
        cur.execute(f"""
            SELECT distrito, COUNT(*) as total
            FROM {esquema}.tabla_materno WHERE {where_sql} AND distrito IS NOT NULL AND distrito != ''
            GROUP BY distrito ORDER BY total DESC LIMIT 15
        """, params)
        by_distrito = [{'distrito': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT microred, COUNT(*) as total
            FROM {esquema}.tabla_materno WHERE {where_sql} AND microred IS NOT NULL AND microred != ''
            GROUP BY microred ORDER BY total DESC LIMIT 15
        """, params)
        by_microred = [{'microred': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT tipo_diagnostico, COUNT(*) as total
            FROM {esquema}.tabla_materno WHERE {where_sql} AND tipo_diagnostico IS NOT NULL AND tipo_diagnostico != ''
            GROUP BY tipo_diagnostico ORDER BY total DESC
        """, params)
        by_tipo_diagnostico = [{'tipo': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT condicion_gestante, COUNT(*) as total
            FROM {esquema}.tabla_materno WHERE {where_sql} AND condicion_gestante IS NOT NULL AND condicion_gestante != ''
            GROUP BY condicion_gestante ORDER BY total DESC LIMIT 10
        """, params)
        by_condicion = [{'condicion': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT codigo_item, COUNT(*) as total, COUNT(DISTINCT dni_paciente) as pacientes
            FROM {esquema}.tabla_materno WHERE {where_sql} AND codigo_item IS NOT NULL AND codigo_item != ''
            GROUP BY codigo_item ORDER BY total DESC LIMIT 15
        """, params)
        by_codigo = [{'codigo': r[0], 'total': r[1], 'pacientes': r[2]} for r in cur.fetchall()]

        cur.execute(f"""
            WITH grupos AS (
                SELECT CASE
                    WHEN edad IS NULL THEN 'Sin dato'
                    WHEN edad < 12 THEN '<12'
                    WHEN edad BETWEEN 12 AND 17 THEN '12-17'
                    WHEN edad BETWEEN 18 AND 29 THEN '18-29'
                    WHEN edad BETWEEN 30 AND 34 THEN '30-34'
                    ELSE '35+'
                END as grupo
                FROM {esquema}.tabla_materno WHERE {where_sql}
            )
            SELECT grupo, COUNT(*) as total
            FROM grupos
            GROUP BY grupo
            ORDER BY CASE grupo WHEN '<12' THEN 1 WHEN '12-17' THEN 2 WHEN '18-29' THEN 3 WHEN '30-34' THEN 4 WHEN '35+' THEN 5 ELSE 6 END
        """, params)
        by_edad = [{'grupo': r[0], 'total': r[1]} for r in cur.fetchall()]

        # Summary
        cur.execute(f"""
            SELECT COUNT(*) as total,
                   COUNT(DISTINCT dni_paciente) as pacientes,
                   COUNT(DISTINCT id_establecimiento) as establecimientos,
                   COUNT(DISTINCT distrito) as distritos,
                   ROUND(AVG(edad)::numeric, 1) as edad_promedio
            FROM {esquema}.tabla_materno WHERE {where_sql}
        """, params)
        row = cur.fetchone()
        summary = {
            'total_atenciones': int(row[0] or 0),
            'pacientes': int(row[1] or 0),
            'establecimientos': int(row[2] or 0),
            'distritos': int(row[3] or 0),
            'edad_promedio': float(row[4]) if row[4] is not None else 0,
        }

        cur.close(); conn.close()
        return jsonify({
            'monthly': monthly,
            'by_establecimiento': by_est,
            'by_distrito': by_distrito,
            'by_microred': by_microred,
            'by_tipo_diagnostico': by_tipo_diagnostico,
            'by_condicion': by_condicion,
            'by_codigo': by_codigo,
            'by_edad': by_edad,
            'summary': summary,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboards/poblacion', methods=['GET'])
def dash_poblacion():
    try:
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        anio = request.args.get('anio', '2024')
        cnv_anio = f"{anio}%"

        # Padron by distrito
        cur.execute(f"""
            SELECT dist_nino, COUNT(*) as total,
                   COUNT(*) FILTER (WHERE latitud IS NOT NULL AND longitud IS NOT NULL) as con_coord
            FROM {esquema}.padron_nominal
            WHERE dist_nino IS NOT NULL AND dist_nino != ''
            GROUP BY dist_nino ORDER BY total DESC
        """)
        by_distrito = [{'distrito': r[0], 'total': r[1], 'con_coord': r[2]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT nombre_eess, COUNT(*) as total
            FROM {esquema}.padron_nominal
            WHERE nombre_eess IS NOT NULL AND nombre_eess != ''
            GROUP BY nombre_eess ORDER BY total DESC LIMIT 15
        """)
        by_eess = [{'nombre': r[0], 'total': r[1]} for r in cur.fetchall()]

        # By gender
        cur.execute(f"""
            SELECT CASE codsexo WHEN '1' THEN 'Masculino' WHEN '2' THEN 'Femenino' ELSE codsexo END as sexo, COUNT(*) as total
            FROM {esquema}.padron_nominal WHERE codsexo IS NOT NULL AND codsexo != ''
            GROUP BY sexo ORDER BY sexo
        """)
        by_genero = [{'genero': r[0], 'total': r[1]} for r in cur.fetchall()]

        # By age group
        cur.execute(f"""
            WITH base AS (
                SELECT CASE
                    WHEN edad_actual ~ '^[0-9]+' THEN substring(edad_actual from '^([0-9]+)')::int
                    ELSE NULL
                END as edad_anios
                FROM {esquema}.padron_nominal
            )
            SELECT
                COUNT(*) FILTER (WHERE edad_anios BETWEEN 0 AND 1) as g_0_1,
                COUNT(*) FILTER (WHERE edad_anios BETWEEN 2 AND 5) as g_2_5,
                COUNT(*) FILTER (WHERE edad_anios BETWEEN 6 AND 11) as g_6_11,
                COUNT(*) FILTER (WHERE edad_anios >= 12) as g_12_mas,
                COUNT(*) FILTER (WHERE edad_anios IS NULL) as sin_edad
            FROM base
        """)
        row = cur.fetchone()
        by_edad = {
            '0-1 años': int(row[0] or 0), '2-5 años': int(row[1] or 0),
            '6-11 años': int(row[2] or 0), '12+ años': int(row[3] or 0),
            'sin_dato': int(row[4] or 0),
        }

        cur.execute(f"""
            SELECT
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE latitud IS NOT NULL AND longitud IS NOT NULL AND latitud != '' AND longitud != '') as con_coord,
                COUNT(*) FILTER (WHERE celular IS NOT NULL AND celular != '') as con_celular,
                COUNT(*) FILTER (WHERE direccion IS NOT NULL AND direccion != '') as con_direccion,
                COUNT(*) FILTER (WHERE nombre_eess IS NOT NULL AND nombre_eess != '') as con_eess,
                COUNT(*) FILTER (WHERE tip_seguro IS NOT NULL AND tip_seguro != '') as con_seguro,
                COUNT(*) FILTER (WHERE prog_social IS NOT NULL AND prog_social != '') as con_prog_social,
                COUNT(*) FILTER (WHERE UPPER(TRIM(menor_visitado)) = 'VISITADO') as visitados,
                COUNT(*) FILTER (WHERE UPPER(TRIM(menor_visitado)) = 'NO VISITADO') as no_visitados
            FROM {esquema}.padron_nominal
        """)
        row = cur.fetchone()
        completitud = {
            'total': int(row[0] or 0),
            'con_coord': int(row[1] or 0),
            'con_celular': int(row[2] or 0),
            'con_direccion': int(row[3] or 0),
            'con_eess': int(row[4] or 0),
            'con_seguro': int(row[5] or 0),
            'con_prog_social': int(row[6] or 0),
            'visitados': int(row[7] or 0),
            'no_visitados': int(row[8] or 0),
        }

        cur.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(tip_seguro), ''), 'Sin dato') as seguro, COUNT(*) as total
            FROM {esquema}.padron_nominal
            GROUP BY seguro ORDER BY total DESC LIMIT 10
        """)
        by_seguro = [{'seguro': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(fuente_datos), ''), 'Sin dato') as fuente, COUNT(*) as total
            FROM {esquema}.padron_nominal
            GROUP BY fuente ORDER BY total DESC LIMIT 10
        """)
        by_fuente = [{'fuente': r[0], 'total': r[1]} for r in cur.fetchall()]

        # Total
        cur.execute(f"SELECT COUNT(*) FROM {esquema}.padron_nominal")
        total = cur.fetchone()[0]

        # CNV indicators for selected year
        cur.execute(f"""
            WITH base AS (
                SELECT
                    NULLIF(regexp_replace(peso_nacido, '[^0-9.]', '', 'g'), '')::numeric as peso,
                    NULLIF(regexp_replace(talla_nacido, '[^0-9.]', '', 'g'), '')::numeric as talla,
                    NULLIF(regexp_replace(apgar_5_nacido, '[^0-9.]', '', 'g'), '')::numeric as apgar,
                    NULLIF(regexp_replace(dur_emb_parto, '[^0-9.]', '', 'g'), '')::int as semanas,
                    NULLIF(regexp_replace(edad_madre, '[^0-9]', '', 'g'), '')::int as edad_madre_num,
                    tipo_parto, condicion_parto, sexo_nacido, lactancia_precoz, malformacion_congenita,
                    dist_madre, nombre_eess, periodo
                FROM {esquema}.cnv_cusco
                WHERE periodo LIKE %s
            )
            SELECT
                COUNT(*) as total,
                COUNT(*) FILTER (WHERE peso < 2500) as bajo_peso,
                COUNT(*) FILTER (WHERE peso < 1500) as muy_bajo_peso,
                COUNT(*) FILTER (WHERE peso >= 4000) as macrosomia,
                COUNT(*) FILTER (WHERE semanas < 37) as prematuros,
                COUNT(*) FILTER (WHERE UPPER(COALESCE(condicion_parto,'')) LIKE '%%CESAREA%%' OR UPPER(COALESCE(tipo_parto,'')) LIKE '%%CESAREA%%') as cesareas,
                COUNT(*) FILTER (WHERE UPPER(COALESCE(lactancia_precoz,'')) IN ('SI','S','1') OR UPPER(COALESCE(lactancia_precoz,'')) LIKE 'SI%%') as lactancia,
                COUNT(*) FILTER (WHERE malformacion_congenita IS NOT NULL AND TRIM(malformacion_congenita) != '' AND UPPER(TRIM(malformacion_congenita)) NOT IN ('NO','NINGUNA')) as malformacion,
                COUNT(*) FILTER (WHERE apgar < 7) as apgar_bajo,
                COUNT(*) FILTER (WHERE edad_madre_num < 18) as madre_menor_18,
                COUNT(*) FILTER (WHERE edad_madre_num >= 35) as madre_35_mas,
                ROUND(AVG(peso)::numeric, 1) as peso_promedio,
                ROUND(AVG(edad_madre_num)::numeric, 1) as edad_madre_promedio
            FROM base
        """, (cnv_anio,))
        row = cur.fetchone()
        cnv_summary = {
            'anio': anio,
            'total': int(row[0] or 0),
            'bajo_peso': int(row[1] or 0),
            'muy_bajo_peso': int(row[2] or 0),
            'macrosomia': int(row[3] or 0),
            'prematuros': int(row[4] or 0),
            'cesareas': int(row[5] or 0),
            'lactancia_precoz': int(row[6] or 0),
            'malformacion': int(row[7] or 0),
            'apgar_bajo': int(row[8] or 0),
            'madre_menor_18': int(row[9] or 0),
            'madre_35_mas': int(row[10] or 0),
            'peso_promedio': float(row[11]) if row[11] is not None else 0,
            'edad_madre_promedio': float(row[12]) if row[12] is not None else 0,
        }

        cur.execute(f"""
            SELECT substring(periodo from 5 for 2)::int as mes,
                   COUNT(*) as total,
                   COUNT(*) FILTER (WHERE NULLIF(regexp_replace(peso_nacido, '[^0-9.]', '', 'g'), '')::numeric < 2500) as bajo_peso
            FROM {esquema}.cnv_cusco
            WHERE periodo LIKE %s AND periodo ~ '^[0-9]{{6}}$'
            GROUP BY mes ORDER BY mes
        """, (cnv_anio,))
        cnv_monthly = [{'mes': int(r[0]), 'total': r[1], 'bajo_peso': r[2]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(sexo_nacido), ''), 'Sin dato') as sexo, COUNT(*) as total
            FROM {esquema}.cnv_cusco WHERE periodo LIKE %s
            GROUP BY sexo ORDER BY total DESC
        """, (cnv_anio,))
        cnv_by_sexo = [{'sexo': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            WITH base AS (
                SELECT NULLIF(regexp_replace(edad_madre, '[^0-9]', '', 'g'), '')::int as edad
                FROM {esquema}.cnv_cusco WHERE periodo LIKE %s
            ), grupos AS (
                SELECT CASE
                        WHEN edad IS NULL THEN 'Sin dato'
                        WHEN edad < 18 THEN '<18'
                        WHEN edad BETWEEN 18 AND 29 THEN '18-29'
                        WHEN edad BETWEEN 30 AND 34 THEN '30-34'
                        ELSE '35+'
                    END as grupo
                FROM base
            )
            SELECT grupo, COUNT(*) as total
            FROM grupos GROUP BY grupo
            ORDER BY CASE grupo WHEN '<18' THEN 1 WHEN '18-29' THEN 2 WHEN '30-34' THEN 3 WHEN '35+' THEN 4 ELSE 5 END
        """, (cnv_anio,))
        cnv_by_edad_madre = [{'grupo': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            WITH base AS (
                SELECT NULLIF(regexp_replace(peso_nacido, '[^0-9.]', '', 'g'), '')::numeric as peso
                FROM {esquema}.cnv_cusco WHERE periodo LIKE %s
            ), grupos AS (
                SELECT CASE
                        WHEN peso IS NULL THEN 'Sin dato'
                        WHEN peso < 1500 THEN '<1500g'
                        WHEN peso < 2500 THEN '1500-2499g'
                        WHEN peso < 4000 THEN '2500-3999g'
                        ELSE '>=4000g'
                    END as grupo
                FROM base
            )
            SELECT grupo, COUNT(*) as total
            FROM grupos GROUP BY grupo
            ORDER BY CASE grupo WHEN '<1500g' THEN 1 WHEN '1500-2499g' THEN 2 WHEN '2500-3999g' THEN 3 WHEN '>=4000g' THEN 4 ELSE 5 END
        """, (cnv_anio,))
        cnv_by_peso = [{'grupo': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(condicion_parto), ''), 'Sin dato') as condicion, COUNT(*) as total
            FROM {esquema}.cnv_cusco WHERE periodo LIKE %s
            GROUP BY condicion ORDER BY total DESC LIMIT 10
        """, (cnv_anio,))
        cnv_by_parto = [{'condicion': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(dist_madre), ''), 'Sin dato') as distrito, COUNT(*) as total
            FROM {esquema}.cnv_cusco WHERE periodo LIKE %s
            GROUP BY distrito ORDER BY total DESC LIMIT 15
        """, (cnv_anio,))
        cnv_by_distrito = [{'distrito': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.execute(f"""
            SELECT COALESCE(NULLIF(TRIM(nombre_eess), ''), 'Sin dato') as nombre, COUNT(*) as total
            FROM {esquema}.cnv_cusco WHERE periodo LIKE %s
            GROUP BY nombre ORDER BY total DESC LIMIT 15
        """, (cnv_anio,))
        cnv_by_eess = [{'nombre': r[0], 'total': r[1]} for r in cur.fetchall()]

        cur.close(); conn.close()
        return jsonify({
            'total': total,
            'by_distrito': by_distrito,
            'by_eess': by_eess,
            'by_genero': by_genero,
            'by_edad': by_edad,
            'completitud': completitud,
            'by_seguro': by_seguro,
            'by_fuente': by_fuente,
            'cnv': {
                'summary': cnv_summary,
                'monthly': cnv_monthly,
                'by_sexo': cnv_by_sexo,
                'by_edad_madre': cnv_by_edad_madre,
                'by_peso': cnv_by_peso,
                'by_parto': cnv_by_parto,
                'by_distrito': cnv_by_distrito,
                'by_eess': cnv_by_eess,
            },
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== EDITOR ====================
EDITOR_CREDENTIALS = {'usuario': 'ivan', 'password': 'ivanar'}

@app.route('/api/editor/login', methods=['POST'])
def editor_login():
    data = request.json
    if data.get('usuario') == EDITOR_CREDENTIALS['usuario'] and data.get('password') == EDITOR_CREDENTIALS['password']:
        session['editor'] = True
        return jsonify({'exito': True})
    return jsonify({'exito': False, 'mensaje': 'Credenciales inválidas'}), 401

@app.route('/api/editor/check', methods=['GET'])
def editor_check():
    return jsonify({'activo': session.get('editor', False)})

@app.route('/api/editor/logout', methods=['POST'])
def editor_logout():
    session.pop('editor', None)
    return jsonify({'exito': True})

@app.route('/api/editor/scripts', methods=['GET'])
def editor_scripts():
    config = _cargar_config_editor()
    return jsonify({
        'reportes': config.get('botones', []),
        'maestros': SCRIPTS_MAESTROS_EDITABLES,
    })

@app.route('/api/editor/script-content', methods=['POST'])
def editor_script_content():
    data = request.json
    ruta = data.get('ruta', '')
    ruta_abs = str(PROJECT_ROOT / ruta)
    if not os.path.exists(ruta_abs):
        return jsonify({'error': 'Archivo no encontrado'}), 404
    with open(ruta_abs, 'r', encoding='utf-8', errors='replace') as f:
        contenido = f.read()
    return jsonify({'contenido': contenido, 'ruta': ruta})

@app.route('/api/editor/script-save', methods=['POST'])
def editor_script_save():
    data = request.json
    ruta = data.get('ruta', '')
    contenido = data.get('contenido', '')
    ruta_abs = str(PROJECT_ROOT / ruta)
    with open(ruta_abs, 'w', encoding='utf-8') as f:
        f.write(contenido)
    return jsonify({'mensaje': 'Guardado correctamente'})

@app.route('/api/editor/restore', methods=['POST'])
def editor_restore():
    config = {'botones': BOTONES_REPORTE_PREDETERMINADOS}
    _guardar_config_editor(config)
    return jsonify({'mensaje': 'Scripts restaurados a valores originales'})

# ==================== TUNNEL (Cloudflare) ====================
_tunnel_proc: Optional[subprocess.Popen] = None
_tunnel_url: str = ""

@app.route('/api/tunnel/start', methods=['POST'])
def tunnel_start():
    global _tunnel_proc, _tunnel_url
    if _tunnel_proc and _tunnel_proc.poll() is None:
        return jsonify({'exito': True, 'url': _tunnel_url, 'mensaje': 'Tunnel ya activo'})
    try:
        import shutil, platform as _plt
        if _plt.system() == "Windows":
            cf_bin = shutil.which("cloudflared") or str(PROJECT_ROOT / "tunnel" / "cloudflared.exe")
        else:
            cf_bin = shutil.which("cloudflared") or "/tmp/cloudflared"
        if not os.path.exists(cf_bin):
            return jsonify({'exito': False, 'mensaje': 'cloudflared no encontrado. Descargalo de https://developers.cloudflare.com/cloudflare-one/connections/connect-apps/'})
        logfile = str(BASE_DIR / "tunnel" / "tunnel.log")
        os.makedirs(os.path.dirname(logfile), exist_ok=True)
        _tunnel_proc = subprocess.Popen(
            [cf_bin, "tunnel", "--url", "http://127.0.0.1:5000", "--logfile", logfile],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        _tunnel_url = ""
        return jsonify({'exito': True, 'mensaje': 'Tunnel iniciando...'})
    except Exception as e:
        return jsonify({'exito': False, 'mensaje': str(e)})

@app.route('/api/tunnel/url', methods=['GET'])
def tunnel_url():
    global _tunnel_proc, _tunnel_url
    if _tunnel_url:
        return jsonify({'url': _tunnel_url, 'listo': True})
    if _tunnel_proc is None or _tunnel_proc.poll() is not None:
        return jsonify({'url': '', 'listo': False, 'mensaje': 'Tunnel no activo'})
    logfile = str(PROJECT_ROOT / "tunnel" / "tunnel.log")
    if os.path.exists(logfile):
        try:
            with open(logfile, 'r') as f:
                content = f.read()
            m = re.search(r'https://[a-z0-9.-]+\.trycloudflare\.com', content)
            if m:
                _tunnel_url = m.group()
                return jsonify({'url': _tunnel_url, 'listo': True})
        except Exception:
            pass
    return jsonify({'url': '', 'listo': False, 'mensaje': 'Esperando URL...'})

@app.route('/api/tunnel/stop', methods=['POST'])
def tunnel_stop():
    global _tunnel_proc
    if _tunnel_proc and _tunnel_proc.poll() is None:
        _tunnel_proc.terminate()
        _tunnel_proc.wait(timeout=5)
        _tunnel_proc = None
        return jsonify({'exito': True, 'mensaje': 'Tunnel detenido'})
    return jsonify({'exito': True, 'mensaje': 'No hay tunnel activo'})

@app.route('/api/tunnel/status', methods=['GET'])
def tunnel_status():
    global _tunnel_proc, _tunnel_url
    activo = _tunnel_proc is not None and _tunnel_proc.poll() is None
    return jsonify({'activo': activo, 'url': _tunnel_url if activo else ''})

# ---------- Dashboard Table Generation ----------
# ---------- Dashboard Check Year ----------
@app.route('/api/dashboards/check-year', methods=['GET'])
def dashboards_check_year():
    anio = request.args.get('anio', '')
    anio = str(anio)
    conn = _db_cursor()
    cur = conn.cursor()
    esquema = _get_esquema()
    checks = {}
    for tipo, tbl_template in [('pai', 'pai_{anio}'), ('cred', 'cred{anio}'),
                                ('cred_1', 'cred{anio}_1'), ('cred_2', 'cred{anio}_2'),
                                ('iras_edas', 'iras_edas_{anio}')]:
        tbl = tbl_template.format(anio=anio)
        exists = _table_exists(cur, esquema, tbl)
        has_year = _table_has_year(cur, esquema, tbl, anio) if exists else False
        count = 0
        if exists:
            cur.execute(f'SELECT COUNT(*) FROM {_qtable(esquema, tbl)}')
            count = int(cur.fetchone()[0])
        checks[tipo + '_table'] = tbl
        checks[tipo + '_exists'] = exists
        checks[tipo + '_has_data'] = has_year
        checks[tipo + '_count'] = count
    for src in ['tabla_vacunas', 'tabla_iras_edas', 'tabla_materno']:
        if _table_exists(cur, esquema, src):
            has_yr = _table_has_year(cur, esquema, src, anio)
            cur.execute(f'SELECT COUNT(*) FROM {_qtable(esquema, src)} WHERE anio::text=%s', (anio,))
            cnt = int(cur.fetchone()[0])
            checks[f'source_{src}'] = {'exists': True, 'has_year': has_yr, 'count': cnt}
        else:
            checks[f'source_{src}'] = {'exists': False}
    cur.close()
    conn.close()
    return jsonify({'anio': anio, 'checks': checks})

# ---------- Reportes MINSA (Formatos Excel) ----------
CIE_SOB_ASMA = {
    'sob_asma': {'label': 'SOB/Asma', 'codes': ['J210','J211','J218','J219','J440','J441','J448','J449','J450','J451','J452','J453','J458','J459','J46X']},
}
CIE_ANEMIA = {
    'anemia_ferropenica': {'label': 'Anemia Ferropénica', 'codes': ['D500','D501','D508','D509']},
    'dosaje_hemoglobina': {'label': 'Dosaje de Hemoglobina', 'codes': ['85018','85018.01']},
}
CIE_PARASITOSIS = {
    'parasitosis': {'label': 'Parasitosis Intestinal', 'codes': ['B680','B681','B689','B700','B701','B760','B761','B779','B780','B79X','B820','B829','A070','A071','B663','B664','B80X','87178']},
}
CIE_DEFUNCIONES = {
    'defuncion': {'label': 'Defunciones', 'codes': ['R96X','R98X','R99X','I469','I490']},
}
CIE_SAL_YODADA = {
    'sal_yodada': {'label': 'Consumo de Sal Yodada', 'codes': ['Z720','Z721','Z722']},
}
CIE_IRAS = {
    'ira_no_complicada': {'label': 'IRA no complicada', 'codes': ['J00X','J040','J041','J042','J060','J068','J069','J209','J205','J206','J210','J211','J218','J219']},
    'faringoamigdalitis': {'label': 'Faringoamigdalitis Aguda', 'codes': ['J020','J029','J030','J038','J039']},
    'oma': {'label': 'Otitis Media Aguda', 'codes': ['H650','H651','H660','H669']},
    'sinusitis': {'label': 'Sinusitis Aguda', 'codes': ['J010','J011','J012','J013','J014','J019']},
    'neumonia_sin_comp': {'label': 'Neumonía sin complicaciones', 'codes': ['J129','J159','J189']},
    'ira_con_complicaciones': {'label': 'IRA con complicaciones', 'codes': ['A369','A370','A371','A378','A379','J120','J121','J122','J128','J129','J181','J188','J189','J100','J101','J108','J109','J110','J111','J118','J119','J158','J159','J168','J180','J189']},
    'neumonia_grave': {'label': 'Neumonía Grave <2m', 'codes': ['J050','J051','J851','J860','J869','J90X','J918','J938','J940','J941','J942','J943','J948','J949']},

}
CIE_EDAS = {
    'eda_acuosa': {'label': 'EDA Acuosa', 'codes': ['A000','A001','A002','A003','A004','A005','A006','A007','A008','A009']},
    'eda_disenterica': {'label': 'EDA Disentérica', 'codes': ['A030','A031','A032','A033','A038','A039','A060','A061','A062','A063','A064','A065','A066','A067','A068','A069','A070','A071','A072','A073','A078','A079']},
    'eda_persistente': {'label': 'EDA Persistente', 'codes': ['A090','A099']},
}

CRED_CODES = ['99381','99381.01','99382','99383',
    '99401','99401.03','99401.05','99401.07','99401.08','99401.09',
    '99401.12','99401.16','99401.24','99401.25',
    '99402','99403','99404',
    '99411.01','99431','99436',
    '99199','99199.26','99199.17','99199.28',
    '99211','P929']

SUP_CODES = ['U1692','59430','U140','99403.01','99199.26','99199.17','99199.27']
ANE_CODES = ['85018.01','85018','C0011']

ATENCION_RN_CODES = ['99436', '99431', '99431.01', '99431.02', '99502', '36416', '94760']
ATENCION_RN_NAMES = {
    '99436': 'Atenci\u00f3n Inmediata RN',
    '99431': 'Tamizaje Neonatal',
    '99431.01': 'Tamizaje Hipoacusia',
    '99431.02': 'Tamizaje Catarata Cong\u00e9nita',
    '99502': 'Cuidado y Evaluaci\u00f3n Neonatal',
    '36416': 'Tamizaje Toma de Muestra',
    '94760': 'Tamizaje Cardiopat\u00eda'
}

SESIONES_CODES = ['99411.01']
SESIONES_LABELS = {'1': '1ra Sesi\u00f3n', '2': '2da Sesi\u00f3n', '3': '3ra Sesi\u00f3n', '4': '4ta Sesi\u00f3n'}

LACTANCIA_CODES = ['99401.03', 'P929']
LACTANCIA_NAMES = {
    '99401.03': 'LME Primera Hora',
    'P929': 'Suspensi\u00f3n LME 6m'
}

EVAL_DESARROLLO_CODES = ['R620']
EVAL_DESARROLLO_AREAS = {
    'LEN': 'Lenguaje',
    'MOT': 'Motora',
    'SOC': 'Social',
    'COO': 'Coordinaci\u00f3n',
    'COG': 'Cognitivo'
}

PLAN_INTEGRAL_CODES = ['C8002', 'Z001']
PLAN_INTEGRAL_NAMES = {
    'C8002': 'Plan AIS Inicio/T\u00e9rmino',
    'Z001': 'Rutina Z001'
}

CONSEJERIA_CODES = ['99401', '99401.05', '99401.07', '99401.08', '99401.09', '99401.12', '99401.16', '99401.24', '99401.25']

SUPLEMENTACION_NAMES = {
    'U1692': 'Sulfato Ferroso',
    '59430': 'Polimaltosado/Hierro Polimaltosado',
    'U140': 'Otros Suplementos',
    '99403.01': 'Multimicronutrientes',
    '99199.26': 'Suplementaci\u00f3n General',
    '99199.17': 'Suplementaci\u00f3n Hierro Ni\u00f1os',
    '99199.27': 'Vitamina A',
}

TX_ANEMIA_NAMES = {
    '85018': 'Dosaje de Hemoglobina',
    '85018.01': 'Dosaje de Hemoglobina',
    'C0011': 'Tratamiento Anemia/Visita Seguimiento',
}

AGE_BANDS = [
    ('< 6 meses', "EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 0 AND EXTRACT(MONTH FROM age(fecha_atencion::date, fecha_nacimiento::date)) < 6"),
    ('6-11m', "EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 0 AND EXTRACT(MONTH FROM age(fecha_atencion::date, fecha_nacimiento::date)) BETWEEN 6 AND 11"),
    ('12-23m', "EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 1"),
    ('24-35m', "EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 2"),
    ('36-47m', "EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 3"),
    ('48-59m', "EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 4"),
    ('5-9 años', "EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) BETWEEN 5 AND 9"),
    ('10-11 años', "EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) BETWEEN 10 AND 11"),
]

def _resolve_tabla(anio, tipo='cred'):
    if tipo == 'iras_edas' and anio == '2025':
        return 'tabla_iras_edas'
    if anio == '2025':
        return 'tabla_vacunas'
    return f'his_proceso_{anio}'

@app.route('/api/reportes-minsa/tipos', methods=['GET'])
def reportes_minsa_tipos():
    return jsonify([
        {'id': 'formato_nino_cred', 'nombre': 'FORMATO NI�O - CRED (Completo)'},
        {'id': 'formato_iras_edas', 'nombre': 'FORMATO IRAS EDAS'},
    ])

@app.route('/api/reportes-minsa/filtros', methods=['GET'])
def reportes_minsa_filtros():
    conn = _db_cursor()
    cur = conn.cursor()
    esquema = _get_esquema()
    result = {'anios_disponibles': ['2024', '2025', '2026']}
    for tbl in ['his_proceso_2026', 'tabla_vacunas', 'his_proceso_2025']:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {_qtable(esquema, tbl)}")
            if cur.fetchone()[0] > 0:
                src = tbl; break
        except Exception:
            continue
    else:
        src = 'tabla_vacunas'
    def get_distinct(col):
        cur.execute(f"SELECT DISTINCT {_qcol(col)} FROM {_qtable(esquema, src)} WHERE {_qcol(col)} IS NOT NULL AND {_qcol(col)} != '' ORDER BY {_qcol(col)}")
        return [r[0] for r in cur.fetchall()]
    for c in ['red', 'microred', 'provincia', 'distrito', 'nombre_establecimiento']:
        try:
            result[c] = get_distinct(c)
        except Exception:
            result[c] = []
    cur.close()
    conn.close()
    return jsonify(result)

def _build_where(anio, meses, filtros):
    clauses = [f"t.anio::text = '{anio}'"]
    if meses:
        meses_str = ','.join(str(m) for m in meses)
        clauses.append(f"t.mes IN ({meses_str})")
    for col, val in filtros.items():
        clauses.append(f"LOWER(t.{_qcol(col)}) LIKE LOWER('%{val.replace(chr(39), chr(39)+chr(39))}%')")
    return ' AND '.join(clauses)

def _build_age_case():
    return """
    CASE
        WHEN fecha_nacimiento IS NULL THEN 'Sin edad'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 0
            AND EXTRACT(MONTH FROM age(fecha_atencion::date, fecha_nacimiento::date)) < 6 THEN '< 6 meses'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 0
            AND EXTRACT(MONTH FROM age(fecha_atencion::date, fecha_nacimiento::date)) BETWEEN 6 AND 11 THEN '6-11m'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 1 THEN '12-23m'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 2 THEN '24-35m'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 3 THEN '36-47m'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 4 THEN '48-59m'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) BETWEEN 5 AND 9 THEN '5-9 a�os'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) BETWEEN 10 AND 11 THEN '10-11 a�os'
        ELSE 'Otros'
    END
    """

AGE_ORDER = {'< 6 meses':1,'6-11m':2,'12-23m':3,'24-35m':4,'36-47m':5,'48-59m':6,'5-9 a�os':7,'10-11 a�os':8,'Otros':9}

def _build_rn_age_case():
    return """
    CASE
        WHEN fecha_nacimiento IS NULL THEN 'Sin edad'
        WHEN (fecha_atencion::date - fecha_nacimiento::date) BETWEEN 3 AND 6 THEN 'RN 3-6d'
        WHEN (fecha_atencion::date - fecha_nacimiento::date) BETWEEN 7 AND 13 THEN 'RN 7-13d'
        WHEN (fecha_atencion::date - fecha_nacimiento::date) BETWEEN 14 AND 21 THEN 'RN 14-21d'
        WHEN (fecha_atencion::date - fecha_nacimiento::date) >= 22 AND (fecha_atencion::date - fecha_nacimiento::date) < 29 THEN 'RN 22-28d'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 0
            AND EXTRACT(MONTH FROM age(fecha_atencion::date, fecha_nacimiento::date)) < 12 THEN '< 1 a\u00f1o'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 1 THEN '1 a\u00f1o'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 2 THEN '2 a\u00f1os'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 3 THEN '3 a\u00f1os'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) = 4 THEN '4 a\u00f1os'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) BETWEEN 5 AND 9 THEN '5-9 a\u00f1os'
        WHEN EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) BETWEEN 10 AND 11 THEN '10-11 a\u00f1os'
        ELSE 'Otros'
    END
    """

@app.route('/api/reportes-minsa/ejecutar', methods=['POST'])
def reportes_minsa_ejecutar():
    data = request.json or {}
    tipo = data.get('tipo', 'formato_nino_cred')
    anio = str(data.get('anio', '2025'))
    meses = data.get('meses', [])
    filtros = {}
    for k in ['red','microred','nombre_establecimiento','provincia','distrito']:
        v = data.get(k, '').strip()
        if v and v != '(Todas)':
            filtros[k] = v
    conn = _db_cursor()
    cur = conn.cursor()
    esquema = _get_esquema()
    try:
        if tipo == 'formato_nino_cred':
            result = _reporte_completo_cred(cur, esquema, anio, meses, filtros)
        elif tipo == 'formato_iras_edas':
            result = _reporte_iras_edas(cur, esquema, anio, meses, filtros)
        else:
            return jsonify({'error': 'Tipo no v�lido'}), 400
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()

# ============ REPORTE COMPLETO CRED (3 p�ginas) ============
def _build_cred_where(anio, meses, filtros, alias=''):
    """Build WHERE clause for cred2024-style tables."""
    clauses = [f"anio = {anio}"]
    if meses:
        ms = ','.join(str(m) for m in meses)
        clauses.append(f"mes IN ({ms})")
    for col in ['red','microred','nombre_establecimiento','provincia','distrito']:
        val = filtros.get(col, '')
        if val:
            clauses.append(f"LOWER({col}) LIKE LOWER('%{val.replace(chr(39), chr(39)+chr(39))}%')")
    return ' AND '.join(clauses)

def _qcred_sum(cur, table, columns, where):
    """Query SUM of multiple columns from a cred2024-style table."""
    sums = ', '.join(f'COALESCE(SUM("{c}"),0) as "{c}"' for c in columns)
    cur.execute(f'SELECT {sums} FROM {table} WHERE {where}')
    row = cur.fetchone()
    return dict(zip([c for c in columns], [r or 0 for r in row]))

def _qcred_cols(cur, tables, where):
    """Query ALL numeric columns from cred2024 tables."""
    result = {}
    for alias, tname in tables.items():
        cur.execute(f"SELECT column_name FROM information_schema.columns WHERE table_schema='es_ivan' AND table_name='{tname}' AND data_type='bigint' ORDER BY ordinal_position")
        cols = [r[0] for r in cur.fetchall()]
        data = _qcred_sum(cur, f'es_ivan."{tname}"', cols, where)
        for k, v in data.items():
            result[f'{alias}_{k}'] = v
    return result

def _cred_flat(cred, cols):
    """Build flat section from cred data dict."""
    items = []
    total = 0
    for key, label, col in cols:
        v = cred.get(col, 0) or 0
        items.append({'INDICADOR': label, 'TOTAL': v})
        total += v
    return {'columnas': ['INDICADOR', 'TOTAL'], 'filas': items, 'total': total}

def _excel_grid(titulo, headers, row_defs):
    """Create a grid section matching Excel multi-column layout.
    headers: list of column header strings (first is row-label header)
    row_defs: list of dicts with 'label' and column values
    """
    col_count = len(headers)
    rows = []
    totals = {h: 0 for h in headers[1:]}
    for rd in row_defs:
        celdas = [rd.get('label', '')]
        for h in headers[1:]:
            val = rd.get(h, 0) or 0
            celdas.append(val)
            totals[h] = totals.get(h, 0) + val
        rows.append(celdas)
    total_celdas = ['TOTAL'] + [totals[h] for h in headers[1:]]
    return {
        'titulo': titulo,
        'tipo': 'grid',
        'grid': {
            'headers': headers,
            'rows': rows,
            'totales': total_celdas,
            'col_count': col_count
        }
    }


def _side_by_side(titulo, izquierda, derecha):
    return {
        'tipo': 'side_by_side',
        'titulo': titulo,
        'izquierda': izquierda,
        'derecha': derecha,
    }


def _sb_subsecciones(titulo, items_data):
    filas = []
    total = 0
    for label, valor in items_data:
        v = valor if isinstance(valor, (int, float)) else 0
        filas.append({'label': label, 'valor': v})
        total += v
    return {
        'titulo': titulo,
        'columnas': ['INDICADOR', 'TOTAL'],
        'filas': filas,
        'total': total
    }


# ============ HTML TABLE BUILDERS (Excel-matching layout) ============
def _esc(v):
    """Format value for HTML display."""
    if v is None: return ''
    if isinstance(v, (int, float)): return f'{v:,.0f}'
    s = str(v)
    s = s.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
    return s

def _btd(v, extra=''):
    return f'<td style="border:1px solid #999;padding:2px 6px;vertical-align:top;{extra}">{_esc(v)}</td>'

def _bth(v, extra=''):
    return f'<th style="border:1px solid #999;padding:3px 6px;text-align:center;font-weight:bold;background:#f0f0f0;{extra}">{_esc(v)}</th>'

def _section_header(titulo, colspan, extra=''):
    return f'<tr><td colspan="{colspan}" style="border:1px solid #999;padding:4px 6px;font-weight:bold;background:#e8e8e8;{extra}">{_esc(titulo)}</td></tr>'

def _build_sec_table(sec):
    """Build HTML table for a single section (flat or grid)."""
    if not sec: return ''
    h = ''
    if sec.get('tipo') == 'grid' and sec.get('grid'):
        g = sec['grid']
        ncols = len(g['headers'])
        h += f'<table style="border-collapse:collapse;width:100%;font-family:Calibri,sans-serif;font-size:9pt;">'
        h += f'<tr>{"".join(_bth(c) for c in g["headers"])}</tr>'
        for row in g['rows']:
            h += '<tr>'
            for i, v in enumerate(row):
                h += _btd(v, 'text-align:right;' if i > 0 else '')
            h += '</tr>'
        if g.get('totales'):
            h += '<tr>'
            for i, v in enumerate(g['totales']):
                h += _btd(v, 'font-weight:bold;text-align:right;' if i > 0 else 'font-weight:bold;')
            h += '</tr>'
        h += '</table>'
        return h
    cols = sec.get('columnas', ['INDICADOR', 'TOTAL'])
    filas = sec.get('filas', [])
    if not filas:
        return ''
    ncols = len(cols)
    h += f'<table style="border-collapse:collapse;width:100%;font-family:Calibri,sans-serif;font-size:9pt;">'
    h += f'<tr>{"".join(_bth(c) for c in cols)}</tr>'
    for item in filas:
        h += '<tr>'
        for i, c in enumerate(cols):
            val = item.get(c)
            if val is None: val = item.get(c.lower())
            if val is None and i == 0: val = item.get('label', '')
            if val is None and i == 1: val = item.get('valor', 0)
            h += _btd(val, 'text-align:right;' if i > 0 else '')
        h += '</tr>'
    total = sec.get('total')
    if total is not None:
        h += '<tr>'
        h += _btd('TOTAL', 'font-weight:bold;')
        h += _btd(total, 'font-weight:bold;text-align:right;')
        h += '</tr>'
    h += '</table>'
    return h

def _build_page1_html(col_names, filas_cred, totales_main, secciones):
    """Build ONE HTML string for Page 1 (FORMATO NIÑO) matching Excel layout."""
    STBL = 'border-collapse:collapse;width:100%;font-family:Calibri,sans-serif;font-size:9pt;'
    html = f'<table style="{STBL}">'
    # Title row
    html += f'<tr><td colspan="12" style="border:1px solid #999;text-align:center;font-weight:bold;font-size:13pt;padding:6px;">FORMATO NIÑO - CRED (2024)</td></tr>'
    # CRED visits header
    html += '<tr>' + ''.join(_bth(c) for c in col_names) + '</tr>'
    # CRED visits data
    for fila in filas_cred:
        html += '<tr>' + _btd(fila.get('EDADES',''))
        for cn in col_names[1:]:
            html += _btd(fila.get(cn, 0), 'text-align:right;')
        html += '</tr>'
    # Total
    html += '<tr>' + _btd('TOTALES', 'font-weight:bold;')
    for cn in col_names[1:]:
        html += _btd(totales_main.get(cn, 0), 'font-weight:bold;text-align:right;')
    html += '</tr>'
    html += '</table>'

    # Each section as its own table
    for sec in secciones:
        html += f'<table style="{STBL}margin-top:4px;">'
        if sec.get('tipo') == 'side_by_side' and sec.get('izquierda') and sec.get('derecha'):
            html += _section_header(sec['titulo'], 2)
            # Nested table for left/right
            html += '<tr><td colspan="2" style="border:none;padding:0;">'
            html += '<table style="width:100%;"><tr>'
            html += '<td style="width:50%;vertical-align:top;padding-right:4px;border:none;">'
            html += _build_sec_table(sec['izquierda'])
            html += '</td>'
            html += '<td style="width:50%;vertical-align:top;padding-left:4px;border:none;">'
            html += _build_sec_table(sec['derecha'])
            html += '</td>'
            html += '</tr></table>'
            html += '</td></tr>'
        elif sec.get('tipo') == 'grid' and sec.get('grid'):
            g = sec['grid']
            html += _section_header(sec['titulo'], len(g['headers']))
            html += '<tr>' + ''.join(_bth(c) for c in g['headers']) + '</tr>'
            for row in g['rows']:
                html += '<tr>'
                for i, v in enumerate(row):
                    html += _btd(v, 'text-align:right;' if i > 0 else '')
                html += '</tr>'
            if g.get('totales'):
                html += '<tr>'
                for i, v in enumerate(g['totales']):
                    html += _btd(v, 'font-weight:bold;text-align:right;' if i > 0 else 'font-weight:bold;')
                html += '</tr>'
        else:
            # Flat section
            cols = sec.get('columnas', ['INDICADOR', 'TOTAL'])
            html += _section_header(sec['titulo'], len(cols))
            html += '<tr>' + ''.join(_bth(c) for c in cols) + '</tr>'
            for item in sec.get('filas', []):
                html += '<tr>'
                for i, c in enumerate(cols):
                    val = item.get(c)
                    if val is None: val = item.get(c.lower())
                    if val is None and i == 0: val = item.get('label', '')
                    if val is None and i == 1: val = item.get('valor', 0)
                    html += _btd(val, 'text-align:right;' if i > 0 else '')
                html += '</tr>'
            total = sec.get('total')
            if total is not None:
                html += '<tr>' + _btd('TOTAL', 'font-weight:bold;') + _btd(total, 'font-weight:bold;text-align:right;') + '</tr>'
        html += '</table>'
    return html


def _reporte_cred_2024(cur, esquema, anio, meses, filtros):
    """Reporte completo usando tablas pre-agregadas cred2024_* (solo a�o 2024)."""
    where = _build_cred_where(2024, meses, filtros)
    qt = lambda t: f'es_ivan."{t}"'

    # Query ALL bigint columns from all 5 tables
    c24 = _qcred_cols(cur, {'C1':'cred2024'}, where)
    c1 = _qcred_cols(cur, {'C1':'cred2024_1'}, where)
    c2 = _qcred_cols(cur, {'C2':'cred2024_2'}, where)
    c3 = _qcred_cols(cur, {'C3':'cred2024_3'}, where)
    c4 = _qcred_cols(cur, {'C4':'cred2024_4'}, where)

    def v(d, k):
        return d.get(k, 0) or 0
    def sec(titulo, cols, flat=True):
        if flat:
            items = []
            total = 0
            for key, label, col in cols:
                val = v(key, col)
                items.append({'label': label, 'valor': val})
                total += val
            return {'titulo': titulo, 'columnas': ['INDICADOR', 'TOTAL'], 'filas': items, 'total': total}
        else:
            items = []
            for key, label, col in cols:
                val = v(key, col)
                items.append({'label': label, 'valor': val})
            return {'titulo': titulo, 'columnas': ['INDICADOR', 'TOTAL'], 'filas': items, 'total': sum(i['valor'] for i in items)}

    secciones = []

    def _flat_sb(dict_list, vfunc):
        return [(l, vfunc(d, c)) for d, l, c in dict_list]

    # ===== 1. I. ATENCIÓN DEL RECIÉN NACIDO (side-by-side) =====
    atenc_inmed = [
        (c24, '1. Atención Inmediata RN Sano', 'C1_atenc_inmediata_rn_sano'),
        (c24, '2. Atención Inmediata RN Prematuro', 'C1_atenc_inmediata_rn_premat'),
        (c24, '3. Atención Inmediata BPN', 'C1_atencion_inmediata_bpn'),
        (c24, '4. Corte Cordón Umbilical', 'C1_corte_cordon_umbilical_cnv'),
        (c24, '5. Lactancia 1ra Hora', 'C1_lactancia_1ra_hora_cnv'),
        (c24, '6. Contacto Piel a Piel', 'C1_contacto_piel_piel'),
        (c24, '7. Examen Físico RN Normal', 'C1_examen_fisico_rn_normal'),
        (c24, '8. Tamizaje Toma Muestra', 'C1_tamizaje_toma_muestra'),
        (c24, '9. Tamizaje Hipoacusia', 'C1_tamizaje_hipoacusia'),
        (c24, '10. Tamizaje Catarata Congénita', 'C1_tamizaje_catarata_congenita'),
        (c24, '11. Tamizaje Cardiopatía', 'C1_tamizaje_cardiopatia'),
    ]
    aloj_conj = [
        (c4, '1. Atención Alojamiento Conjunto', 'C4_atencion_alojamiento_conjunto'),
        (c4, '2. Atención Inmediata Niño Sano', 'C4_atenc_inmed_nino_sano'),
        (c4, '3. Evaluación Médica RN', 'C4_evaluacion_medica_rn'),
        (c4, '4. BCG 12h', 'C4_bcg_12horas'),
        (c4, '5. BCG 12-24h', 'C4_bcg_12_24h'),
        (c4, '6. BCG 1-11m', 'C4_bcg_1_11m'),
        (c4, '7. HVB 12-24h', 'C4_hvb_12_24h'),
        (c4, '8. HVB 24h', 'C4_hvb_24h'),
    ]
    secciones.append(_side_by_side('I. ATENCIÓN DEL RECIÉN NACIDO',
        _sb_subsecciones('A) Atención Inmediata', _flat_sb(atenc_inmed, v)),
        _sb_subsecciones('C) Atención de Recién Nacido en Alojamiento Conjunto', _flat_sb(aloj_conj, v))))

    # ===== 2. B) Condición de Nacimiento / CONSEJERÍA (side-by-side) =====
    cond_nac = [
        (c4, '1. Extremadamente bajo peso', 'C4_peso_extremadamente_bajo'),
        (c4, '2. Muy bajo peso', 'C4_muy_bajo_peso'),
        (c4, '3. Bajo peso al nacer', 'C4_bajo_peso'),
        (c4, '4. Macrosómico', 'C4_macrosomico'),
        (c4, '5. Microcefalia', 'C4_microcefalia'),
        (c4, '6. Prematuro', 'C4_prematuro'),
        (c4, '7. Post-término', 'C4_post_termino'),
    ]
    cons_brief = [
        (c4, '1. Consejería en atención temprana del desarrollo', 'C4_consej_atc_tempra_desarrollo'),
        (c4, '2. Consejería en inmunizaciones', 'C4_consej_inmunizaciones'),
        (c4, '3. Consejería de identificación de signos de alarma', 'C4_conse_signos_alarma'),
        (c4, '4. Consejería para la prevención de muerte súbita del lactante', 'C4_conse_prev_muerte_subita_lactant'),
        (c4, '5. Consejería para la prevención de enfermedades prevalentes', 'C4_conse_prev_enf_prevalentes_ira_eda'),
        (c4, '6. Consejería en salud ocular', 'C4_conse_salud_ocular'),
        (c4, '7. Consejería en pautas de crianza', 'C4_conse_pautas_crianza'),
        (c4, '8. Consejería nutricional', 'C4_conse_aliment_saludable'),
        (c4, '9. Consejería en Lactancia Materna', 'C4_conse_lme'),
    ]
    secciones.append(_side_by_side('B) Condición de Nacimiento / CONSEJERÍA',
        _sb_subsecciones('B) Condición de Nacimiento', _flat_sb(cond_nac, v)),
        _sb_subsecciones('CONSEJERÍA', _flat_sb(cons_brief, v))))

    # ===== 3. B) Tamizaje Neonatal / E) Atención RN en VD (side-by-side) =====
    tamiz_diag = [
        (c4, '1. Hipotiroidismo Congénito', 'C4_hipotiroidismo_congenito_sin_bocio'),
        (c4, '2. Fenilcetonuria Clásica', 'C4_fenilcetonuria_clasica'),
        (c4, '3. Hiperplasia Suprarrenal Congénita', 'C4_hiperplasia_suprarrenal_congenita'),
        (c4, '4. Cardiopatía Congénita Tipo 1', 'C4_cardiopatia_congenita_tipo1'),
        (c4, '5. Cardiopatía Congénita Tipo 2', 'C4_cardiopatia_congenita_tipo2'),
        (c4, '6. Fibrosis Quística', 'C4_fibrosis_quistica_sin_otra_especificacion'),
        (c4, '7. Catarata Congénita', 'C4_catarata_congenita'),
        (c4, '8. Hipoacusia Conductiva', 'C4_hipoacusia_conductiva'),
    ]
    atenc_vd = [
        (c4, '1. VD Cuidado y Evaluación Neonatal', 'C4_vd_cuidado_y_evaluacion_neonatal'),
        (c4, '2. Anamnesis y Ex. Físico RN Normal', 'C4_anamnesis_y_ex_fisico_rn_normal'),
        (c1, '3. Visita Domiciliaria 1 BPN', 'C1_vis_domic_1_bpn'),
        (c1, '4. Visita Domiciliaria 2 BPN', 'C1_vis_domic_2_bpn'),
        (c1, '5. Visita Domiciliaria 3 BPN', 'C1_vis_domic_3_bpn'),
        (c1, '6. Visita Domiciliaria 6-23m', 'C1_vis_domic1_6_23m'),
        (c1, '7. Visita Domiciliaria 2 6-23m', 'C1_vis_domic2_6_23m'),
        (c1, '8. Visita Domiciliaria 3 6-23m', 'C1_vis_domic3_6_23m'),
        (c1, '9. Visita Domiciliaria 36-59m', 'C1_vis_domic1_36_59m'),
        (c1, '10. Visita Domiciliaria <1m', 'C1_vis_domic_1_men1a'),
        (c1, '11. Visita Domiciliaria 2 <1m', 'C1_vis_domic_2_men1a'),
    ]
    secciones.append(_side_by_side('B) Resultados del Tamizaje Neonatal / E) Atención RN en VD',
        _sb_subsecciones('B) Resultados del Tamizaje Neonatal', _flat_sb(tamiz_diag, v)),
        _sb_subsecciones('E) Atención de Recién Nacido en la Visita Domiciliaria', _flat_sb(atenc_vd, v))))

    # ===== 4. IX. EVALUACIÓN DEL DESARROLLO (grid) =====
    dev_headers = ['Edad', 'Dx Lenguaje', 'Dx Motor', 'Dx Social', 'Dx Coordinación', 'Dx Cognitivo',
                   'Recup Lenguaje', 'Recup Motor', 'Recup Social', 'Recup Coordinación', 'Recup Cognitivo']
    dev_areas = ['len', 'mot', 'soc', 'coo', 'cog']
    area_map = {'len': 'Lenguaje', 'mot': 'Motor', 'soc': 'Social', 'coo': 'Coordinación', 'cog': 'Cognitivo'}
    dev_row_defs = []
    for edad_label, edad_suf in [('< 1 año', 'm1a'), ('01 año', '1a'), ('02 años', '2a')]:
        mapped = {'label': edad_label}
        for area in dev_areas:
            mapped[f'Dx {area_map[area]}'] = v(c1, f'C1_retardo_desarrollo_{area}_{edad_suf}')
            mapped[f'Recup {area_map[area]}'] = v(c1, f'C1_rec_retardo_desarrollo_{area}_{edad_suf}')
        dev_row_defs.append(mapped)
    secciones.append(_excel_grid('IX. EVALUACIÓN DEL DESARROLLO', dev_headers, dev_row_defs))

    # ===== 5. II. SESIONES DE ATENCIÓN TEMPRANA / VI. LACTANCIA MATERNA (side-by-side) =====
    ses_headers = ['Sesiones', 'RN', '< 1 año', '01 año', '02 años', '03 años', 'Total']
    ses_grupos = [('RN', 'rn', c24, 'C1_sesion_est_temprana_rn_'),
        ('< 1 año', 'menor_1a', c1, 'C1_sesion_est_temprana_menor_1a_'),
        ('01 año', '1a', c1, 'C1_sesion_est_temprana_1a_'),
        ('02 años', '2a', c1, 'C1_sesion_est_temprana_2a_'),
        ('03 años', '3a', c1, 'C1_sesion_est_temprana_3a_')]
    ses_row_defs = []
    for si in range(1, 6):
        row = {'label': f'Sesión {si}', 'Total': 0}
        for label, age, d, pref in ses_grupos:
            val = v(d, f'{pref}{si}')
            row[label] = val
            row['Total'] += val
        ses_row_defs.append(row)
    izq_sesiones = _excel_grid('II. SESIONES DE ATENCIÓN TEMPRANA', ses_headers, ses_row_defs)
    der_lactancia = _sb_subsecciones('VI. LACTANCIA MATERNA', _flat_sb([
        (c24, 'LME 1ra Hora', 'C1_lme_1ra_hora'),
        (c24, 'Suspensión LME 6m', 'C1_suspencion_lme_6m'),
        (c4, 'Lactancia 1ra Hora CNV', 'C4_lactancia_1ra_hora_cnv'),
        (c4, 'Contacto Piel a Piel', 'C4_contacto_piel_piel'),
        (c4, 'Suspensión LME 6m', 'C4_suspencion_lme_6m'),
    ], v))
    secciones.append(_side_by_side('II. SESIONES DE ATENCIÓN TEMPRANA / VI. LACTANCIA MATERNA', izq_sesiones, der_lactancia))

    # ===== 6. X. PLAN DE ATENCIÓN INTEGRAL =====
    secciones.append({'titulo': 'X. PLAN DE ATENCIÓN INTEGRAL', **_cred_flat({**c24, **c4}, [
        (c24, 'Plan AIS Inicio RN', 'C1_plan_ais_ini_rn'),
        (c24, 'Plan AIS Término RN', 'C1_plan_ais_ta_rn'),
        (c4, 'Plan AIS Inicio 1m', 'C4_plan_ais_ini_1m'),
        (c4, 'Plan AIS Término 7m', 'C4_plan_ais_termino_7m'),
        (c4, 'Plan AIS Inicio 1a', 'C4_plan_ais_ini_1a'),
        (c4, 'Plan AIS Término 1a', 'C4_plan_ais_termino_1a'),
        (c4, 'Plan AIS Inicio 2a', 'C4_plan_ais_ini_2a'),
        (c4, 'Plan AIS Término 2a', 'C4_plan_ais_termino_2a'),
        (c4, 'Plan AIS Inicio 3a', 'C4_plan_ais_ini_3a'),
        (c4, 'Plan AIS Término 3a', 'C4_plan_ais_termino_3a'),
        (c4, 'Plan AIS Inicio 4a', 'C4_plan_ais_ini_4a'),
        (c4, 'Plan AIS Término 4a', 'C4_plan_ais_termino_4a'),
        (c4, 'Plan AIS Inicio 5-11a', 'C4_plan_ais_ini_5a'),
        (c4, 'Plan AIS Término 5-11a', 'C4_plan_ais_ta_5a'),
    ])})

    # ===== 7. IV. CONSEJERÍA EN LA ATENCIÓN DEL NIÑO(A) (grid) =====
    cons_headers = ['Tipos / Edades', 'Total', 'RN', '<1 año', '01 año', '02 años', '03 años', '04 años',
                    '5 años', '6 años', '7 años', '8 años', '9 años', '10 años', '11 años']
    cons_age_order = ['rn', 'men_1a', '1a', '2a', '3a', '4a', '5a', '6a', '7a', '8a', '9a', '10a', '11a']
    cons_types = [
        'Consejería en atención temprana del desarrollo',
        'Consejería en inmunizaciones',
        'Consejería de identificación de signos de alarma',
        'Consejería para la prevención de muerte súbita del lactante',
        'Consejería para la prevención de enfermedades prevalentes (EDA, IRA, entre otras)',
        'Consejería en salud ocular',
        'Consejería en pautas de crianza, buen trato, comunicación y cuidados adecuados',
        'Consejería nutricional: Alimentación saludable',
        'Consejería en Lactancia Materna Exclusiva hasta los 06 meses',
    ]
    cons_prefixes = {
        'Consejería en atención temprana del desarrollo': 'C4_consej_atc_tempra_desarrollo',
        'Consejería en inmunizaciones': 'C4_consej_inmunizaciones',
        'Consejería de identificación de signos de alarma': 'C4_conse_signos_alarma',
        'Consejería para la prevención de muerte súbita del lactante': 'C4_conse_prev_muerte_subita_lactant',
        'Consejería para la prevención de enfermedades prevalentes (EDA, IRA, entre otras)': 'C4_conse_prev_enf_prevalentes_ira_eda',
        'Consejería en salud ocular': 'C4_conse_salud_ocular',
        'Consejería en pautas de crianza, buen trato, comunicación y cuidados adecuados': 'C4_conse_pautas_crianza',
        'Consejería nutricional: Alimentación saludable': 'C4_conse_aliment_saludable',
        'Consejería en Lactancia Materna Exclusiva hasta los 06 meses': 'C4_conse_lme',
    }
    cons_row_defs = []
    for ct in cons_types:
        pref = cons_prefixes[ct]
        row = {'label': ct, 'Total': 0}
        age_vals = []
        for i, age in enumerate(cons_age_order):
            val = v(c4, f'{pref}_{age}')
            age_vals.append(val)
            row['Total'] += val
        for i, h in enumerate(cons_headers[2:]):
            row[h] = age_vals[i]
        cons_row_defs.append(row)
    secciones.append(_excel_grid('IV. CONSEJERÍA EN LA ATENCIÓN DEL NIÑO(A)', cons_headers, cons_row_defs))

    # ===== 8. V. EVALUACIÓN NUTRICIONAL (grid) =====
    nut_headers = ['Indicador', '< 1 año', '01 año', '02 años', '03 años', '04 años', '5-11 años', 'Total']
    nut_age_suf = ['men1a', '1a', '2a', '3a', '4a', '5_11a']
    nut_age_h = ['< 1 año', '01 año', '02 años', '03 años', '04 años', '5-11 años']
    nut_mapped_rows = []
    for nt, pref in [('Desnutrición Global', 'C1_desnutric_global'), ('Desnutrición Aguda', 'C1_desnutric_aguda'),
        ('Desnutrición Crónica', 'C1_desnutric_cronica'), ('Desnutrición Severa', 'C1_desnutric_severa'),
        ('Obesidad', 'C1_obeso'), ('Sobrepeso', 'C1_sobre_peso')]:
        vals = [v(c1, f'{pref}_{s}') for s in nut_age_suf]
        if not any(vals):
            continue
        row = {'label': nt}
        for i, s in enumerate(nut_age_suf):
            row[nut_age_h[i]] = vals[i]
        row['Total'] = sum(vals)
        nut_mapped_rows.append(row)
    for label, pref, suffixes in [('Gan. Inadec. Peso', 'C1_g_inadecuada_pe', ['men1a']),
        ('Gan. Inadec. Talla', 'C1_g_inadecuada_talla', ['men1a']),
        ('Talla Alta', 'C1_te_alto', ['5_11a'])]:
        row = {'label': label, 'Total': 0}
        for i, s in enumerate(nut_age_suf):
            val = v(c1, f'{pref}_{s}') if s in suffixes else 0
            row[nut_age_h[i]] = val
            row['Total'] += val
        nut_mapped_rows.append(row)
    secciones.append(_excel_grid('V. EVALUACIÓN NUTRICIONAL', nut_headers, nut_mapped_rows))

    # ===== 9. VIII. PARASITOSIS / EXÁMENES DE LABORATORIO =====
    lab_items = [
        (c1, 'Test Graham <1a', 'C1_test_graham_1a'),
        (c1, 'Test Graham 1a', 'C1_test_graham_men1a'),
        (c1, 'Test Graham 2a', 'C1_test_graham_2a'),
        (c1, 'Test Graham 3a', 'C1_test_graham_3a'),
        (c1, 'Test Graham 4a', 'C1_test_graham_4a'),
        (c1, 'Test Graham 5-11a', 'C1_test_graham_5_11a'),
        (c1, 'Test Graham Positivo <1a', 'C1_test_graham_posit_1a'),
        (c1, 'Test Graham Positivo 2a', 'C1_test_graham_posit_2a'),
        (c1, 'Test Graham Positivo 3a', 'C1_test_graham_posit_3a'),
        (c1, 'Test Graham Positivo 4a', 'C1_test_graham_posit_4a'),
        (c1, 'Test Graham Positivo 5-11a', 'C1_test_graham_posit_5_11a'),
        (c1, 'Test Graham Tratado <1a', 'C1_test_graham_tto_1a'),
        (c1, 'Test Graham Tratado 2a', 'C1_test_graham_tto_2a'),
        (c1, 'Test Graham Tratado 3a', 'C1_test_graham_tto_3a'),
        (c1, 'Test Graham Tratado 4a', 'C1_test_graham_tto_4a'),
        (c1, 'Test Graham Tratado 5-11a', 'C1_test_graham_tto_5_11a'),
        (c1, 'Seriado Heces <1a', 'C1_seriado_heces_1a'),
        (c1, 'Seriado Heces 2a', 'C1_seriado_heces_2a'),
        (c1, 'Seriado Heces 3a', 'C1_seriado_heces_3a'),
        (c1, 'Seriado Heces 4a', 'C1_seriado_heces_4a'),
        (c1, 'Seriado Heces 5-11a', 'C1_seriado_heces_5_11a'),
        (c1, 'Seriado Heces Positivo <1a', 'C1_seriado_heces_positivo_1a'),
        (c1, 'Seriado Heces Positivo 2a', 'C1_seriado_heces_positivo_2a'),
        (c1, 'Seriado Heces Positivo 3a', 'C1_seriado_heces_positivo_3a'),
        (c1, 'Seriado Heces Positivo 4a', 'C1_seriado_heces_positivo_4a'),
        (c1, 'Seriado Heces Positivo 5-11a', 'C1_seriado_heces_positivo_5_11a'),
        (c1, 'SH Positivo Tratado <1a', 'C1_sh_positivo_tto_1a'),
        (c1, 'SH Positivo Tratado 2a', 'C1_sh_positivo_tto_2a'),
        (c1, 'SH Positivo Tratado 3a', 'C1_sh_positivo_tto_3a'),
        (c1, 'SH Positivo Tratado 4a', 'C1_sh_positivo_tto_4a'),
        (c1, 'SH Positivo Tratado 5-11a', 'C1_sh_positivo_tto_5_11a'),
    ]
    secciones.append({'titulo': 'VIII. PARASITOSIS / EXÁMENES DE LABORATORIO', **_cred_flat(c1, lab_items)})

    # ===== 10. IX. ADMINISTRACIÓN DE PROFILAXIS ANTIPARASITARIA (grid) =====
    para_headers = ['Dosis', '< 1 año', '02 años', '03 años', '04 años', '5-11 años', 'Total']
    para_ages = {'< 1 año': '1a', '02 años': '2a', '03 años': '3a', '04 años': '4a', '5-11 años': '5_11a'}
    para_row_defs = []
    for dose_num, dose_label in [(1, '1ra Dosis'), (2, '2da Dosis')]:
        row = {'label': dose_label, 'Total': 0}
        for header, suf in para_ages.items():
            val = v(c1, f'C1_antiparasitaria_{dose_num}_{suf}')
            row[header] = val
            row['Total'] += val
        para_row_defs.append(row)
    secciones.append(_excel_grid('IX. ADMINISTRACIÓN DE PROFILAXIS ANTIPARASITARIA', para_headers, para_row_defs))

    # ===== 11. XII. VISITA DOMICILIARIA (grid) =====
    vd_headers = ['Tipo', '< 1 año', '1-5 años', 'Total']
    vd_row_defs = []
    for label, col in [('Visita Domiciliaria 1 BPN', 'C1_vis_domic_1_bpn'),
        ('Visita Domiciliaria 2 BPN', 'C1_vis_domic_2_bpn'),
        ('Visita Domiciliaria 3 BPN', 'C1_vis_domic_3_bpn'),
        ('Visita Domiciliaria 6-23m', 'C1_vis_domic1_6_23m'),
        ('Visita Domiciliaria 2 6-23m', 'C1_vis_domic2_6_23m'),
        ('Visita Domiciliaria 3 6-23m', 'C1_vis_domic3_6_23m'),
        ('Visita Domiciliaria 36-59m', 'C1_vis_domic1_36_59m'),
        ('Visita Domiciliaria <1m', 'C1_vis_domic_1_men1a'),
        ('Visita Domiciliaria 2 <1m', 'C1_vis_domic_2_men1a'),
        ('VD Cuidado y Evaluación Neonatal', 'C4_vd_cuidado_y_evaluacion_neonatal'),
        ('Anamnesis y Ex. Físico RN Normal', 'C4_anamnesis_y_ex_fisico_rn_normal'),
    ]:
        d = c1 if col.startswith('C1_') else c4
        val = v(d, col)
        vd_row_defs.append({'label': label, '< 1 año': val, '1-5 años': 0, 'Total': val})
    secciones.append(_excel_grid('XII. VISITA DOMICILIARIA', vd_headers, vd_row_defs))

    # ===== 12. XV. SALUD MENTAL (Tamizajes) =====
    secciones.append({'titulo': 'XV. SALUD MENTAL', **_cred_flat(c3, [
        (c3, 'Tamizaje Neurodesarrollo 2a', 'C3_tamizaje_nd_2a'),
        (c3, 'Tamizaje Trastorno Mental 6-9a', 'C3_tamizaje_tm_6_9a'),
        (c3, 'Tamizaje Trastorno Mental 10-11a', 'C3_tamizaje_tm_10_11a'),
    ])})

    # ===== 13. TAMIZAJES =====
    secciones.append({'titulo': 'TAMIZAJES', **_cred_flat(c3, [
        (c3, 'Tamizaje Violencia 1-2a', 'C3_tamizaje_viol_1_2a'),
        (c3, 'Tamizaje Violencia 3-5a', 'C3_tamizaje_viol_3_5a'),
        (c3, 'Tamizaje Violencia 6-9a', 'C3_tamizaje_viol_6_9a'),
        (c3, 'Tamizaje Violencia 10-11a', 'C3_tamizaje_viol_10_11a'),
        (c3, 'Tamizaje Alcohol/Drogas 6-9a', 'C3_tamizaje_ad_6_9a'),
        (c3, 'Tamizaje Alcohol/Drogas 10-11a', 'C3_tamizaje_ad_10_11a'),
        (c3, 'Tamizaje Trastorno Depresivo 10-11a', 'C3_tamizaje_td_10_11a'),
    ])})

    # ===== 14. TAMIZAJES POSITIVOS =====
    secciones.append({'titulo': 'TAMIZAJES POSITIVOS', **_cred_flat(c3, [
        (c3, 'Tamizaje Violencia POSITIVO 1-2a', 'C3_tamizaje_viol_posit_1_2a'),
        (c3, 'Tamizaje Violencia POSITIVO 3-5a', 'C3_tamizaje_viol_posit_3_5a'),
        (c3, 'Tamizaje Violencia POSITIVO 6-9a', 'C3_tamizaje_viol_posit_6_9a'),
        (c3, 'Tamizaje Violencia POSITIVO 10-11a', 'C3_tamizaje_viol_posit_10_11a'),
        (c3, 'Tamizaje Alcohol POSITIVO 10-11a', 'C3_tamizaje_ad_alcohol_posit_10_11a'),
        (c3, 'Tamizaje Tabaco POSITIVO 10-11a', 'C3_tamizaje_ad_tabaco_posit_10_11a'),
        (c3, 'Tamizaje Drogas POSITIVO 10-11a', 'C3_tamizaje_ad_drogas_posit_10_11a'),
        (c3, 'Tamizaje TD POSITIVO 10-11a', 'C3_tamizaje_td_posit_10_11a'),
    ])})

    # ===== 15. XV. RETINOPATÍA DE LA PREMATURIDAD - ROP =====
    secciones.append({'titulo': 'XV. RETINOPATÍA DE LA PREMATURIDAD - ROP', **_cred_flat(c3, [
        (c3, 'Tamizaje RN Factores Riesgo Sano', 'C3_o_tamrn_fr_s_0_29d'),
        (c3, 'Tamizaje RN Factores Riesgo Normal', 'C3_o_tamrn_fr_n_0_29d'),
        (c3, 'Tamizaje RN Factores Riesgo Referido', 'C3_o_tamrn_fr_r_0_29d'),
        (c3, 'Dx Retinopatía Prematuros RN', 'C3_o_dx_retinoprema_c_0_29d'),
        (c3, 'Dx Retinopatía Prematuros 6m', 'C3_o_dx_retinoprema_c_6m'),
        (c3, 'Dx Retinopatía Prematuros <3a', 'C3_o_dx_retinoprema_c_1_3a'),
        (c3, 'Tto Láser Retinopatía', 'C3_o_tto_retinoprema_ct_l_0_29d'),
        (c3, 'Tto Antiangiogénico Retinopatía', 'C3_o_tto_retinoprema_ct_i_0_29d'),
        (c3, 'Tto Láser+Antiangiogénico', 'C3_o_tto_retinoprema_ct_lm_0_29d'),
    ])})

    # ===== 16. ATENCIÓN DE SALUD OCULAR / AGUDEZA VISUAL =====
    secciones.append({'titulo': 'ATENCIÓN DE SALUD OCULAR - AGUDEZA VISUAL', **_cred_flat(c3, [
        (c3, 'Determinación Agudeza Visual 3-4a', 'C3_o_determ_agudeza_visual_3_4a'),
        (c3, 'Determinación Agudeza Visual 5-7a', 'C3_o_determ_agudeza_visual_5_7a'),
        (c3, 'Determinación Agudeza Visual 8-11a', 'C3_o_determ_agudeza_visual_8_11a'),
        (c3, 'Evaluación Agudeza Visual', 'C3_o_determ_agudeza_visual_eva_total'),
        (c3, 'Referencia Agudeza Visual', 'C3_o_determ_agudeza_visual_ref_total'),
        (c3, 'Examen Ojos/Visión Normal', 'C3_o_ex_ojo_vis_n_0_5a_total'),
        (c3, 'Examen Ojos/Visión Anormal', 'C3_o_ex_ojo_vis_a_0_5a_total'),
        (c3, 'Examen Ojos/Visión Ref.', 'C3_o_ex_ojo_vis_rf_0_5a_total'),
    ])})

    # ===== 17. ERRORES DE REFRACCIÓN =====
    secciones.append({'titulo': 'ERRORES DE REFRACCIÓN', **_cred_flat(c3, [
        (c3, 'Dx Hipermetropía 3-4a', 'C3_o_dx_errr_hip_3_4a'),
        (c3, 'Dx Miopía 3-4a', 'C3_o_dx_errr_mio_3_4a'),
        (c3, 'Dx Astigmatismo 3-4a', 'C3_o_dx_errr_ast_3_4a'),
        (c3, 'Dx Anisometropía 3-4a', 'C3_o_dx_errr_ani_3_4a'),
        (c3, 'Tto/Provisión Anteojos 3-4a', 'C3_o_tto_prov_anteo_3_4a'),
        (c3, 'Dx Hipermetropía 5-7a', 'C3_o_dx_errr_hip_5_7a'),
        (c3, 'Dx Miopía 5-7a', 'C3_o_dx_errr_mio_5_7a'),
        (c3, 'Dx Astigmatismo 5-7a', 'C3_o_dx_errr_ast_5_7a'),
        (c3, 'Dx Anisometropía 5-7a', 'C3_o_dx_errr_ani_5_7a'),
        (c3, 'Tto/Provisión Anteojos 5-7a', 'C3_o_tto_prov_anteo_5_7a'),
        (c3, 'Dx Hipermetropía 8-11a', 'C3_o_dx_errr_hip_8_11a'),
        (c3, 'Dx Miopía 8-11a', 'C3_o_dx_errr_mio_8_11a'),
        (c3, 'Dx Astigmatismo 8-11a', 'C3_o_dx_errr_ast_8_11a'),
        (c3, 'Dx Anisometropía 8-11a', 'C3_o_dx_errr_ani_8_11a'),
        (c3, 'Tto/Provisión Anteojos 8-11a', 'C3_o_tto_prov_anteo_8_11a'),
    ])})

    # ===== 18. ATENCIÓN EN SALUD BUCAL =====
    secciones.append({'titulo': 'ATENCIÓN EN SALUD BUCAL', **_cred_flat(c3, [
        (c3, 'Instrucción Higiene Oral', 'C3_b_iho_i_0_28d'),
        (c3, 'Cepillado Dental 5-11a', 'C3_b_iho_5_11a'),
        (c3, 'Asesoría Nutricional Control Enf. Dentales', 'C3_b_anced_i_0_28d'),
        (c3, 'Aplicación Flúor Barniz', 'C3_b_abf_i_0_28d'),
        (c3, 'Profilaxis Dental', 'C3_b_pd_i_0_28d'),
        (c3, 'Aplicación Sellantes', 'C3_b_aseln_ct_2a'),
    ])})

    # ===== 19. SEGUIMIENTO =====
    seg_items = [
        (c1, 'Seguimiento Control CRED', 'C1_seguim_control_cred_1a'),
        (c1, 'Seguim. Prob. Nutricionales RN', 'C1_seguim_problemas_nutric_rn'),
        (c1, 'Seguim. Prob. Nutricionales <1a', 'C1_seguim_problemas_nutric_men_1a'),
        (c1, 'Seguim. Prob. Nutricionales 1a', 'C1_seguim_problemas_nutric_1a'),
        (c1, 'Seguim. Prob. Nutricionales 2a', 'C1_seguim_problemas_nutric_2a'),
        (c1, 'Seguim. Prob. Nutricionales 3a', 'C1_seguim_problemas_nutric_3a'),
        (c1, 'Seguim. Prob. Nutricionales 4a', 'C1_seguim_problemas_nutric_4a'),
        (c1, 'Seguim. Prob. Nutricionales 5-11a', 'C1_seguim_problemas_nutric_5_11a'),
        (c1, 'Seguim. Prob. Desarrollo RN', 'C1_seguim_problemas_desarr_rn'),
        (c1, 'Seguim. Prob. Desarrollo <1a', 'C1_seguim_problemas_desarr_men_1a'),
        (c1, 'Seguim. Prob. Desarrollo 1a', 'C1_seguim_problemas_desarr_1a'),
        (c1, 'Seguim. Prob. Desarrollo 2a', 'C1_seguim_problemas_desarr_2a'),
        (c1, 'Seguim. Prob. Desarrollo 3a', 'C1_seguim_problemas_desarr_3a'),
        (c1, 'Seguim. Prob. Desarrollo 4a', 'C1_seguim_problemas_desarr_4a'),
        (c1, 'Seguim. Prob. Desarrollo 5-11a', 'C1_seguim_problemas_desarr_5_11a'),
    ]
    secciones.append({'titulo': 'SEGUIMIENTO - VISITA DOMICILIARIA', **_cred_flat({**c1, **c24}, seg_items)})

    # ---- PAGE 1: FORMATO NIÑO ----
    # CRED visits main table from cred2024 columns
    cred_row_names = [
        ('Reci�n Nacido Sano 3-6d', ['C1_cred1_rn_3_6d','C1_cred2_rn_7_13d','C1_cred3_rn_14_21d','C1_cred4_rn_22_a_mas_dias']),
        ('< 1 a�o', ['C1_cred1_29_59d_men1a','C1_cred2_60_89d_men1a','C1_cred3_90_119d_men1a','C1_cred4_120_149d_men1a','C1_cred5_180_209d_men1a','C1_cred6_210_239d_men1a','C1_cred7_270_299d_men1a','C1_cred8_men1a','C1_cred9_men1a','C1_cred10_men1a','C1_cred11_men1a']),
        ('1 a�o', ['C1_cred1_360_389d_1a','C1_cred2_450_479d_1a','C1_cred3_540_569d_1a','C1_cred4_630_659d_1a','C1_cred5_1a','C1_cred6_1a']),
        ('2 a�os', ['C1_cred1_2a','C1_cred2_2a','C1_cred3_2a','C1_cred4_2a']),
        ('3 a�os', ['C1_cred1_3a','C1_cred2_3a','C1_cred3_3a','C1_cred4_3a']),
        ('4 a�os', ['C1_cred1_4a','C1_cred2_4a','C1_cred3_4a','C1_cred4_4a']),
        ('5-9 a�os', ['C1_cred1_5a','C1_cred1_6a','C1_cred1_7a','C1_cred1_8a','C1_cred1_9a']),
        ('10-11 a�os', ['C1_cred1_10a','C1_cred1_11a']),
    ]
    col_names = ['EDADES','1ER.CONTROL','2DO.CONTROL','3ER.CONTROL','4TO.CONTROL',
                 '5TO.CONTROL','6TO.CONTROL','7MO.CONTROL','8VO.CONTROL',
                 'TOTAL ATENCIONES','TOTAL NI�OS ATENDIDOS']
    totales_main = {cn:0 for cn in col_names[1:]}
    filas_cred = []
    for grupo, cols in cred_row_names:
        total = sum(v({**c24, **c1}, col) for col in cols)
        fila = {'EDADES': grupo}
        # Distribute total across columns (we don't have per-control breakdown in cred2024)
        fila['1ER.CONTROL'] = total
        for cn in col_names[2:]:
            fila[cn] = 0
        fila['TOTAL ATENCIONES'] = total
        fila['TOTAL NI�OS ATENDIDOS'] = total
        for cn in col_names[1:]:
            totales_main[cn] += fila.get(cn, 0)
        filas_cred.append(fila)

    data_composite = {}
    for d in [c24, c1, c2, c3, c4]:
        data_composite.update({k.replace('C1_','').replace('C2_','').replace('C3_','').replace('C4_',''): v for k, v in d.items()})
    tabla_html = build_page1_html(col_names, filas_cred, totales_main, secciones,
                                  c24, c1, c2, c3, c4,
                                  data_composite=data_composite,
                                  filtros=filtros)
    pages = [{
        'id': 'formato_nino', 'titulo': 'FORMATO NI�O - CRED',
        'tabla_html': tabla_html,
        # Keep original data for Excel export
        'columnas': col_names, 'filas': filas_cred, 'totales': totales_main,
        'secciones': secciones
    }]

    # ---- PAGE 2: SUPLEMENTACION ----
    sup_secciones = []

    # 1. Suplementaci�n <6m (RN BPN/Prematuro + 4-5m sano)
    sup_secciones.append(sec('1. Sup. Preventiva <6m', [
        (c24, 'RN Bajo Peso 1ra', 'C1_ta_suplem_bpn'),
        (c24, 'RN Prematuro', 'C1_ta_suplem_bpn'),
        (c24, '4m Sano', 'C1_suplem_4m_sano'),
        (c24, '5m Sano', 'C1_suplem_5m_sano'),
        (c24, 'TA 5-6m Sano', 'C1_ta_suplem_5_6m_sano'),
    ]))

    # 2-3. Multimicronutriente 6-59m + Sulfato Ferroso
    sup_edades = [('6-11m', '6_11m'), ('1a', '1a'), ('2a', '2a'), ('3a', '3a'), ('4a', '4a'), ('5-11a', '5_11a')]
    for grupo, suf in sup_edades:
        sup_secciones.append(sec(f'Multimicronutrientes {grupo}', [
            (c24, f'1ra Entrega', f'C1_suplem_1ra_{suf}'),
            (c24, f'2da Entrega', f'C1_suplem_2da_{suf}'),
            (c24, f'3ra Entrega', f'C1_suplem_3ra_{suf}'),
            (c24, f'4ta Entrega', f'C1_suplem_4ta_{suf}'),
            (c24, f'5ta Entrega', f'C1_suplem_5ta_{suf}'),
            (c24, f'6ta Entrega', f'C1_suplem_6ta_{suf}'),
            (c24, f'TA', f'C1_ta_suplem_{suf}'),
        ]))

    # Vitamina A
    sup_secciones.append(sec('4. Vitamina A', [
        (c24, 'VA1 6-11m', 'C1_va1_6_11m'),
        (c24, 'VA1 1a', 'C1_va1_1a'),
        (c24, 'VA2 1a', 'C1_va2_1a'),
        (c24, 'VA1 2a', 'C1_va1_2a'),
        (c24, 'VA2 2a', 'C1_va2_2a'),
        (c24, 'VA1 3a', 'C1_va1_3a'),
        (c24, 'VA2 3a', 'C1_va2_3a'),
        (c24, 'VA1 4a', 'C1_va1_4a'),
        (c24, 'VA2 4a', 'C1_va2_4a'),
    ]))

    # Suplementaci�n Gestantes, Pu�rperas, MEF
    sup_secciones.append(sec('9. Sup. Hierro MEF/Gestantes/Puerperas', [
        (c24, 'Suplem Gestante 1', 'C1_suplem_gest1'),
        (c24, 'Suplem Gestante 2', 'C1_suplem_gest2'),
        (c24, 'Suplem Gestante 3', 'C1_suplem_gest3'),
        (c24, 'Suplem Gestante 4', 'C1_suplem_gest4'),
        (c24, 'Suplem Gestante 5', 'C1_suplem_gest5'),
        (c24, 'Suplem Gestante 6', 'C1_suplem_gest6'),
        (c24, 'Suplem Gestante TA', 'C1_suplem_gest_ta'),
        (c24, 'Suplem Pu�rpera 1', 'C1_suplem_puer1'),
        (c24, 'Suplem Pu�rpera 2', 'C1_suplem_puer2'),
        (c24, 'Suplem Pu�rpera 3', 'C1_suplem_puer3'),
        (c24, 'Suplem Pu�rpera 4', 'C1_suplem_puer4'),
        (c24, 'Suplem Pu�rpera 5', 'C1_suplem_puer5'),
        (c24, 'Suplem Pu�rpera 6', 'C1_suplem_puer6'),
        (c24, 'Suplem Pu�rpera 7', 'C1_suplem_puer7'),
        (c24, 'Suplem Pu�rpera TA', 'C1_suple_puer_ta'),
        (c24, 'Suplem MEF 1', 'C1_suplem_mef1'),
        (c24, 'Suplem MEF 2', 'C1_suplem_mef2'),
        (c24, 'Suplem MEF 3', 'C1_suplem_mef3'),
        (c24, 'Suplem MEF TA', 'C1_suplem_mef_ta'),
        (c24, 'Suplem Adol. 12-17 1', 'C1_suplem1_12_17'),
        (c24, 'Suplem Adol. 12-17 2', 'C1_suplem2_12_17'),
        (c24, 'Suplem Adol. 12-17 3', 'C1_suplem3_12_17'),
        (c24, 'Suplem Adol. 12-17 TA', 'C1_suplem4_12_17ta'),
    ]))

    # 6. Dosaje HB
    sup_secciones.append(sec('6. DOSAJE DE HEMOGLOBINA', [
        (c2, '1er Dosaje BPN', 'C2_primer_dosaje_hb_bpn'),
        (c2, '2do Dosaje BPN', 'C2_segundo_dosaje_hb_bpn'),
        (c2, 'HB 6-11m 1er', 'C2_hb_6_11m_primer'),
        (c2, 'HB 6-11m 2do', 'C2_hb_6_11m_segundo'),
        (c2, 'HB 12-23m 1er', 'C2_hb_12_23m_primer'),
        (c2, 'HB 12-23m 2do', 'C2_hb_12_23m_segundo'),
        (c2, 'HB 12-23m 3er', 'C2_hb_12_23m_tercer'),
        (c2, 'HB 24-35m 1er', 'C2_hb_24_35m_primer'),
        (c2, 'HB 24-35m 2do', 'C2_hb_24_35m_segundo'),
        (c2, 'HB 36-59m 1er', 'C2_hb_36_59m_primer'),
        (c2, 'HB 36-59m 2do', 'C2_hb_36_59m_segundo'),
        (c2, 'HB 5-11a', 'C2_hb_5_11a'),
        (c2, 'HB Adolescente 1er', 'C2_hb_adolescente_primer'),
        (c2, 'HB Adolescente 2do', 'C2_hb_adolescente_segundo'),
        (c2, 'HB Gestante 1er', 'C2_hb_gestante_primer'),
        (c2, 'HB Gestante 2do', 'C2_hb_gestante_segundo'),
        (c2, 'HB Gestante 3er', 'C2_hb_gestante_tercero'),
        (c2, 'HB MEF', 'C2_hb_mef'),
        (c2, 'HB Pu�rpera', 'C2_hb_puerpera'),
    ]))

    # 7. Consulta Nutricional
    con_items = [
        (c4, 'Consulta Nutricional BPN', 'C4_consulta_nutric_bpn'),
        (c4, 'Consulta Nutricional 4-5m', 'C4_consulta_nutric1_4_5m'),
        (c4, 'Consulta Nutricional 6-11m', 'C4_consulta_nutric_1_6_11m'),
        (c4, 'Consulta Nutricional 1a', 'C4_consulta_nutric_1_1a'),
        (c4, 'Consulta Nutricional 2a', 'C4_consulta_nutric_1_2a'),
        (c4, 'Consulta Nutricional 3a', 'C4_consulta_nutric_1_3a'),
        (c4, 'Consulta Nutricional 4a', 'C4_consulta_nutric_1_4a'),
        (c4, 'Consulta Nutricional 5-11a', 'C4_consulta_nutric_1_5_11a'),
        (c24, 'Cons. Nutr. Gestante 1', 'C1_cons_nutric1_gest'),
        (c24, 'Cons. Nutr. Gestante 2', 'C1_cons_nutric2_gest'),
        (c24, 'Cons. Nutr. Gestante 3', 'C1_cons_nutric3_gest'),
        (c24, 'Cons. Nutr. Gestante 4', 'C1_cons_nutric4_gest'),
        (c24, 'Cons. Nutr. Gestante 5', 'C1_cons_nutric5_gest'),
        (c24, 'Cons. Nutr. Gestante 6', 'C1_cons_nutric6_gest'),
        (c24, 'Cons. Nutr. Pu�rpera 1', 'C1_consulta_nutric1_gest'),
        (c24, 'Cons. Nutr. MEF 1', 'C1_cons_mef1'),
        (c24, 'Cons. Nutr. MEF 2', 'C1_cons_mef2'),
        (c24, 'Cons. Nutr. Adolescente 1', 'C1_cons_nutri1_12_17'),
        (c24, 'Cons. Nutr. Adolescente 2', 'C1_cons_nutri2_12_17'),
        (c24, 'Cons. Nutr. Adolescente 3', 'C1_cons_nutri3_12_17'),
    ]
    sup_secciones.append(sec('7. CONSULTA NUTRICIONAL', con_items))

    # 5. Consejer�a Nutricional (from cred2024_4)
    cn_items = []
    for suf in ['bpn', '4_5m', '1_6_11m', '2_6_11m', '3_6_11m', '4_6_11m',
                '1_1a', '2_1a', '3_1a', '4_1a', '5_1a', '6_1a',
                '1_2a', '2_2a', '3_2a', '4_2a', '5_2a', '6_2a',
                '1_3a', '2_3a', '3_3a', '4_3a',
                '1_4a', '2_4a', '3_4a', '4_4a',
                '1_5_11a', '2_5_11a', '3_5_11a', '4_5_11a', '5_5_11a', '6_5_11a']:
        col = f'C1_cons_nutric_{suf}'
        label = f'Cons. Nutr. {suf}'
        if v(c24, col):
            cn_items.append((c24, label, col))
    if cn_items:
        sup_secciones.append(sec('5. CONSEJER\u00cdA NUTRICIONAL', cn_items))
    sup_sec = sup_secciones if sup_secciones else [{'titulo':'Sin datos','columnas':['INDICADOR','TOTAL'],'filas':[],'total':0}]
    pages.append({
        'id': 'suplementacion', 'titulo': 'SUPLEMENTACION PREVENTIVA',
        'tabla_html': build_suplementacion_html(sup_sec),
        'secciones': sup_sec
    })

    # ---- PAGE 3: Tx. ANEMIA ----
    ane_secciones = []

    # 1. Tratamiento Sulfato Ferroso
    ane_edades = ['6_11m', '1a', '2a', '3a', '4a']
    for suf in ane_edades:
        ane_secciones.append(sec(f'Adm. Tx. Sulfato Ferroso {suf}', [
            (c24, f'Suplem 1ra {suf}', f'C1_suplem_1ra_{suf}'),
            (c24, f'Suplem 2da {suf}', f'C1_suplem_2da_{suf}'),
            (c24, f'Suplem 3ra {suf}', f'C1_suplem_3ra_{suf}'),
            (c24, f'Suplem 4ta {suf}', f'C1_suplem_4ta_{suf}'),
            (c24, f'Suplem 5ta {suf}', f'C1_suplem_5ta_{suf}'),
            (c24, f'Suplem 6ta {suf}', f'C1_suplem_6ta_{suf}'),
            (c24, f'TA', f'C1_ta_suplem_{suf}'),
        ]))

    # 2. Dosaje Hemoglobina Control
    ane_secciones.append(sec('2. DOSAJE HEMOGLOBINA CONTROL', [
        (c2, 'HB BPN 1er', 'C2_primer_dosaje_hb_bpn'),
        (c2, 'HB BPN 2do', 'C2_segundo_dosaje_hb_bpn'),
        (c2, 'HB 6-11m 1er', 'C2_hb_6_11m_primer'),
        (c2, 'HB 6-11m 2do', 'C2_hb_6_11m_segundo'),
        (c2, 'HB 12-23m 1er', 'C2_hb_12_23m_primer'),
        (c2, 'HB 12-23m 2do', 'C2_hb_12_23m_segundo'),
        (c2, 'HB 12-23m 3er', 'C2_hb_12_23m_tercer'),
        (c2, 'HB 24-35m 1er', 'C2_hb_24_35m_primer'),
        (c2, 'HB 24-35m 2do', 'C2_hb_24_35m_segundo'),
        (c2, 'HB 36-59m 1er', 'C2_hb_36_59m_primer'),
        (c2, 'HB 36-59m 2do', 'C2_hb_36_59m_segundo'),
        (c2, 'HB 5-11a', 'C2_hb_5_11a'),
        (c2, 'HB Adolescente 1er', 'C2_hb_adolescente_primer'),
        (c2, 'HB Adolescente 2do', 'C2_hb_adolescente_segundo'),
        (c2, 'HB Gestante 1er', 'C2_hb_gestante_primer'),
        (c2, 'HB Gestante 2do', 'C2_hb_gestante_segundo'),
        (c2, 'HB Gestante 3er', 'C2_hb_gestante_tercero'),
        (c2, 'HB MEF', 'C2_hb_mef'),
        (c2, 'HB Pu�rpera', 'C2_hb_puerpera'),
    ]))

    # 3. Consulta M�dica
    ane_secciones.append(sec('3. CONSULTA M�DICA', [
        (c4, 'Anamnesis y Ex. F�sico RN', 'C4_anamnesis_y_ex_fisico_rn_normal'),
        (c4, 'Evaluaci�n M�dica RN', 'C4_evaluacion_medica_rn'),
    ]))

    # 4. Consulta Nutricional (Tx Anemia)
    ane_secciones.append(sec('4. CONSULTA NUTRICIONAL (Tx Anemia)', [
        (c4, 'Consulta Nutricional BPN', 'C4_consulta_nutric_bpn'),
        (c4, 'Consulta Nutricional 4-5m', 'C4_consulta_nutric1_4_5m'),
        (c4, 'Consulta Nutricional 1', 'C4_consulta_nutric_1_1a'),
        (c4, 'Consulta Nutricional 6-11m 1', 'C4_consulta_nutric_1_6_11m'),
        (c4, 'Consulta Nutricional 2a 1', 'C4_consulta_nutric_1_2a'),
        (c4, 'Consulta Nutricional 3a 1', 'C4_consulta_nutric_1_3a'),
        (c4, 'Consulta Nutricional 4a 1', 'C4_consulta_nutric_1_4a'),
        (c4, 'Consulta Nutricional 5-11a 1', 'C4_consulta_nutric_1_5_11a'),
    ]))
    ane_sec = ane_secciones if ane_secciones else [{'titulo':'Sin datos','columnas':['INDICADOR','TOTAL'],'filas':[],'total':0}]
    pages.append({
        'id': 'tx_anemia', 'titulo': 'TRATAMIENTO DE ANEMIA',
        'tabla_html': build_tx_anemia_html(ane_sec),
        'secciones': ane_sec
    })

    return {
        'tipo': 'formato_nino_cred', 'anio': anio, 'filtros': filtros, 'meses': meses,
        'paginas': pages
    }

# Old reporte for non-2024 years (using his_proceso tables)
def _reporte_completo_cred_old(cur, esquema, anio, meses, filtros):
    tabla = _resolve_tabla(anio, 'cred')
    where = _build_where(anio, meses, filtros)
    qt = lambda t: _qtable(esquema, t)
    ace = _build_age_case()
    rn_ace = _build_rn_age_case()

    AGE_ORDER_LOCAL = {'< 6 meses':1,'6-11m':2,'12-23m':3,'24-35m':4,'36-47m':5,'48-59m':6,'5-9 a\u00f1os':7,'10-11 a\u00f1os':8,'Otros':9}
    RN_AGE_ORDER = {'RN 3-6d':1,'RN 7-13d':2,'RN 14-21d':3,'RN 22-28d':4,'< 1 a\u00f1o':5,'1 a\u00f1o':6,'2 a\u00f1os':7,'3 a\u00f1os':8,'4 a\u00f1os':9,'5-9 a\u00f1os':10,'10-11 a\u00f1os':11,'Sin edad':0,'Otros':12}

    pages = []

    def _qflat(codes, name_map, age_case_override=None, age_order_override=None):
        ac = age_case_override or ace
        ao = age_order_override or AGE_ORDER_LOCAL
        cs = "','".join(codes)
        cur.execute(f"""
            SELECT codigo_item, {ac} as grupo_edad, COUNT(*) as total
            FROM {qt(tabla)} t
            WHERE {where} AND codigo_item IN ('{cs}')
            GROUP BY codigo_item, grupo_edad
            ORDER BY codigo_item, grupo_edad
        """)
        rows = cur.fetchall()
        agg = {}
        all_ages = []
        seen_ages = set()
        for code, ge, tot in rows:
            if ge not in seen_ages:
                seen_ages.add(ge)
                all_ages.append(ge)
            if code not in agg:
                agg[code] = {}
            agg[code][ge] = agg[code].get(ge, 0) + tot
        all_ages.sort(key=lambda x: ao.get(x, 99))
        filas = []
        totales = {a:0 for a in all_ages}
        for code in codes:
            name = name_map.get(code, code)
            if code not in agg:
                continue
            row = {'Item': name}
            total = 0
            for a in all_ages:
                v = agg[code].get(a, 0)
                row[a] = v
                totales[a] = totales.get(a, 0) + v
                total += v
            row['Total'] = total
            filas.append(row)
        totales['Total'] = sum(totales.values())
        columnas = ['Item'] + all_ages + ['Total']
        return {'columnas': columnas, 'filas': filas, 'totales': totales}

    def _qval(codes, valor_labels, label_map, age_case_override=None, age_order_override=None):
        ac = age_case_override or ace
        ao = age_order_override or AGE_ORDER_LOCAL
        cs = "','".join(codes)
        vs = "','".join(valor_labels)
        cur.execute(f"""
            SELECT {ac} as grupo_edad, valor_lab, COUNT(*) as total
            FROM {qt(tabla)} t
            WHERE {where} AND codigo_item IN ('{cs}') AND valor_lab IN ('{vs}')
            GROUP BY grupo_edad, valor_lab
            ORDER BY grupo_edad, valor_lab
        """)
        rows = cur.fetchall()
        agg = {}
        for ge, vl, tot in rows:
            if ge not in agg:
                agg[ge] = {}
            agg[ge][vl] = agg[ge].get(vl, 0) + tot
        col_labels = [label_map.get(vl, vl) for vl in valor_labels]
        filas = []
        totales = {l:0 for l in col_labels}
        for ge in sorted(agg.keys(), key=lambda x: ao.get(x, 99)):
            row = {'Grupo Edad': ge}
            total = 0
            for vl in valor_labels:
                lb = label_map.get(vl, vl)
                v = agg[ge].get(vl, 0)
                row[lb] = v
                totales[lb] = totales.get(lb, 0) + v
                total += v
            row['Total'] = total
            filas.append(row)
        totales['Total'] = sum(totales.values())
        columnas = ['Grupo Edad'] + col_labels + ['Total']
        return {'columnas': columnas, 'filas': filas, 'totales': totales}

    # ---- PAGE 1: FORMATO NI\u00d1O ----
    cn = "','".join(CRED_CODES)
    # age_case_cte: uses CTE aliases (edad_dias/edad_anios/edad_meses) - for main query
    age_case_cte = """
        CASE
            WHEN edad_dias BETWEEN 3 AND 6 THEN 'Reci\u00e9n Nacido Sano 3-6d'
            WHEN edad_dias BETWEEN 7 AND 13 THEN 'Reci\u00e9n Nacido Sano 7-13d'
            WHEN edad_dias BETWEEN 14 AND 21 THEN 'Reci\u00e9n Nacido Sano 14-21d'
            WHEN edad_dias >= 22 AND edad_dias < 29 THEN 'Reci\u00e9n Nacido Sano may_22d'
            WHEN edad_anios = 0 AND edad_meses < 12 THEN '< 1 a\u00f1o'
            WHEN edad_anios = 1 THEN '1 a\u00f1o'
            WHEN edad_anios = 2 THEN '2 a\u00f1os'
            WHEN edad_anios = 3 THEN '3 a\u00f1os'
            WHEN edad_anios = 4 THEN '4 a\u00f1os'
            WHEN edad_anios BETWEEN 5 AND 9 THEN '5-9 a\u00f1os'
            WHEN edad_anios BETWEEN 10 AND 11 THEN '10-11 a\u00f1os'
            ELSE 'Otros'
        END"""
    # age_case_direct: uses raw column computation - for _qmonth_pivot (no CTE)
    _ed = '(fecha_atencion::date - fecha_nacimiento::date)'
    _ea = 'EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date))'
    _em = 'EXTRACT(MONTH FROM age(fecha_atencion::date, fecha_nacimiento::date))'
    age_case_direct = f"""
        CASE
            WHEN {_ed} BETWEEN 3 AND 6 THEN 'Reci\u00e9n Nacido Sano 3-6d'
            WHEN {_ed} BETWEEN 7 AND 13 THEN 'Reci\u00e9n Nacido Sano 7-13d'
            WHEN {_ed} BETWEEN 14 AND 21 THEN 'Reci\u00e9n Nacido Sano 14-21d'
            WHEN {_ed} >= 22 AND {_ed} < 29 THEN 'Reci\u00e9n Nacido Sano may_22d'
            WHEN {_ea} = 0 AND {_em} < 12 THEN '< 1 a\u00f1o'
            WHEN {_ea} = 1 THEN '1 a\u00f1o'
            WHEN {_ea} = 2 THEN '2 a\u00f1os'
            WHEN {_ea} = 3 THEN '3 a\u00f1os'
            WHEN {_ea} = 4 THEN '4 a\u00f1os'
            WHEN {_ea} BETWEEN 5 AND 9 THEN '5-9 a\u00f1os'
            WHEN {_ea} BETWEEN 10 AND 11 THEN '10-11 a\u00f1os'
            ELSE 'Otros'
        END"""
    age_case = age_case_cte
    sql = f"""
    WITH base AS (
        SELECT dni_paciente,
            (fecha_atencion::date - fecha_nacimiento::date) as edad_dias,
            EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) as edad_anios,
            EXTRACT(MONTH FROM age(fecha_atencion::date, fecha_nacimiento::date)) as edad_meses,
            ROW_NUMBER() OVER (PARTITION BY dni_paciente ORDER BY fecha_atencion, id_cita) as nro_control
        FROM {qt(tabla)} t
        WHERE {where} AND codigo_item IN ('{cn}')
    )
    SELECT * FROM (
        SELECT
            {age_case} as grupo_etareo,
            SUM(CASE WHEN nro_control = 1 THEN 1 ELSE 0 END) as c1,
            SUM(CASE WHEN nro_control = 2 THEN 1 ELSE 0 END) as c2,
            SUM(CASE WHEN nro_control = 3 THEN 1 ELSE 0 END) as c3,
            SUM(CASE WHEN nro_control = 4 THEN 1 ELSE 0 END) as c4,
            SUM(CASE WHEN nro_control = 5 THEN 1 ELSE 0 END) as c5,
            SUM(CASE WHEN nro_control = 6 THEN 1 ELSE 0 END) as c6,
            SUM(CASE WHEN nro_control = 7 THEN 1 ELSE 0 END) as c7,
            SUM(CASE WHEN nro_control = 8 THEN 1 ELSE 0 END) as c8,
            SUM(CASE WHEN nro_control > 8 THEN 1 ELSE 0 END) as c9mas,
            COUNT(*) as total_atenciones,
            COUNT(DISTINCT dni_paciente) as total_ninos
        FROM base
        GROUP BY {age_case}
    ) sub
    ORDER BY
        CASE sub.grupo_etareo
            WHEN 'Reci\u00e9n Nacido Sano 3-6d' THEN 1
            WHEN 'Reci\u00e9n Nacido Sano 7-13d' THEN 2
            WHEN 'Reci\u00e9n Nacido Sano 14-21d' THEN 3
            WHEN 'Reci\u00e9n Nacido Sano may_22d' THEN 4
            WHEN '< 1 a\u00f1o' THEN 5
            WHEN '1 a\u00f1o' THEN 6
            WHEN '2 a\u00f1os' THEN 7
            WHEN '3 a\u00f1os' THEN 8
            WHEN '4 a\u00f1os' THEN 9
            WHEN '5-9 a\u00f1os' THEN 10
            WHEN '10-11 a\u00f1os' THEN 11
            ELSE 12
        END
    """
    cur.execute(sql)
    filas_brutas = cur.fetchall()
    pg_cols = [desc[0] for desc in cur.description]
    col_names = ['EDADES','1ER.CONTROL','2DO.CONTROL','3ER.CONTROL','4TO.CONTROL',
                 '5TO.CONTROL','6TO.CONTROL','7MO.CONTROL','8VO.CONTROL',
                 'TOTAL ATENCIONES','TOTAL NI\u00d1OS ATENDIDOS']
    display_to_pg = {
        '1ER.CONTROL': 'c1', '2DO.CONTROL': 'c2', '3ER.CONTROL': 'c3',
        '4TO.CONTROL': 'c4', '5TO.CONTROL': 'c5', '6TO.CONTROL': 'c6',
        '7MO.CONTROL': 'c7', '8VO.CONTROL': 'c8',
        'TOTAL ATENCIONES': 'total_atenciones',
        'TOTAL NI\u00d1OS ATENDIDOS': 'total_ninos',
    }

    filas = []
    totales_main = {cn:0 for cn in col_names[1:]}
    for row in filas_brutas:
        d = dict(zip(pg_cols, row))
        fila = {'EDADES': d.get('grupo_etareo', '')}
        for cname in col_names[1:]:
            pg_key = display_to_pg.get(cname, '')
            val = d.get(pg_key, 0) or 0
            fila[cname] = val
            totales_main[cname] += val
        filas.append(fila)

    secciones = []

    rn_data = _qflat(ATENCION_RN_CODES, ATENCION_RN_NAMES, rn_ace, RN_AGE_ORDER)
    secciones.append({'titulo': 'ATENCI\u00d3N DEL RECI\u00c9N NACIDO', **rn_data})

    ses_data = _qval(SESIONES_CODES, ['1','2','3','4'], SESIONES_LABELS)
    secciones.append({'titulo': 'SESIONES DE ESTIMULACI\u00d3N TEMPRANA', **ses_data})

    lac_data = _qflat(LACTANCIA_CODES, LACTANCIA_NAMES)
    secciones.append({'titulo': 'LACTANCIA MATERNA', **lac_data})

    eval_data = _qval(EVAL_DESARROLLO_CODES, ['LEN','MOT','SOC','COO','COG'], EVAL_DESARROLLO_AREAS)
    secciones.append({'titulo': 'EVALUACI\u00d3N DEL DESARROLLO', **eval_data})

    plan_data = _qflat(PLAN_INTEGRAL_CODES, PLAN_INTEGRAL_NAMES)
    secciones.append({'titulo': 'PLAN INTEGRAL', **plan_data})

    consejeria_names = {c: c for c in CONSEJERIA_CODES}
    cons_data = _qflat(CONSEJERIA_CODES, consejeria_names)
    secciones.append({'titulo': 'CONSEJER\u00cdA', **cons_data})

    bpn_codes = ['P070','P071','P0711','P0712','P0713','P072','P073']
    bpn_cs = "','".join(bpn_codes)
    cur.execute(f"""
        SELECT COUNT(DISTINCT dni_paciente) as total_pacientes
        FROM {qt(tabla)} t
        WHERE {where} AND codigo_item IN ('{bpn_cs}')
    """)
    bpn_total = cur.fetchone()[0] or 0
    secciones.append({'titulo': 'RECI\u00c9N NACIDO BAJO PESO / PREMATURO',
        'columnas': ['INDICADOR', 'TOTAL'],
        'filas': [{'INDICADOR': 'Pacientes con diagn\u00f3stico RN BPN/Prematuro', 'TOTAL': bpn_total}],
        'totales': {'TOTAL': bpn_total}})

    # ---- EVALUACI\u00d3N NUTRICIONAL (sub-secci\u00f3n) ----
    def _qmonth_pivot(age_case_label, age_case_sql, age_order):
        cs = "','".join(CRED_CODES)
        cur.execute(f"""
            WITH age_cte AS (
                SELECT EXTRACT(MONTH FROM fecha_atencion::date) as mes,
                       (fecha_atencion::date - fecha_nacimiento::date) as edad_dias,
                       EXTRACT(YEAR FROM age(fecha_atencion::date, fecha_nacimiento::date)) as edad_anios,
                       EXTRACT(MONTH FROM age(fecha_atencion::date, fecha_nacimiento::date)) as edad_meses,
                       dni_paciente
                FROM {qt(tabla)} t
                WHERE {where} AND codigo_item IN ('{cs}')
            )
            SELECT mes, {age_case_sql} as grupo, COUNT(DISTINCT dni_paciente) as total
            FROM age_cte
            GROUP BY mes, grupo
            ORDER BY mes, grupo
        """)
        rows = cur.fetchall()
        months = [1,2,3,4,5,6]
        all_groups = sorted(set(r[1] for r in rows if r[1]), key=lambda x: age_order.get(x, 99))
        pivot = {}
        for m in months:
            pivot[m] = {g: 0 for g in all_groups}
        for m, g, t in rows:
            if g and m in pivot:
                pivot[int(m)][g] = pivot[int(m)].get(g, 0) + t
        filas = []
        totales = {g: 0 for g in all_groups}
        for m in months:
            row = {'Mes': f'Mes {m}'}
            total = 0
            for g in all_groups:
                v = pivot[m][g]
                row[g] = v
                totales[g] += v
                total += v
            row['Total'] = total
            filas.append(row)
        totales['Total'] = sum(totales.values())
        total_row = {'Mes': 'Total'}
        total_row.update(totales)
        filas.append(total_row)
        columnas = ['Mes'] + all_groups + ['Total']
        return {'columnas': columnas, 'filas': filas, 'totales': totales}

    rn_age_groups = {
        'Reci\u00e9n Nacido Sano 3-6d': 1,
        'Reci\u00e9n Nacido Sano 7-13d': 2,
        'Reci\u00e9n Nacido Sano 14-21d': 3,
        'Reci\u00e9n Nacido Sano may_22d': 4
    }
    eval_nutric_data = _qmonth_pivot('RN', age_case_cte, rn_age_groups)
    secciones.append({'titulo': 'EVALUACI\u00d3N NUTRICIONAL', **eval_nutric_data})

    consulta_nutric_age_case = """
        CASE
            WHEN edad_dias BETWEEN 0 AND 27 THEN 'Reci\u00e9n Nacido'
            WHEN edad_anios = 0 AND edad_meses BETWEEN 4 AND 5 THEN '4-5m'
            WHEN edad_anios = 0 AND edad_meses BETWEEN 6 AND 11 THEN '6-11m'
            WHEN edad_anios = 1 THEN '1 a\u00f1o'
            WHEN edad_anios = 2 THEN '2 a\u00f1os'
            WHEN edad_anios = 3 THEN '3 a\u00f1os'
            WHEN edad_anios = 4 THEN '4 a\u00f1os'
            WHEN edad_anios BETWEEN 5 AND 11 THEN '5-11 a\u00f1os'
            ELSE 'Otros'
        END"""
    consulta_ao = {'Reci\u00e9n Nacido':1,'4-5m':2,'6-11m':3,'1 a\u00f1o':4,'2 a\u00f1os':5,'3 a\u00f1os':6,'4 a\u00f1os':7,'5-11 a\u00f1os':8}
    consulta_nutric_data = _qmonth_pivot('Consulta', consulta_nutric_age_case, consulta_ao)
    secciones.append({'titulo': 'CONSULTA NUTRICIONAL', **consulta_nutric_data})

    pages.append({
        'id': 'formato_nino', 'titulo': 'FORMATO NI\u00d1O - CONTROL DE CRECIMIENTO Y DESARROLLO',
        'columnas': col_names, 'filas': filas, 'totales': totales_main,
        'secciones': secciones
    })

    # ---- PAGE 2: SUPLEMENTACION ----
    sup_secciones = []

    for code, name in [('U1692', 'Sulfato Ferroso'), ('59430', 'Polimaltosado/Hierro Polimaltosado'), ('U140', 'Otros Suplementos')]:
        cur.execute(f"""
            SELECT {ace} as grupo_edad, COUNT(*) as total
            FROM {qt(tabla)} t
            WHERE {where} AND codigo_item = '{code}'
            GROUP BY grupo_edad
            ORDER BY grupo_edad
        """)
        srows = cur.fetchall()
        sitems = []
        stotal = 0
        for ge, tot in srows:
            sitems.append({'label': ge, 'valor': tot})
            stotal += tot
        sup_secciones.append({
            'titulo': name,
            'columnas': ['GRUPO EDAD', 'TOTAL'],
            'filas': sitems,
            'total': stotal
        })

    supl_entrega = [
        ('99199.26', 'SUPLEMENTACI\u00d3N GENERAL', ['1','2','3','4','5','6','7','TA'], {'1':'1ra','2':'2da','3':'3ra','4':'4ta','5':'5ta','6':'6ta','7':'7ma','TA':'TA'}),
        ('99199.17', 'SUPLEMENTACI\u00d3N HIERRO NI\u00d1OS', ['1','2','3','4','5','6','TA'], {'1':'1ra','2':'2da','3':'3ra','4':'4ta','5':'5ta','6':'6ta','TA':'TA'}),
        ('99199.27', 'VITAMINA A', ['VA1','VA2'], {'VA1':'1ra Dosis','VA2':'2da Dosis'}),
        ('99403.01', 'MULTIMICRONUTRIENTES', ['1','2','3','4','5','6'], {'1':'1ra','2':'2da','3':'3ra','4':'4ta','5':'5ta','6':'6ta'}),
    ]
    for code, titulo, vlabels, vmap in supl_entrega:
        data = _qval([code], vlabels, vmap)
        sup_secciones.append({'titulo': titulo, **data})

    pages.append({
        'id': 'suplementacion', 'titulo': 'SUPLEMENTACION PREVENTIVA',
        'secciones': sup_secciones if sup_secciones else [{'titulo':'Sin datos','columnas':['INDICADOR','TOTAL'],'filas':[],'total':0}]
    })

    # ---- PAGE 3: Tx. ANEMIA ----
    ane_secciones = []

    for code, name in [('85018.01', 'Dosaje de Hemoglobina'), ('85018', 'Dosaje de Hemoglobina')]:
        cur.execute(f"""
            SELECT {ace} as grupo_edad, COUNT(*) as total
            FROM {qt(tabla)} t
            WHERE {where} AND codigo_item = '{code}'
            GROUP BY grupo_edad
            ORDER BY grupo_edad
        """)
        arows = cur.fetchall()
        aitems = []
        atotal = 0
        for ge, tot in arows:
            aitems.append({'label': ge, 'valor': tot})
            atotal += tot
        ane_secciones.append({
            'titulo': name,
            'columnas': ['GRUPO EDAD', 'TOTAL'],
            'filas': aitems,
            'total': atotal
        })

    cur.execute(f"""
        SELECT {ace} as grupo_edad, COUNT(*) as total
        FROM {qt(tabla)} t
        WHERE {where} AND codigo_item = 'C0011'
        GROUP BY grupo_edad
        ORDER BY grupo_edad
    """)
    arows = cur.fetchall()
    aitems = []
    atotal = 0
    for ge, tot in arows:
        aitems.append({'label': ge, 'valor': tot})
        atotal += tot
    ane_secciones.append({
        'titulo': 'Tratamiento Anemia/Visita Seguimiento',
        'columnas': ['GRUPO EDAD', 'TOTAL'],
        'filas': aitems,
        'total': atotal
    })

    pages.append({
        'id': 'tx_anemia', 'titulo': 'TRATAMIENTO DE ANEMIA',
        'secciones': ane_secciones if ane_secciones else [{'titulo':'Sin datos','columnas':['INDICADOR','TOTAL'],'filas':[],'total':0}]
    })

    # Add tabla_html for all pages
    for p in pages:
        secs = p.get('secciones', [])
        if secs and not p.get('tabla_html'):
            h = '<style>' + _REPORT_CSS + '</style>'
            h += '<div class="page">'
            h += _sec_header({**filtros, 'anio': str(anio)})
            for s in secs:
                h += _render_seccion_table(s)
            h += '</div>'
            p['tabla_html'] = _wrap_tables(h)

    return {
        'tipo': 'formato_nino_cred', 'anio': anio, 'filtros': filtros, 'meses': meses,
        'paginas': pages
    }


def _reporte_completo_cred(cur, esquema, anio, meses, filtros):
    """Route to appropriate implementation based on year."""
    if str(anio) == '2024':
        try:
            return _reporte_cred_2024(cur, esquema, anio, meses, filtros)
        except Exception as e:
            print(f"cred2024 tables not available: {e}")
            return _reporte_completo_cred_old(cur, esquema, anio, meses, filtros)
    return _reporte_completo_cred_old(cur, esquema, anio, meses, filtros)

# ============ REPORTE IRAS EDAS (2024) ============
def _query_age_groups(cur, anio, codigos, filtros):
    """Query CPT/procedure codes from his_proceso_{anio} grouped by age.

    Returns dict: {code: {'men2m':N, '2_11m':N, '1a':N, '2a_4a':N, '5_11a':N}}
    """
    tbl = f'es_ivan.his_proceso_{anio}'
    where_parts = [f"anio = {anio}"]
    if filtros:
        for col in ['red','microred','nombre_establecimiento','provincia','distrito']:
            val = filtros.get(col, '')
            if val:
                where_parts.append(f"LOWER({col}) LIKE LOWER('%{val.replace(chr(39), chr(39)+chr(39))}%')")
    where = ' AND '.join(where_parts)
    codes_str = ','.join(f"'{c}'" for c in codigos)
    sql = f"""
        SELECT codigo_item,
            CASE
                WHEN tip_edad = 'D' THEN 'men2m'
                WHEN tip_edad = 'M' AND edad <= 11 THEN '2_11m'
                WHEN tip_edad = 'A' AND edad = 1 THEN '1a'
                WHEN tip_edad = 'A' AND edad >= 2 AND edad <= 4 THEN '2a_4a'
                WHEN tip_edad = 'A' AND edad >= 5 AND edad <= 11 THEN '5_11a'
                ELSE 'men2m'
            END as grupo,
            COUNT(*) as cnt
        FROM {tbl}
        WHERE {where} AND codigo_item IN ({codes_str})
        GROUP BY codigo_item, grupo
        ORDER BY codigo_item, grupo
    """
    cur.execute(sql)
    result = {}
    for code in codigos:
        result[code] = {'men2m':0, '2_11m':0, '1a':0, '2a_4a':0, '5_11a':0}
    for row in cur.fetchall():
        code, grupo, cnt = row
        if code in result:
            if grupo in result[code]:
                result[code][grupo] = (result[code][grupo] or 0) + cnt
            else:
                result[code][grupo] = cnt
    return result


def _query_iras_ref(cur, anio, filtros):
    """Query his_proceso_{anio} with exact ref HTML age groups for ALL CIE10 diagnosis codes.

    Age groups: <29d, 29d_59d, 2_11m, 1_4a, 5_11a
    Returns dict: {category: {age_group: count, subtotal: N, grand_total: N}}
    """
    tbl = f'es_ivan.his_proceso_{anio}'
    where_parts = [f"anio = {anio}"]
    if filtros:
        for col in ['red','microred','nombre_establecimiento','provincia','distrito']:
            val = filtros.get(col, '')
            if val:
                where_parts.append(f"LOWER({col}) LIKE LOWER('%{val.replace(chr(39), chr(39)+chr(39))}%')")
    where = ' AND '.join(where_parts)

    # Map of category -> list of CIE10 codes
    CATEGORIES = {
        'ira_no_compl': ['J00X','J040','J041','J042','J060','J068','J069','J209'],
        'faringo': ['J020','J029','J030','J038','J039'],
        'oma': ['H650','H651','H660','H669'],
        'sinusitis': ['J010','J011','J012','J013','J014','J019'],
        'neumonia_sin_compl': ['J129','J159','J189'],
        'iras_con_compl': ['A369','A370','A371','A378','A379','J120','J121','J122','J123','J128','J13X','J14X',
                          'J150','J151','J152','J153','J154','J157','J158','J159','J160','J168'],
        'neumonia_grave_men2m': ['J050','J051','J851','J860','J869','J90X','J939','J100','J110'],
        'neumonia_emg_2m_4a': ['J050','J051','J851','J860','J869','J90X','J939','J100','J110'],
        'sob_asma': ['J210','J211','J218','J219','J440','J441','J448','J449','J450','J451','J458','J459','J4591','J46X'],
    }

    all_codes = list(set(sum(CATEGORIES.values(), [])))
    codes_str = ','.join(f"'{c}'" for c in all_codes)

    AGE_CASE = '''
        CASE
            WHEN tip_edad = 'D' AND edad < 29 THEN 'men29d'
            WHEN tip_edad = 'D' AND edad >= 29 AND edad <= 59 THEN '29d_59d'
            WHEN (tip_edad = 'M' AND edad <= 11) OR (tip_edad = 'D' AND edad > 59) THEN '2_11m'
            WHEN tip_edad = 'A' AND edad >= 1 AND edad <= 4 THEN '1_4a'
            WHEN tip_edad = 'A' AND edad >= 5 AND edad <= 11 THEN '5_11a'
            WHEN tip_edad = 'M' AND edad >= 12 THEN '1_4a'
            ELSE 'men29d'
        END'''

    AGE_CASE_SOB = '''
        CASE
            WHEN tip_edad = 'D' AND edad < 29 THEN 'men29d'
            WHEN tip_edad = 'D' AND edad >= 29 AND edad <= 59 THEN '29d_59d'
            WHEN (tip_edad = 'M' AND edad <= 11) OR (tip_edad = 'D' AND edad > 59) THEN '2_11m'
            WHEN (tip_edad = 'M' AND edad >= 12 AND edad <= 23) OR (tip_edad = 'A' AND edad = 1) THEN '12_23m'
            WHEN tip_edad = 'A' AND edad = 2 THEN '2a'
            WHEN tip_edad = 'A' AND edad >= 3 AND edad <= 4 THEN '3_4a'
            WHEN tip_edad = 'A' AND edad >= 5 AND edad <= 11 THEN '5_11a'
            WHEN tip_edad = 'M' AND edad >= 24 THEN '2a'
            ELSE 'men29d'
        END'''

    # Query IRA/EDA codes with IRA age groups
    sql_ira = f"""
        SELECT codigo_item,
            {AGE_CASE} as grupo,
            COUNT(*) as cnt
        FROM {tbl}
        WHERE {where} AND codigo_item IN ({codes_str})
        GROUP BY codigo_item, grupo
        ORDER BY codigo_item, grupo
    """
    cur.execute(sql_ira)
    ira_rows = {}
    for row in cur.fetchall():
        code, grupo, cnt = row
        if code not in ira_rows:
            ira_rows[code] = {}
        ira_rows[code][grupo] = (ira_rows[code].get(grupo, 0) or 0) + cnt

    # Query SOB codes with SOB age groups
    sob_codes = ','.join(f"'{c}'" for c in CATEGORIES['sob_asma'])
    sql_sob = f"""
        SELECT codigo_item,
            {AGE_CASE_SOB} as grupo,
            COUNT(*) as cnt
        FROM {tbl}
        WHERE {where} AND codigo_item IN ({sob_codes})
        GROUP BY codigo_item, grupo
        ORDER BY codigo_item, grupo
    """
    cur.execute(sql_sob)
    sob_rows = {}
    for row in cur.fetchall():
        code, grupo, cnt = row
        if code not in sob_rows:
            sob_rows[code] = {}
        sob_rows[code][grupo] = (sob_rows[code].get(grupo, 0) or 0) + cnt

    AGE_KEYS = ['men29d', '29d_59d', '2_11m', '1_4a', '5_11a']
    SOB_AGE_KEYS = ['men29d', '29d_59d', '2_11m', '12_23m', '2a', '3_4a', '5_11a']

    def _agg(rows, codes, age_keys):
        totals = {k: 0 for k in age_keys}
        for code in codes:
            for ak in age_keys:
                totals[ak] += rows.get(code, {}).get(ak, 0)
        subtotal = sum(totals.values())
        return {**totals, 'subtotal': subtotal}

    result = {}
    for cat, codes in CATEGORIES.items():
        if cat == 'sob_asma':
            result[cat] = _agg(sob_rows, codes, SOB_AGE_KEYS)
        else:
            result[cat] = _agg(ira_rows, codes, AGE_KEYS)

    # Compute grand total for IRA without complications (sum of 5 sub-items)
    ira_sin_compl_keys = ['ira_no_compl', 'faringo', 'oma', 'sinusitis', 'neumonia_sin_compl']
    result['ira_sin_compl'] = {}
    for k in AGE_KEYS + ['subtotal']:
        result['ira_sin_compl'][k] = sum(result.get(cat, {}).get(k, 0) for cat in ira_sin_compl_keys)

    # Grand total for IRA without complications = sum of all subtotals
    result['ira_sin_compl']['grand_total'] = sum(result.get(cat, {}).get('subtotal', 0) for cat in ira_sin_compl_keys)

    # IRA with complications
    ira_con_compl_keys = ['iras_con_compl', 'neumonia_grave_men2m', 'neumonia_emg_2m_4a']
    result['ira_con_compl'] = {}
    for k in AGE_KEYS + ['subtotal']:
        result['ira_con_compl'][k] = sum(result.get(cat, {}).get(k, 0) for cat in ira_con_compl_keys)
    result['ira_con_compl']['grand_total'] = sum(result.get(cat, {}).get('subtotal', 0) for cat in ira_con_compl_keys)

    # Total de Casos de IRA = ira_sin_compl + ira_con_compl
    result['ira_total'] = {}
    for k in AGE_KEYS + ['subtotal']:
        result['ira_total'][k] = result['ira_sin_compl'].get(k, 0) + result['ira_con_compl'].get(k, 0)
    result['ira_total']['grand_total'] = result['ira_sin_compl']['grand_total'] + result['ira_con_compl']['grand_total']

    return result


def _query_resumen_mensual(cur, anio, filtros):
    """Query monthly summary data for Resumen Acumulado section.

    Returns list of dicts: {mes, eda_acuosa_men1a, eda_acuosa_1a, ...}
    """
    tbl = f'es_ivan.his_proceso_{anio}'
    where_parts = [f"anio = {anio}"]
    if filtros:
        for col in ['red','microred','nombre_establecimiento','provincia','distrito']:
            val = filtros.get(col, '')
            if val:
                where_parts.append(f"LOWER({col}) LIKE LOWER('%{val.replace(chr(39), chr(39)+chr(39))}%')")
    where = ' AND '.join(where_parts)

    # EDA acuosa age groups: <1a, 1a, 2a, 3a, 4a (from the table columns)
    # We'll aggregate by month and age from his_proceso
    # Use simplified monthly counts from iras_edas_2024 by month
    sql = f"""
        SELECT 
            nt.mes,
            SUM(CASE WHEN nt.codigo_item IN ('A009','A000','A001') 
                     AND (nt.tip_edad='D' OR (nt.tip_edad='M' AND nt.edad < 12) OR (nt.tip_edad='A' AND nt.edad = 0))
                THEN 1 ELSE 0 END) as eda_acuosa_men1a,
            SUM(CASE WHEN nt.codigo_item IN ('A009','A000','A001') 
                     AND nt.tip_edad='A' AND nt.edad = 1
                THEN 1 ELSE 0 END) as eda_acuosa_1a,
            SUM(CASE WHEN nt.codigo_item IN ('A009','A000','A001') 
                     AND nt.tip_edad='A' AND nt.edad = 2
                THEN 1 ELSE 0 END) as eda_acuosa_2a,
            SUM(CASE WHEN nt.codigo_item IN ('A009','A000','A001') 
                     AND nt.tip_edad='A' AND nt.edad = 3
                THEN 1 ELSE 0 END) as eda_acuosa_3a,
            SUM(CASE WHEN nt.codigo_item IN ('A009','A000','A001') 
                     AND nt.tip_edad='A' AND nt.edad = 4
                THEN 1 ELSE 0 END) as eda_acuosa_4a,
            SUM(CASE WHEN nt.codigo_item IN ('J00X','J040','J041','J042','J060','J068','J069','J209') 
                     AND nt.tip_edad='A' AND nt.edad < 5
                THEN 1 ELSE 0 END) as ira_no_compl_men5a,
            SUM(CASE WHEN nt.codigo_item IN ('J020','J029','J030','J038','J039') 
                     AND nt.tip_edad='A' AND nt.edad < 5
                THEN 1 ELSE 0 END) as faringo_men5a,
            SUM(CASE WHEN nt.codigo_item IN ('H650','H651','H660','H669') 
                     AND nt.tip_edad='A' AND nt.edad < 5
                THEN 1 ELSE 0 END) as oma_men5a,
            SUM(CASE WHEN nt.codigo_item IN ('J129','J159','J189') 
                     AND nt.tip_edad='A' AND nt.edad < 5
                THEN 1 ELSE 0 END) as neumonia_men5a,
            SUM(CASE WHEN nt.codigo_item IN ('A369','A370','A371','A378','A379','J120','J121','J122','J123','J128','J13X','J14X','J150','J151','J152','J153','J154','J157','J158','J159','J160','J168')
                THEN 1 ELSE 0 END) as iras_compl,
            COUNT(*) as foni
        FROM {tbl} nt
        WHERE {where}
        GROUP BY nt.mes
        ORDER BY nt.mes
    """
    try:
        cur.execute(sql)
        rows = []
        keys = ['eda_acuosa_men1a','eda_acuosa_1a','eda_acuosa_2a','eda_acuosa_3a','eda_acuosa_4a',
                'ira_no_compl_men5a','faringo_men5a','oma_men5a','neumonia_men5a','iras_compl','foni']
        for r in cur.fetchall():
            row = {'mes': r[0]}
            for i, k in enumerate(keys):
                row[k] = r[i+1] or 0
            rows.append(row)
        return rows
    except Exception as e:
        print(f"Resumen query error: {e}")
        return []


def _reporte_iras_edas_2024(cur, esquema, anio, meses, filtros):
    """IRAS EDAS report matching ref HTML exactly using his_proceso_2024."""
    where = _build_cred_where(2024, meses, filtros)
    ie = _qcred_cols(cur, {'IE':'iras_edas_2024'}, where)

    # Query ALL CIE10 diagnosis data with ref age groups
    diag = _query_iras_ref(cur, 2024, filtros)

    # Query CPT procedure data (oxigenoterapia, oximetria, nebulizacion)
    proc_raw = _query_age_groups(cur, 2024, ['94799.02','94760','94664'], filtros)
    def _proc_sum(code, age_map):
        d = proc_raw.get(code, {})
        return {ref_k: d.get(db_k, 0) for ref_k, db_k in age_map.items()}
    oxi_map = {'men29d': 'men2m', '29d_59d': 'men2m', '2_11m': '2_11m', '1_4a': '1a', '5_11a': '5_11a'}
    sob_oxi_map = {'men29d': 'men2m', '29d_59d': 'men2m', '2_11m': '2_11m', '12_23m': '1a', '2a': '2a_4a', '3_4a': '2a_4a', '5_11a': '5_11a'}
    proc = {
        'oxigeno_ira': _proc_sum('94799.02', oxi_map),
        'oximetria_ira': _proc_sum('94760', oxi_map),
        'oxigeno_sob': _proc_sum('94799.02', sob_oxi_map),
        'nebulizacion_sob': _proc_sum('94664', sob_oxi_map),
    }

    # Query Resumen Acumulado
    resumen = _query_resumen_mensual(cur, 2024, filtros)

    from html_report_builder import build_iras_edas_html
    tabla_html = build_iras_edas_html({
        'filtros': {**filtros, 'anio': anio},
        'anio': anio,
        'ie': ie,
        'diag': diag,
        'proc': proc,
        'resumen': resumen,
    })

    return {
        'tipo': 'formato_iras_edas', 'anio': anio, 'filtros': filtros, 'meses': meses,
        'paginas': [
            {
                'id': 'iras_edas', 'titulo': 'FORMATO IRAS - EDAS',
                'tabla_html': tabla_html,
            },
        ]
    }


def _build_cred_suplementacion(cred):
    """Build suplementacion sections from cred2024 data dict (C1/C2/C4 prefixed)."""
    v = lambda k: cred.get(k, 0) or 0
    secciones = []

    def _e(titulo, cols):
        items = []
        total = 0
        for label, col in cols:
            val = v(col)
            items.append({'INDICADOR': label, 'TOTAL': val})
            total += val
        return {'titulo': titulo, 'columnas': ['INDICADOR', 'TOTAL'], 'filas': items, 'total': total}

    secciones.append(_e('1. Sup. Preventiva <6m', [
        ('RN Bajo Peso 1ra', 'C1_ta_suplem_bpn'),
        ('4m Sano', 'C1_suplem_4m_sano'),
        ('5m Sano', 'C1_suplem_5m_sano'),
        ('TA 5-6m Sano', 'C1_ta_suplem_5_6m_sano'),
    ]))

    for grupo, suf in [('6-11m','6_11m'),('1a','1a'),('2a','2a'),('3a','3a'),('4a','4a'),('5-11a','5_11a')]:
        secciones.append(_e(f'Multimicronutrientes {grupo}', [
            (f'1ra Entrega', f'C1_suplem_1ra_{suf}'),
            (f'2da Entrega', f'C1_suplem_2da_{suf}'),
            (f'3ra Entrega', f'C1_suplem_3ra_{suf}'),
            (f'4ta Entrega', f'C1_suplem_4ta_{suf}'),
            (f'5ta Entrega', f'C1_suplem_5ta_{suf}'),
            (f'6ta Entrega', f'C1_suplem_6ta_{suf}'),
            (f'TA', f'C1_ta_suplem_{suf}'),
        ]))

    secciones.append(_e('4. Vitamina A', [
        ('VA1 6-11m', 'C1_va1_6_11m'),
        ('VA1 1a', 'C1_va1_1a'), ('VA2 1a', 'C1_va2_1a'),
        ('VA1 2a', 'C1_va1_2a'), ('VA2 2a', 'C1_va2_2a'),
        ('VA1 3a', 'C1_va1_3a'), ('VA2 3a', 'C1_va2_3a'),
        ('VA1 4a', 'C1_va1_4a'), ('VA2 4a', 'C1_va2_4a'),
    ]))

    secciones.append(_e('9. Sup. Hierro MEF/Gestantes/Puerperas', [
        ('Gestante 1', 'C1_suplem_gest1'), ('Gestante 2', 'C1_suplem_gest2'),
        ('Gestante 3', 'C1_suplem_gest3'), ('Gestante 4', 'C1_suplem_gest4'),
        ('Gestante 5', 'C1_suplem_gest5'), ('Gestante 6', 'C1_suplem_gest6'),
        ('Gestante TA', 'C1_suplem_gest_ta'),
        ('Pu\u00e9rpera 1', 'C1_suplem_puer1'), ('Pu\u00e9rpera 2', 'C1_suplem_puer2'),
        ('Pu\u00e9rpera 3', 'C1_suplem_puer3'), ('Pu\u00e9rpera 4', 'C1_suplem_puer4'),
        ('Pu\u00e9rpera 5', 'C1_suplem_puer5'), ('Pu\u00e9rpera 6', 'C1_suplem_puer6'),
        ('Pu\u00e9rpera 7', 'C1_suplem_puer7'), ('Pu\u00e9rpera TA', 'C1_suple_puer_ta'),
        ('MEF 1', 'C1_suplem_mef1'), ('MEF 2', 'C1_suplem_mef2'),
        ('MEF 3', 'C1_suplem_mef3'), ('MEF TA', 'C1_suplem_mef_ta'),
        ('Adol. 12-17 1', 'C1_suplem1_12_17'), ('Adol. 12-17 2', 'C1_suplem2_12_17'),
        ('Adol. 12-17 3', 'C1_suplem3_12_17'), ('Adol. 12-17 TA', 'C1_suplem4_12_17ta'),
    ]))

    secciones.append(_e('6. DOSAJE DE HEMOGLOBINA', [
        ('1er Dosaje BPN', 'C2_primer_dosaje_hb_bpn'),
        ('2do Dosaje BPN', 'C2_segundo_dosaje_hb_bpn'),
        ('HB 6-11m 1er', 'C2_hb_6_11m_primer'),
        ('HB 6-11m 2do', 'C2_hb_6_11m_segundo'),
        ('HB 12-23m 1er', 'C2_hb_12_23m_primer'),
        ('HB 12-23m 2do', 'C2_hb_12_23m_segundo'),
        ('HB 12-23m 3er', 'C2_hb_12_23m_tercer'),
        ('HB 24-35m 1er', 'C2_hb_24_35m_primer'),
        ('HB 24-35m 2do', 'C2_hb_24_35m_segundo'),
        ('HB 36-59m 1er', 'C2_hb_36_59m_primer'),
        ('HB 36-59m 2do', 'C2_hb_36_59m_segundo'),
        ('HB 5-11a', 'C2_hb_5_11a'),
        ('HB Adolesc. 1er', 'C2_hb_adolescente_primer'),
        ('HB Adolesc. 2do', 'C2_hb_adolescente_segundo'),
        ('HB Gestante 1er', 'C2_hb_gestante_primer'),
        ('HB Gestante 2do', 'C2_hb_gestante_segundo'),
        ('HB Gestante 3er', 'C2_hb_gestante_tercero'),
        ('HB MEF', 'C2_hb_mef'),
        ('HB Pu\u00e9rpera', 'C2_hb_puerpera'),
    ]))

    secciones.append(_e('7. CONSULTA NUTRICIONAL', [
        ('Cons. Nutr. BPN', 'C4_consulta_nutric_bpn'),
        ('Cons. Nutr. 4-5m', 'C4_consulta_nutric1_4_5m'),
        ('Cons. Nutr. 6-11m', 'C4_consulta_nutric_1_6_11m'),
        ('Cons. Nutr. 1a', 'C4_consulta_nutric_1_1a'),
        ('Cons. Nutr. 2a', 'C4_consulta_nutric_1_2a'),
        ('Cons. Nutr. 3a', 'C4_consulta_nutric_1_3a'),
        ('Cons. Nutr. 4a', 'C4_consulta_nutric_1_4a'),
        ('Cons. Nutr. 5-11a', 'C4_consulta_nutric_1_5_11a'),
        ('Cons. Nutr. Gestante 1', 'C1_cons_nutric1_gest'),
        ('Cons. Nutr. Gestante 2', 'C1_cons_nutric2_gest'),
        ('Cons. Nutr. Puerpera 1', 'C1_consulta_nutric1_gest'),
        ('Cons. Nutr. MEF 1', 'C1_cons_mef1'),
        ('Cons. Nutr. MEF 2', 'C1_cons_mef2'),
        ('Cons. Nutr. Adolesc. 1', 'C1_cons_nutri1_12_17'),
        ('Cons. Nutr. Adolesc. 2', 'C1_cons_nutri2_12_17'),
        ('Cons. Nutr. Adolesc. 3', 'C1_cons_nutri3_12_17'),
    ]))

    return secciones


def _build_cred_tx_anemia(cred):
    v = lambda k: cred.get(k, 0) or 0
    secciones = []

    def _e(titulo, cols):
        items = []
        total = 0
        for label, col in cols:
            val = v(col)
            items.append({'INDICADOR': label, 'TOTAL': val})
            total += val
        return {'titulo': titulo, 'columnas': ['INDICADOR', 'TOTAL'], 'filas': items, 'total': total}

    for ane_suf in ['6_11m', '1a', '2a', '3a', '4a']:
        secciones.append(_e(f'Adm. Tx. Sulfato Ferroso {ane_suf}', [
            (f'1ra', f'C1_suplem_1ra_{ane_suf}'),
            (f'2da', f'C1_suplem_2da_{ane_suf}'),
            (f'3ra', f'C1_suplem_3ra_{ane_suf}'),
            (f'4ta', f'C1_suplem_4ta_{ane_suf}'),
            (f'5ta', f'C1_suplem_5ta_{ane_suf}'),
            (f'6ta', f'C1_suplem_6ta_{ane_suf}'),
            (f'TA', f'C1_ta_suplem_{ane_suf}'),
        ]))

    secciones.append(_e('2. DOSAJE HEMOGLOBINA CONTROL', [
        ('1er Dosaje BPN', 'C2_primer_dosaje_hb_bpn'),
        ('2do Dosaje BPN', 'C2_segundo_dosaje_hb_bpn'),
        ('HB 6-11m 1er', 'C2_hb_6_11m_primer'),
        ('HB 6-11m 2do', 'C2_hb_6_11m_segundo'),
        ('HB 12-23m 1er', 'C2_hb_12_23m_primer'),
        ('HB 12-23m 2do', 'C2_hb_12_23m_segundo'),
        ('HB 12-23m 3er', 'C2_hb_12_23m_tercer'),
        ('HB 24-35m 1er', 'C2_hb_24_35m_primer'),
        ('HB 24-35m 2do', 'C2_hb_24_35m_segundo'),
        ('HB 36-59m 1er', 'C2_hb_36_59m_primer'),
        ('HB 36-59m 2do', 'C2_hb_36_59m_segundo'),
        ('HB 5-11a', 'C2_hb_5_11a'),
        ('HB Adolesc. 1er', 'C2_hb_adolescente_primer'),
        ('HB Adolesc. 2do', 'C2_hb_adolescente_segundo'),
        ('HB Gestante 1er', 'C2_hb_gestante_primer'),
        ('HB Gestante 2do', 'C2_hb_gestante_segundo'),
        ('HB Gestante 3er', 'C2_hb_gestante_tercero'),
        ('HB MEF', 'C2_hb_mef'),
        ('HB Pu\u00e9rpera', 'C2_hb_puerpera'),
    ]))

    secciones.append(_e('3. CONSULTA MEDICA', [
        ('Anamnesis y Ex. F\u00edsico RN', 'C4_anamnesis_y_ex_fisico_rn_normal'),
        ('Evaluaci\u00f3n M\u00e9dica RN', 'C4_evaluacion_medica_rn'),
    ]))

    secciones.append(_e('4. CONSULTA NUTRICIONAL (Tx Anemia)', [
        ('Cons. Nutr. BPN', 'C4_consulta_nutric_bpn'),
        ('Cons. Nutr. 4-5m', 'C4_consulta_nutric1_4_5m'),
        ('Cons. Nutr. 6-11m 1', 'C4_consulta_nutric_1_6_11m'),
        ('Cons. Nutr. 1a 1', 'C4_consulta_nutric_1_1a'),
        ('Cons. Nutr. 2a 1', 'C4_consulta_nutric_1_2a'),
        ('Cons. Nutr. 3a 1', 'C4_consulta_nutric_1_3a'),
        ('Cons. Nutr. 4a 1', 'C4_consulta_nutric_1_4a'),
        ('Cons. Nutr. 5-11a 1', 'C4_consulta_nutric_1_5_11a'),
    ]))

    return secciones


# ============ REPORTE IRAS EDAS (dispatch) ============
def _reporte_iras_edas(cur, esquema, anio, meses, filtros):
    """Route to appropriate implementation based on year."""
    if str(anio) == '2024':
        try:
            return _reporte_iras_edas_2024(cur, esquema, anio, meses, filtros)
        except Exception as e:
            print(f"iras_edas_2024 tables not available: {e}")
            import traceback; traceback.print_exc()
            return _reporte_iras_edas_old(cur, esquema, anio, meses, filtros)
    return _reporte_iras_edas_old(cur, esquema, anio, meses, filtros)


def _reporte_iras_edas_old(cur, esquema, anio, meses, filtros):
    tabla = _resolve_tabla(anio, 'iras_edas')
    where = _build_where(anio, meses, filtros)
    qt = lambda t: _qtable(esquema, t)
    ace = _build_age_case()

    def _query_cie(cie_dict):
        rows = []
        for key, cfg in cie_dict.items():
            cs = "','".join(cfg['codes'])
            cur.execute(f"""
                SELECT {ace} as grupo_edad, COUNT(*) as total
                FROM {qt(tabla)} t
                WHERE {where} AND codigo_item IN ('{cs}')
                GROUP BY grupo_edad ORDER BY grupo_edad
            """)
            items = [{'label': r[0], 'valor': r[1]} for r in cur.fetchall()]
            total = sum(it['valor'] for it in items)
            rows.append({'diagnostico': cfg['label'], 'items': items, 'total': total})
        return rows

    rows_ira = _query_cie(CIE_IRAS)
    rows_sob = _query_cie(CIE_SOB_ASMA)
    rows_eda = _query_cie(CIE_EDAS)
    rows_anemia = _query_cie(CIE_ANEMIA)
    rows_parasitos = _query_cie(CIE_PARASITOSIS)
    rows_def = _query_cie(CIE_DEFUNCIONES)
    rows_sal = _query_cie(CIE_SAL_YODADA)

    secciones = [
        {'titulo': 'INFECCION RESPIRATORIA AGUDA (IRA)', 'diagnosticos': rows_ira},
        {'titulo': 'SOB / ASMA', 'diagnosticos': rows_sob},
        {'titulo': 'ENFERMEDAD DIARREICA AGUDA (EDA)', 'diagnosticos': rows_eda},
        {'titulo': 'ANEMIA', 'diagnosticos': rows_anemia},
        {'titulo': 'PARASITOSIS INTESTINAL', 'diagnosticos': rows_parasitos},
        {'titulo': 'DEFUNCIONES', 'diagnosticos': rows_def},
        {'titulo': 'CONSUMO DE SAL YODADA', 'diagnosticos': rows_sal},
    ]

    return {
        'tipo': 'formato_iras_edas', 'anio': anio, 'filtros': filtros, 'meses': meses,
        'paginas': [{
            'id': 'iras_edas', 'titulo': 'FORMATO IRAS - EDAS',
            'secciones': secciones
        }]
    }

# ============ EXPORTAR A EXCEL ============
@app.route('/api/reportes-minsa/exportar', methods=['POST'])
def reportes_minsa_exportar():
    data = request.json or {}
    rd = data.get('data', {})
    from io import BytesIO
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        return jsonify({'error': 'openpyxl no instalado'}), 500

    wb = Workbook()
    bold = Font(bold=True, size=11)
    bold14 = Font(bold=True, size=14)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

    def style_range(ws, row_start, row_end, ncols):
        for r in range(row_start, row_end + 1):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin
                cell.alignment = center

    for pi, pagina in enumerate(rd.get('paginas', [])):
        ws = wb.create_sheet(title=pagina.get('titulo', f'Pagina{pi+1}')[:31]) if pi > 0 else wb.active
        if pi > 0:
            wb.active = ws
        else:
            ws.title = pagina.get('titulo', 'Reporte')[:31]

        r = 1
        # Title
        ws.cell(row=r, column=1, value=f'{pagina.get("titulo","")} - A�o {rd.get("anio","")}').font = bold14
        r += 2

        if pagina.get('id') == 'formato_nino':
            cols = pagina['columnas']
            # Header
            for ci, cn in enumerate(cols, 1):
                ws.cell(row=r, column=ci, value=cn)
            style_header(ws, r, len(cols))
            r += 1
            # Data
            for fila in pagina['filas']:
                for ci, cn in enumerate(cols, 1):
                    ws.cell(row=r, column=ci, value=fila.get(cn, ''))
                r += 1
            style_range(ws, 3, r-1, len(cols))
            # Totals
            totes = pagina.get('totales', {})
            for ci, cn in enumerate(cols, 1):
                cell = ws.cell(row=r, column=ci)
                if cn == 'EDADES':
                    cell.value = 'TOTALES'
                else:
                    cell.value = totes.get(cn, 0)
                cell.font = bold
                cell.fill = total_fill
                cell.border = thin
            r += 2
            # Sub-secciones
            for sec in pagina.get('secciones', []):
                if sec.get('tipo') == 'side_by_side':
                    ws.cell(row=r, column=1, value=sec.get('titulo','')).font = bold
                    r += 1
                    izq = sec.get('izquierda', {})
                    der = sec.get('derecha', {})
                    start_r = r
                    # Left panel
                    ws.cell(row=r, column=1, value=izq.get('titulo','')).font = bold
                    if izq.get('columnas') and izq.get('filas'):
                        r += 1
                        cn_i = izq.get('columnas', ['INDICADOR','TOTAL'])
                        for ci, cn in enumerate(cn_i, 1):
                            ws.cell(row=r, column=ci, value=cn)
                        style_header(ws, r, len(cn_i))
                        r += 1
                        ld = r
                        for item in izq.get('filas', []):
                            for ci, cn in enumerate(cn_i, 1):
                                ws.cell(row=r, column=ci, value=item.get(cn, ''))
                            r += 1
                        style_range(ws, ld-1, r-1, len(cn_i))
                    # Right panel  
                    ws.cell(row=r, column=8, value=der.get('titulo','')).font = bold
                    if der.get('columnas') and der.get('filas'):
                        r_col = r
                        cn_d = der.get('columnas', ['INDICADOR','TOTAL'])
                        for ci, cn in enumerate(cn_d, 8):
                            ws.cell(row=r_col, column=ci, value=cn)
                        for ci, cn in enumerate(cn_d, 8):
                            cell = ws.cell(row=r_col, column=ci)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center
                            cell.border = thin
                        r_col += 1
                        for item in der.get('filas', []):
                            for ci, cn in enumerate(cn_d, 8):
                                ws.cell(row=r_col, column=ci, value=item.get(cn, ''))
                            r_col += 1
                        style_range(ws, r+1, r_col-1, 8+len(cn_d)-1)
                        r = max(r, r_col)
                    r += 1
                    continue
                if sec.get('tipo') == 'grid' and sec.get('grid'):
                    g = sec['grid']
                    headers = g.get('headers', [])
                    for ci, cn in enumerate(headers, 1):
                        ws.cell(row=r, column=ci, value=cn)
                    style_header(ws, r, len(headers))
                    r += 1
                    data_start = r
                    for row_data in g.get('rows', []):
                        for ci, val in enumerate(row_data, 1):
                            ws.cell(row=r, column=ci, value=val)
                        r += 1
                    totales = g.get('totales', [])
                    if totales:
                        for ci, val in enumerate(totales, 1):
                            cell = ws.cell(row=r, column=ci, value=val)
                            cell.font = bold
                            cell.fill = total_fill
                            cell.border = thin
                        r += 1
                    style_range(ws, data_start - 1, r - 1, len(headers))
                    r += 1
                    continue
                colnames = sec.get('columnas', [])
                if not colnames:
                    continue
                for ci, cn in enumerate(colnames, 1):
                    ws.cell(row=r, column=ci, value=cn)
                style_header(ws, r, len(colnames))
                r += 1
                data_start = r
                items = sec.get('filas', [])
                for item in items:
                    for ci, cn in enumerate(colnames, 1):
                        ws.cell(row=r, column=ci, value=item.get(cn, ''))
                    r += 1
                totes = sec.get('totales', {})
                if totes:
                    for ci, cn in enumerate(colnames, 1):
                        cell = ws.cell(row=r, column=ci)
                        if cn == colnames[0]:
                            cell.value = 'TOTALES'
                        else:
                            cell.value = totes.get(cn, 0)
                        cell.font = bold
                        cell.fill = total_fill
                        cell.border = thin
                    r += 1
                style_range(ws, data_start - 1, r - 1, len(colnames))
                r += 1

        elif pagina.get('id') in ('suplementacion', 'tx_anemia'):
            for sec in pagina.get('secciones', []):
                ws.cell(row=r, column=1, value=sec.get('titulo','')).font = bold
                r += 1
                if sec.get('tipo') == 'grid' and sec.get('grid'):
                    g = sec['grid']
                    headers = g.get('headers', [])
                    for ci, cn in enumerate(headers, 1):
                        ws.cell(row=r, column=ci, value=cn)
                    style_header(ws, r, len(headers))
                    r += 1
                    data_start = r
                    for row_data in g.get('rows', []):
                        for ci, val in enumerate(row_data, 1):
                            ws.cell(row=r, column=ci, value=val)
                        r += 1
                    totales = g.get('totales', [])
                    if totales:
                        for ci, val in enumerate(totales, 1):
                            cell = ws.cell(row=r, column=ci, value=val)
                            cell.font = bold
                            cell.fill = total_fill
                            cell.border = thin
                        r += 1
                    style_range(ws, data_start - 1, r - 1, len(headers))
                    r += 1
                    continue
                colnames = sec.get('columnas', ['INDICADOR', 'TOTAL'])
                for ci, cn in enumerate(colnames, 1):
                    ws.cell(row=r, column=ci, value=cn)
                style_header(ws, r, len(colnames))
                r += 1
                data_start = r
                items = sec.get('filas', [])
                for item in items:
                    for ci, cn in enumerate(colnames, 1):
                        ws.cell(row=r, column=ci, value=item.get(cn, ''))
                    r += 1
                totes = sec.get('totales', {})
                if sec.get('total') is not None:
                    ws.cell(row=r, column=1, value='TOTAL').font = bold
                    ws.cell(row=r, column=2, value=sec['total']).font = bold
                    ws.cell(row=r, column=1).fill = total_fill
                    ws.cell(row=r, column=2).fill = total_fill
                    r += 1
                elif totes:
                    for ci, cn in enumerate(colnames, 1):
                        cell = ws.cell(row=r, column=ci)
                        if cn == colnames[0]:
                            cell.value = 'TOTALES'
                        else:
                            cell.value = totes.get(cn, 0)
                        cell.font = bold
                        cell.fill = total_fill
                        cell.border = thin
                    r += 1
                style_range(ws, data_start - 1, r - 1, len(colnames))
                r += 1

        elif pagina.get('id') == 'iras_edas':
            for sec in pagina.get('secciones', []):
                ws.cell(row=r, column=1, value=sec.get('titulo','')).font = bold
                r += 1
                for diag in sec.get('diagnosticos', []):
                    ws.cell(row=r, column=1, value=diag['diagnostico']).font = Font(bold=True, italic=True)
                    r += 1
                    if diag.get('items'):
                        ws.cell(row=r, column=2, value='Edad')
                        ws.cell(row=r, column=3, value='Total')
                        style_header(ws, r, 2)
                        r += 1
                        for item in diag['items']:
                            ws.cell(row=r, column=2, value=item.get('label',''))
                            ws.cell(row=r, column=3, value=item.get('valor',0))
                            r += 1
                    ws.cell(row=r, column=2, value=f"Total {diag['diagnostico']}:").font = bold
                    ws.cell(row=r, column=3, value=diag.get('total',0)).font = bold
                    r += 1
                r += 1

    # Remove default sheet if we created a named one
    if 'Reporte' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Reporte']

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import Response as FlaskResponse
    return FlaskResponse(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=reporte_minsa.xlsx'}
    )

# ---------- Filesystem (native dialogs) ----------
@app.route('/api/fs/select-folder', methods=['POST'])
def fs_select_folder():
    """Abre el dialogo nativo del SO en un proceso separado usando selector_carpeta.py."""
    try:
        script = str(SCRIPTS_BI / 'selector_carpeta.py')
        if getattr(sys, 'frozen', False):
            cmd = [sys.executable, '--run-script', script]
        else:
            cmd = [sys.executable, script]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        carpeta = result.stdout.strip()
        if carpeta:
            return jsonify({'path': carpeta})
        return jsonify({'path': ''})
    except subprocess.TimeoutExpired:
        return jsonify({'path': '', 'error': 'timeout'}), 408
    except Exception as e:
        return jsonify({'path': '', 'error': str(e)}), 500

# ---------- Vista Previa (standalone HTML) ----------
@app.route('/api/reportes-minsa/vista-previa', methods=['GET'])
def reportes_minsa_vista_previa():
    """Devuelve el reporte como página HTML standalone para vista previa directa."""
    try:
        anio = request.args.get('anio', '2024', type=str)
        meses_str = request.args.get('meses', '1,2,3,4,5,6,7,8,9,10,11,12')
        meses = [int(m) for m in meses_str.split(',') if m.strip().isdigit()]
        conn = _db_cursor()
        cur = conn.cursor()
        esquema = _get_esquema()
        try:
            data = _reporte_completo_cred(cur, esquema, anio, meses, {})
        finally:
            cur.close()
            conn.close()
        if not data or 'paginas' not in data:
            return '<h1>Error: No se pudo generar el reporte</h1>', 500, {'Content-Type': 'text/html; charset=utf-8'}
        pages_html = ''
        for i, p in enumerate(data['paginas']):
            if p.get('tabla_html'):
                pages_html += f'<h2 style="text-align:center;margin:10px 0;">{p.get("titulo", f"Página {i+1}")}</h2>'
                pages_html += p['tabla_html']
                if i < len(data['paginas']) - 1:
                    pages_html += '<hr style="margin:20px 0;border:1px dashed #ccc;">'
        full = f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Vista Previa - CRED {anio}</title>
<style>
  body {{ font-family: Arial, sans-serif; margin: 0; padding: 10px; background: #fff; color: #000; }}
</style>
</head>
<body>
{pages_html}
</body>
</html>'''
        return full, 200, {'Content-Type': 'text/html; charset=utf-8'}
    except Exception as e:
        import traceback
        return f'<h1>Error</h1><pre>{traceback.format_exc()}</pre>', 500, {'Content-Type': 'text/html; charset=utf-8'}

# ---------- Main ----------
if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(host='0.0.0.0', port=5000, debug=debug_mode, threaded=True)
